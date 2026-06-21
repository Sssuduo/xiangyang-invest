import io
from datetime import date, datetime
from flask import request, jsonify, send_file
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from models import ExportFieldConfigActivity, InvestmentActivity, InvestmentProject
from extensions import db
from routes import admin_activity_export_bp


# ============================================================
# 导出字段配置 CRUD
# ============================================================

@admin_activity_export_bp.route('/activity-export/fields', methods=['GET'])
@login_required
def get_export_fields():
    """获取所有导出字段配置"""
    fields = ExportFieldConfigActivity.query.order_by(ExportFieldConfigActivity.sort_order).all()
    return jsonify({'code': 0, 'data': [f.to_dict() for f in fields]})


@admin_activity_export_bp.route('/activity-export/fields', methods=['PUT'])
@login_required
def update_export_fields():
    """批量更新导出字段配置"""
    data = request.get_json()
    if not data or not isinstance(data, list):
        return jsonify({'code': 1, 'message': '请求数据格式错误'}), 400

    for item in data:
        field = ExportFieldConfigActivity.query.get(item.get('id'))
        if field:
            if 'field_label' in item:
                field.field_label = item['field_label']
            if 'is_visible' in item:
                field.is_visible = bool(item['is_visible'])
            if 'column_width' in item:
                field.column_width = int(item['column_width'])
            if 'sort_order' in item:
                field.sort_order = int(item['sort_order'])

    db.session.commit()
    fields = ExportFieldConfigActivity.query.order_by(ExportFieldConfigActivity.sort_order).all()
    return jsonify({'code': 0, 'data': [f.to_dict() for f in fields], 'message': '配置已保存'})


# ============================================================
# 导出预览 & 下载
# ============================================================

def _resolve_activity_row(a):
    """将动态对象解析为展示用 dict"""
    files = []
    if a.files:
        try:
            import json
            files = json.loads(a.files)
        except (json.JSONDecodeError, TypeError):
            files = []

    return {
        'project_name': a.project.project_name if a.project else '',
        'date': a.date.strftime('%Y-%m-%d %H:%M') if a.date else '',
        'content': a.content or '',
        'files': ', '.join(files) if files else '',
    }


@admin_activity_export_bp.route('/activity-export/preview', methods=['POST'])
@login_required
def export_preview():
    """导出预览：返回表头+前3条数据"""
    data = request.get_json() or {}
    activity_ids = data.get('activity_ids', [])

    fields = ExportFieldConfigActivity.query.filter_by(is_visible=True)\
        .order_by(ExportFieldConfigActivity.sort_order).all()

    headers = [{'field_key': f.field_key, 'field_label': f.field_label, 'column_width': f.column_width}
               for f in fields]

    q = InvestmentActivity.query.join(InvestmentProject)
    if activity_ids:
        q = q.filter(InvestmentActivity.id.in_(activity_ids))
    activities = q.order_by(InvestmentActivity.date.desc()).limit(3).all()

    rows = [_resolve_activity_row(a) for a in activities]

    return jsonify({'code': 0, 'data': {'headers': headers, 'rows': rows, 'total': q.count()}})


@admin_activity_export_bp.route('/activity-export/download', methods=['GET'])
@login_required
def export_download():
    """生成并下载 Excel 文件"""
    ids_str = request.args.get('activity_ids', '')
    activity_ids = [int(x) for x in ids_str.split(',') if x.strip().isdigit()] if ids_str else []

    fields = ExportFieldConfigActivity.query.filter_by(is_visible=True)\
        .order_by(ExportFieldConfigActivity.sort_order).all()
    if not fields:
        return jsonify({'code': 1, 'message': '未配置导出字段'}), 400

    q = InvestmentActivity.query.join(InvestmentProject)
    if activity_ids:
        q = q.filter(InvestmentActivity.id.in_(activity_ids))
    activities = q.order_by(InvestmentActivity.date.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = '招商动态库'

    # 样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1A3A5C', end_color='1A3A5C', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='微软雅黑', size=10)
    cell_align = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )

    # 写入表头
    for col_idx, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_idx, value=field.field_label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = field.column_width / 7

    # 写入数据行
    for row_idx, a in enumerate(activities, 2):
        row_data = _resolve_activity_row(a)
        for col_idx, field in enumerate(fields, 1):
            val = row_data.get(field.field_key, '')
            if isinstance(val, (date, datetime)):
                val = val.isoformat()
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

    # 冻结首行
    ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='招商动态库.xlsx'
    )
