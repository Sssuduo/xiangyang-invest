import io
import json
import re
from datetime import datetime
from flask import request, jsonify, send_file
from flask_login import login_required
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from models import ImportFieldConfigActivity, InvestmentActivity, InvestmentProject, FollowStatusDict
from extensions import db
from routes import admin_activity_import_bp


# ============================================================
# 文本拆分：根据日期模式将动态内容拆分为独立行
# ============================================================

# 日期正则：匹配 X月X日、X月X号、XXXX年X月X日、XXXX年X月X号
_DATE_PATTERN = re.compile(
    r'(\d{4})\s*[年/-]\s*(\d{1,2})\s*[月/-]\s*(\d{1,2})\s*[日号]?'  # 2024年6月4日
    r'|'
    r'(?<!\d)(\d{1,2})\s*[月/-]\s*(\d{1,2})\s*[日号](?!\d)'         # 6月4日
)

# 编号前缀正则：1、 2. （1） ① 等
_NUMBER_PREFIX = re.compile(
    r'^\s*(?:\d+[\.\、\)）]\s*|（[一二三四五六七八九十\d]+）\s*|[①②③④⑤⑥⑦⑧⑨⑩])\s*'
)


def _extract_date_from_text(text):
    """从文本中提取第一个日期，返回 (日期字符串, 匹配对象) 或 (None, None)"""
    m = _DATE_PATTERN.search(text)
    if not m:
        return None, None
    if m.group(1):  # 完整日期：2024年6月4日
        date_str = f'{m.group(1)}年{m.group(2)}月{m.group(3)}日'
    else:  # 简短日期：6月4日
        date_str = f'{m.group(4)}月{m.group(5)}日'
    return date_str, m


def split_activity_content(text, default_year=None):
    """
    将动态内容按日期模式拆分为多条独立动态。

    算法：
    1. 按换行和编号标记将文本拆分为句段，去除编号前缀
    2. 含日期的句段 → 新分组起点
    3. 不含日期的句段 → 归入当前最近的分组
    4. 每个分组输出一行 {year, month_day, content}
       - content 保留完整原始语句（含日期本身，不做加工修饰）
       - year 优先从文本日期中提取，否则使用 default_year
       - month_day 从文本日期中提取

    示例输入：
      "1、5月27日下午，已在农高区项目进展推进会上向赵区长汇报。
       2、5月28日日上午，区政协主席刘畅到襄州经开区调研。"
    示例输出：
      [{year:2026, month_day:"5月27日", content:"5月27日下午，已在农高区..."},
       {year:2026, month_day:"5月28日", content:"5月28日日上午，区政协主席..."}]
    """
    if default_year is None:
        default_year = datetime.now().year

    if not text or not str(text).strip():
        return [{'year': default_year, 'month_day': '', 'content': ''}]

    text = str(text).strip()

    # 第一步：按换行拆分
    raw_lines = re.split(r'\n+', text)

    # 第二步：进一步按编号标记拆分并去除编号前缀
    segments = []
    for line in raw_lines:
        line = line.strip()
        if not line:
            continue
        # 先按编号拆分
        parts = re.split(
            r'(?:^|(?<=\s))(?:\d+[\.\、\)）]\s*|（[一二三四五六七八九十\d]+）\s*|[①②③④⑤⑥⑦⑧⑨⑩])',
            line
        )
        for part in parts:
            part = part.strip()
            # 额外清理：去除残留的编号前缀（处理拆分后仍残留的情况）
            part = _NUMBER_PREFIX.sub('', part).strip()
            if part:
                segments.append(part)

    if not segments:
        return [{'year': default_year, 'month_day': '', 'content': text}]

    # 第三步：按日期分组——保留完整语句，不做修饰
    groups = []
    current_year = default_year
    current_month_day = ''
    current_parts = []

    for seg in segments:
        date_str, match = _extract_date_from_text(seg)

        if date_str:
            # 保存上一个分组
            if current_parts:
                groups.append({
                    'year': current_year,
                    'month_day': current_month_day,
                    'content': '；'.join(current_parts)
                })

            # 开始新分组——解析年月日
            if match.group(1):  # 完整日期：2024年6月4日
                current_year = int(match.group(1))
                current_month_day = f'{int(match.group(2))}月{int(match.group(3))}日'
            else:  # 简短日期：6月4日
                current_year = default_year
                current_month_day = f'{int(match.group(4))}月{int(match.group(5))}日'

            # ★ content 保留完整原始语句（含日期），不做任何加工修饰
            current_parts = [seg]
        else:
            current_parts.append(seg)

    # 保存最后一个分组
    if current_parts:
        groups.append({
            'year': current_year,
            'month_day': current_month_day,
            'content': '；'.join(current_parts)
        })

    # 如果没有识别到任何日期，整段作为一个无日期行
    if not groups:
        return [{'year': default_year, 'month_day': '', 'content': text}]

    return groups


# ============================================================
# 导入字段配置 CRUD
# ============================================================

@admin_activity_import_bp.route('/activity-import/fields', methods=['GET'])
@login_required
def get_import_fields():
    fields = ImportFieldConfigActivity.query.order_by(ImportFieldConfigActivity.sort_order).all()
    return jsonify({'code': 0, 'data': [f.to_dict() for f in fields]})


@admin_activity_import_bp.route('/activity-import/fields', methods=['PUT'])
@login_required
def update_import_fields():
    data = request.get_json()
    if not data or not isinstance(data, list):
        return jsonify({'code': 1, 'message': '格式错误'}), 400
    for item in data:
        f = ImportFieldConfigActivity.query.get(item.get('id'))
        if f:
            if 'field_label' in item:
                f.field_label = item['field_label']
            if 'is_enabled' in item:
                f.is_enabled = bool(item['is_enabled'])
            if 'is_required' in item:
                f.is_required = bool(item['is_required'])
            if 'sort_order' in item:
                f.sort_order = int(item['sort_order'])
    db.session.commit()
    fields = ImportFieldConfigActivity.query.order_by(ImportFieldConfigActivity.sort_order).all()
    return jsonify({'code': 0, 'data': [f.to_dict() for f in fields], 'message': '已保存'})


# ============================================================
# 模板项目列表（供模板下载弹窗使用）
# ============================================================

@admin_activity_import_bp.route('/activity-import/projects-for-template', methods=['GET'])
@login_required
def projects_for_template():
    """返回可用于模板的项目列表，支持跟进状态筛选"""
    follow_status = request.args.get('follow_status', '').strip()

    q = InvestmentProject.query.filter_by(is_deleted=False)
    if follow_status:
        q = q.filter_by(follow_status_code=follow_status)

    projects = q.order_by(InvestmentProject.order_no.asc()).all()

    # 获取字典名称
    follow_map = {d.code: d.name for d in FollowStatusDict.query.all()}

    result = []
    for p in projects:
        result.append({
            'id': p.id,
            'project_name': p.project_name,
            'follow_status_code': p.follow_status_code,
            'follow_status_name': follow_map.get(p.follow_status_code, ''),
            'invest_enterprise': p.invest_enterprise,
        })

    return jsonify({'code': 0, 'data': result})


# ============================================================
# 下载导入模板
# ============================================================

@admin_activity_import_bp.route('/activity-import/template', methods=['GET'])
@login_required
def download_template():
    """
    下载导入模板。
    可选参数 project_ids=1,2,3 用于预填选中的项目名称到第一列，
    同时为项目列添加下拉数据验证（包含所有活跃项目）。
    """
    fields = ImportFieldConfigActivity.query.filter_by(is_enabled=True)\
        .order_by(ImportFieldConfigActivity.sort_order).all()
    if not fields:
        return jsonify({'code': 1, 'message': '未配置导入模板'}), 400

    project_ids_str = request.args.get('project_ids', '')
    project_ids = []
    if project_ids_str:
        try:
            project_ids = [int(x) for x in project_ids_str.split(',') if x.strip().isdigit()]
        except (ValueError, TypeError):
            project_ids = []

    # 公共样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='3A7ABD', end_color='3A7ABD', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )
    data_font = Font(name='微软雅黑', size=10)
    data_align = Alignment(vertical='center', wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = '动态导入模板'

    # 第1行：表头
    for col_idx, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_idx, value=field.field_label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 列宽
    for col_idx, field in enumerate(fields, 1):
        if field.field_key == 'project_id':
            ws.column_dimensions[get_column_letter(col_idx)].width = 10
        elif field.field_key == 'project_name':
            ws.column_dimensions[get_column_letter(col_idx)].width = 26
        elif field.field_key == 'content':
            ws.column_dimensions[get_column_letter(col_idx)].width = 50
        elif field.field_key == 'files':
            ws.column_dimensions[get_column_letter(col_idx)].width = 22
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # 找到 project_id 和 project_name 的列索引（1-based）
    id_col = None
    name_col = None
    for col_idx, field in enumerate(fields, 1):
        if field.field_key == 'project_id':
            id_col = col_idx
        elif field.field_key == 'project_name':
            name_col = col_idx

    if project_ids:
        # ========== 有项目列表：预填项目ID + 项目名称 ==========
        selected_projects = InvestmentProject.query.filter(
            InvestmentProject.id.in_(project_ids),
            InvestmentProject.is_deleted == False
        ).order_by(InvestmentProject.order_no.asc()).all()

        # 获取所有活跃项目（供下拉验证用）
        all_projects = InvestmentProject.query.filter_by(is_deleted=False)\
            .order_by(InvestmentProject.project_name).all()
        all_project_names = [p.project_name for p in all_projects]

        # 第2行：必填/选填标记
        for col_idx, field in enumerate(fields, 1):
            mark = '（必填）' if field.is_required else '（选填）'
            cell = ws.cell(row=2, column=col_idx, value=mark)
            cell.font = Font(name='微软雅黑', size=9, color='C00000' if field.is_required else '909090')

        # 第3行起：写入选中的项目ID和项目名称
        for i, proj in enumerate(selected_projects):
            row = 3 + i
            # 项目ID 列
            if id_col:
                cell = ws.cell(row=row, column=id_col, value=proj.id)
                cell.font = data_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            # 项目名称 列
            if name_col:
                cell = ws.cell(row=row, column=name_col, value=proj.project_name)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border
            # 其余列空
            for col_idx in range(1, len(fields) + 1):
                if col_idx not in (id_col, name_col):
                    cell = ws.cell(row=row, column=col_idx, value='')
                    cell.font = data_font
                    cell.alignment = data_align
                    cell.border = thin_border

        # 下拉数据验证（项目名称列）
        if all_project_names and name_col:
            ws_hidden = wb.create_sheet('项目列表')
            ws_hidden.sheet_state = 'hidden'
            for i, name in enumerate(all_project_names, 1):
                ws_hidden.cell(row=i, column=1, value=name)

            dv = DataValidation(
                type='list',
                formula1=f'=项目列表!$A$1:$A${len(all_project_names)}',
                allow_blank=True
            )
            dv.error = '请从下拉列表中选择项目名称'
            dv.errorTitle = '无效项目'
            dv.prompt = '请选择'
            dv.promptTitle = '所属项目'
            last_row = max(3 + len(selected_projects), 500)
            name_letter = get_column_letter(name_col)
            dv.add(f'{name_letter}3:{name_letter}{last_row}')
            ws.add_data_validation(dv)

        ws.freeze_panes = 'A3'

    else:
        # ========== 无项目列表：空模板 ==========
        sample = {
            'project_id': '',
            'project_name': '示例项目名称',
            'date': '2026-01-01',
            'content': '示例动态内容',
            'files': 'https://example.com/doc.pdf',
        }
        for col_idx, field in enumerate(fields, 1):
            val = sample.get(field.field_key, '')
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.font = Font(name='微软雅黑', size=10, color='909090')
            cell.border = thin_border

        # 第3行：必填标记
        for col_idx, field in enumerate(fields, 1):
            mark = '（必填）' if field.is_required else '（选填）'
            cell = ws.cell(row=3, column=col_idx, value=mark)
            cell.font = Font(name='微软雅黑', size=9, color='C00000' if field.is_required else '909090')

        ws.freeze_panes = 'A4'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='招商动态导入模板.xlsx'
    )


# ============================================================
# 导入预览（含文本拆分）
# ============================================================

def _get_import_fields():
    return ImportFieldConfigActivity.query.filter_by(is_enabled=True)\
        .order_by(ImportFieldConfigActivity.sort_order).all()


def _resolve_project_id(project_name):
    """根据项目名称查找项目ID"""
    if not project_name:
        return None
    p = InvestmentProject.query.filter_by(project_name=project_name.strip(), is_deleted=False).first()
    return p.id if p else None


def _parse_activity_datetime(val):
    """将字符串解析为 datetime"""
    if not val:
        return None
    s = str(val).strip()
    # 尝试多种格式
    for fmt in [
        '%Y-%m-%d %H:%M',
        '%Y-%m-%d',
        '%Y/%m/%d',
        '%Y年%m月%d日',
        '%m月%d日',
        '%Y-%m-%dT%H:%M',
        '%Y-%m-%dT%H:%M:%S',
    ]:
        try:
            return datetime.strptime(s, fmt)
        except (ValueError, TypeError):
            continue
    # 尝试 ISO 格式
    try:
        return datetime.fromisoformat(s)
    except (ValueError, TypeError):
        return None


def _validate_activity_row(row_data, field_configs, existing_project_names, row_idx):
    """校验单行数据，返回错误列表（date 改为选填）"""
    errors = []
    for f in field_configs:
        if f.field_key == 'date':
            continue  # 日期选填，不校验必填
        val = row_data.get(f.field_key, '')
        if f.is_required and (val is None or str(val).strip() == ''):
            errors.append(f'第{row_idx}行：{f.field_label} 为必填项')

    # 校验所属项目是否存在
    pname = str(row_data.get('project_name', '') or '').strip()
    if pname and pname not in existing_project_names:
        errors.append(f'第{row_idx}行：项目「{pname}」在数据库中不存在')

    return errors


# 导入预览的列头（含项目ID和年份月日拆分列）
_PREVIEW_HEADERS = ['项目ID', '所属项目', '年份', '月日', '动态内容', '附件(URL)']


@admin_activity_import_bp.route('/activity-import/preview', methods=['POST'])
@login_required
def import_preview():
    """上传 Excel 并返回预览数据（含文本智能拆分 + 年份/月日分离）"""
    if 'file' not in request.files:
        return jsonify({'code': 1, 'message': '未找到上传文件'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'code': 1, 'message': '未选择文件'}), 400

    try:
        wb = load_workbook(file, read_only=True)
        ws = wb.active
    except Exception:
        return jsonify({'code': 1, 'message': '无法读取文件，请确认是有效的 Excel 文件'}), 400

    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value is not None:
            headers.append(str(cell.value).strip())

    # 获取配置的导入字段
    import_fields = _get_import_fields()
    expected_headers = [f.field_label for f in import_fields]

    # 校验表头是否一致
    if headers != expected_headers:
        return jsonify({
            'code': 1,
            'message': '模板有误，请选择正确的导入模板',
            'data': {'expected': expected_headers, 'actual': headers}
        }), 400

    # 构建 field_label -> field_key 映射
    field_map = {f.field_label: f.field_key for f in import_fields}

    # 获取已存在的项目名称
    existing_project_names = set(
        p.project_name for p in InvestmentProject.query.filter_by(is_deleted=False).all()
    )

    # 默认年份 = 当前年份
    default_year = datetime.now().year

    # 读取数据行
    raw_rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or str(v).strip() == '' for v in row):
            continue

        row_data = {}
        for col_idx, val in enumerate(row):
            if col_idx < len(headers):
                field_label = headers[col_idx]
                field_key = field_map.get(field_label, '')
                str_val = str(val).strip() if val is not None else ''
                row_data[field_key] = str_val

        # 跳过标记行
        first_val = list(row_data.values())[0] if row_data else ''
        if first_val in ('（必填）', '（选填）'):
            continue

        raw_rows.append({
            'row': row_idx,
            'data': row_data
        })

    # ========== 文本智能拆分（保留完整语句 + 年/月日分离） ==========
    split_rows = []
    all_errors = []
    split_counter = 0

    for raw in raw_rows:
        content = raw['data'].get('content', '')
        # 从原始日期列尝试提取年份
        excel_date = raw['data'].get('date', '')
        excel_year = default_year
        if excel_date:
            dm = _extract_date_from_text(excel_date)
            if dm[0]:
                m = _DATE_PATTERN.search(excel_date)
                if m and m.group(1):
                    excel_year = int(m.group(1))

        split_results = split_activity_content(content, default_year=excel_year)

        # 获取原始行中的项目ID和项目名称
        raw_project_id = str(raw['data'].get('project_id', '') or '').strip()
        raw_project_name = raw['data'].get('project_name', '')

        # 优先通过项目ID解析，确保对应准确
        resolved_project_name = raw_project_name
        if raw_project_id and raw_project_id.isdigit():
            proj = InvestmentProject.query.get(int(raw_project_id))
            if proj and not proj.is_deleted:
                resolved_project_name = proj.project_name

        for sr in split_results:
            split_counter += 1
            new_data = {
                'project_id': raw_project_id,
                'project_name': resolved_project_name,
                'year': sr['year'],
                'month_day': sr['month_day'],
                'content': sr['content'],
                'files': raw['data'].get('files', ''),
            }

            row_errors = _validate_activity_row(
                new_data, import_fields, existing_project_names, split_counter
            )
            all_errors.extend(row_errors)

            split_rows.append({
                'row': split_counter,
                '_source_row': raw['row'],
                'data': new_data,
                'errors': row_errors,
                '_valid': len(row_errors) == 0,
                '_split': len(split_results) > 1
            })

    # 统计
    valid_count = sum(1 for r in split_rows if r['_valid'])
    error_count = len(split_rows) - valid_count

    return jsonify({
        'code': 0,
        'data': {
            'headers': _PREVIEW_HEADERS,
            'rows': split_rows,
            'total': len(split_rows),
            'valid_count': valid_count,
            'error_count': error_count,
            'all_errors': all_errors,
            'default_year': default_year,
            'split_info': f'原始 {len(raw_rows)} 行，拆分后 {len(split_rows)} 行'
        }
    })


# ============================================================
# 执行导入
# ============================================================

@admin_activity_import_bp.route('/activity-import/execute', methods=['POST'])
@login_required
def import_execute():
    """将预览通过的数据写入数据库"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 1, 'message': '数据为空'}), 400

    rows = data.get('rows', [])
    if not rows:
        return jsonify({'code': 1, 'message': '无数据可导入'}), 400

    valid_rows = [r for r in rows if r.get('_valid')]
    if not valid_rows:
        return jsonify({'code': 1, 'message': '没有有效数据行'}), 400

    # 预加载项目名称->ID映射
    project_map = {
        p.project_name: p.id
        for p in InvestmentProject.query.filter_by(is_deleted=False).all()
    }

    imported = 0
    for row in valid_rows:
        row_data = row['data']

        # 优先通过项目ID直接获取（100%准确对应）
        pid_str = str(row_data.get('project_id', '') or '').strip()
        if pid_str and pid_str.isdigit():
            project_id = int(pid_str)
            # 验证项目是否存在
            proj = InvestmentProject.query.get(project_id)
            if not proj or proj.is_deleted:
                project_id = None
        else:
            # 回退：通过项目名称查找
            project_id = project_map.get(row_data.get('project_name', '').strip())

        if not project_id:
            continue

        # 组合日期：年份 + 月日 → "2026年5月27日"
        year = row_data.get('year', '')
        month_day = row_data.get('month_day', '')
        if year and month_day:
            date_str = f'{year}年{month_day}'
        elif month_day:
            date_str = month_day
        elif row_data.get('date', ''):
            date_str = row_data.get('date', '')
        else:
            date_str = ''

        activity = InvestmentActivity(
            project_id=project_id,
            date=_parse_activity_datetime(date_str) if date_str else None,
            content=row_data.get('content', ''),
            files=json.dumps(
                [u.strip() for u in str(row_data.get('files', '')).split(',') if u.strip()],
                ensure_ascii=False
            ) if row_data.get('files', '').strip() else '[]'
        )
        db.session.add(activity)
        imported += 1

    db.session.commit()
    return jsonify({'code': 0, 'message': f'成功导入 {imported} 条记录', 'data': {'count': imported}})
