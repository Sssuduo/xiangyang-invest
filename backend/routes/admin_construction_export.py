import io
import json
import re as _re
from datetime import date, datetime, timedelta
from flask import request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from models import PrintTemplate, PrintFieldConfig, PrintTemplate, PrintFieldConfig
from models import ConstructionProject, WorkProgress, ProjectIssue, WorkRoadmapItem
from models import ConstructionProjectTypeDict, DispatchStatusDict, OrganizationDict, IssueTypeDict, ResolutionStatusDict
from extensions import db
from routes import admin_construction_export_bp
from routes.business_auth import dual_login_required, visitor_block


def _sanitize_filename(name):
    """移除文件名中的非法字符（仅去除 Windows 真正禁止的字符）"""
    name = _re.sub(r'[\[\]*?"<>|\\/:]', '', name)
    name = _re.sub(r'_+', '_', name)
    return name.strip() or f'导出文件_{date.today().strftime("%Y%m%d")}'


# ============================================================
# 模板 CRUD
# ============================================================

@admin_construction_export_bp.route('/construction-export/templates', methods=['GET'])
@dual_login_required
def get_templates():
    """列出所有导出模板"""
    entity_type = request.args.get('entity_type', 'construction').strip()
    templates = PrintTemplate.query.filter_by(entity_type=entity_type)\
        .order_by(PrintTemplate.id.asc()).all()
    return jsonify({'code': 0, 'data': [t.to_dict() for t in templates]})


@admin_construction_export_bp.route('/construction-export/templates', methods=['POST'])
@dual_login_required
@visitor_block
def create_template():
    """新建导出模板"""
    data = request.get_json()
    if not data or not data.get('name'):
        return jsonify({'code': 1, 'message': '模板名称不能为空'}), 400
    entity_type = data.get('entity_type', 'construction')
    template = PrintTemplate(name=data['name'].strip(), entity_type=entity_type)
    db.session.add(template)
    db.session.flush()

    # 从已有模板（排除自身）复制字段配置
    default_template = PrintTemplate.query.filter_by(entity_type=entity_type)\
        .filter(PrintTemplate.id != template.id)\
        .order_by(PrintTemplate.id.asc()).first()
    if default_template:
        default_fields = PrintFieldConfig.query.filter_by(template_id=default_template.id)\
            .order_by(PrintFieldConfig.sort_order).all()
        for f in default_fields:
            db.session.add(PrintFieldConfig(
                template_id=template.id,
                field_key=f.field_key,
                field_label=f.field_label,
                is_visible=True,
                column_width=f.column_width,
                sort_order=f.sort_order
            ))

    db.session.commit()
    return jsonify({'code': 0, 'data': template.to_dict(), 'message': '模板已创建'})


@admin_construction_export_bp.route('/construction-export/templates/<int:template_id>', methods=['PUT'])
@dual_login_required
@visitor_block
def update_template(template_id):
    """重命名模板"""
    tpl = PrintTemplate.query.get(template_id)
    if not tpl:
        return jsonify({'code': 1, 'message': '模板不存在'}), 404
    data = request.get_json()
    if not data or not data.get('name'):
        return jsonify({'code': 1, 'message': '模板名称不能为空'}), 400
    tpl.name = data['name'].strip()
    db.session.commit()
    return jsonify({'code': 0, 'data': tpl.to_dict(), 'message': '模板已更新'})


@admin_construction_export_bp.route('/construction-export/templates/<int:template_id>', methods=['DELETE'])
@dual_login_required
@visitor_block
def delete_template(template_id):
    """删除模板及其字段配置"""
    tpl = PrintTemplate.query.get(template_id)
    if not tpl:
        return jsonify({'code': 1, 'message': '模板不存在'}), 404
    # 至少保留一个模板
    remaining = PrintTemplate.query.filter_by(entity_type=tpl.entity_type).count()
    if remaining <= 1:
        return jsonify({'code': 1, 'message': '至少保留一个导出模板'}), 400
    PrintFieldConfig.query.filter_by(template_id=template_id).delete()
    db.session.delete(tpl)
    db.session.commit()
    return jsonify({'code': 0, 'message': '模板已删除'})


# ============================================================
# 导出字段配置 CRUD
# ============================================================

def _get_default_template_id():
    """获取默认模板ID（entity_type='construction' 的第一个模板）"""
    tpl = PrintTemplate.query.filter_by(entity_type='construction').order_by(PrintTemplate.id).first()
    return tpl.id if tpl else 1


@admin_construction_export_bp.route('/construction-export/fields', methods=['GET'])
@dual_login_required
def get_export_fields():
    """获取指定模板的导出字段配置"""
    template_id = request.args.get('template_id', 0, type=int) or _get_default_template_id()
    fields = PrintFieldConfig.query.filter_by(template_id=template_id)\
        .order_by(PrintFieldConfig.sort_order).all()
    return jsonify({'code': 0, 'data': [f.to_dict() for f in fields]})


@admin_construction_export_bp.route('/construction-export/fields', methods=['PUT'])
@dual_login_required
@visitor_block
def update_export_fields():
    """批量更新导出字段配置"""
    data = request.get_json()
    if not data or not isinstance(data, list):
        return jsonify({'code': 1, 'message': '请求数据格式错误'}), 400

    template_id = None
    saved_ids = set()
    for item in data:
        field_id = item.get('id')
        if field_id and field_id > 0:
            saved_ids.add(field_id)
            field = PrintFieldConfig.query.get(field_id)
            if field:
                template_id = field.template_id
        elif item.get('_new') and item.get('template_id'):
            # 新建字段（自定义列）
            template_id = int(item['template_id'])
            field = PrintFieldConfig(
                template_id=template_id,
                field_key=item.get('field_key', ''),
                field_label=item.get('field_label', ''),
                is_visible=item.get('is_visible', True),
                is_custom=item.get('is_custom', True),
                column_width=item.get('column_width', 120),
                sort_order=item.get('sort_order', 0)
            )
            db.session.add(field)
            continue

        if field:
            if 'field_label' in item:
                field.field_label = item['field_label']
            if 'is_visible' in item:
                field.is_visible = bool(item['is_visible'])
            if 'column_width' in item:
                field.column_width = int(item['column_width'])
            if 'sort_order' in item:
                field.sort_order = int(item['sort_order'])

    # 删除在前端被移除的自定义字段
    if template_id:
        existing_ids = {f.id for f in PrintFieldConfig.query.filter_by(template_id=template_id).all()}
        to_delete = existing_ids - saved_ids
        if to_delete:
            PrintFieldConfig.query.filter(PrintFieldConfig.id.in_(to_delete)).delete(synchronize_session=False)

    db.session.commit()
    fields = PrintFieldConfig.query.filter_by(template_id=template_id)\
        .order_by(PrintFieldConfig.sort_order).all() if template_id else []
    return jsonify({'code': 0, 'data': [f.to_dict() for f in fields], 'message': '配置已保存'})


# ============================================================
# 导出预览 & 下载 — 辅助函数
# ============================================================

def _get_last2_windows(project_ids=None):
    """查询最近2条工作进展的时间窗口（限定在勾选项目范围内），返回 (本周窗口, 上周窗口) 元组
    每个窗口为 (start_date, end_date)，无进展时对应窗口为 None"""
    q = WorkProgress.query
    if project_ids:
        q = q.filter(WorkProgress.project_id.in_(project_ids))
    recent = q.order_by(WorkProgress.start_date.desc()).limit(2).all()
    this_week = (recent[0].start_date, recent[0].end_date) if len(recent) >= 1 else None
    last_week = (recent[1].start_date, recent[1].end_date) if len(recent) >= 2 else None
    return this_week, last_week


def _aggregate_work_progresses(project, progress_range='', last2_windows=None):
    """聚合项目工作进展，按开始日期倒序，序号分段"""
    if progress_range == 'last2':
        return _aggregate_work_progresses_last2(project, last2_windows)

    q = WorkProgress.query.filter_by(project_id=project.id)
    if progress_range == 'last1m':
        since = datetime.utcnow() - timedelta(days=30)
        q = q.filter(WorkProgress.start_date >= since)
    elif progress_range == 'last3m':
        since = datetime.utcnow() - timedelta(days=90)
        q = q.filter(WorkProgress.start_date >= since)

    q = q.order_by(WorkProgress.start_date.desc())
    if progress_range == 'last5':
        progresses = q.limit(5).all()
    else:
        progresses = q.all()

    if not progresses:
        return ''

    lines = []
    for i, p in enumerate(progresses, 1):
        start = p.start_date.isoformat() if p.start_date else ''
        end = p.end_date.isoformat() if p.end_date else ''
        content = p.content or ''
        lines.append(f'{i}、{start}~{end}: {content}')
    return '\n'.join(lines)


def _aggregate_work_progresses_last2(project, last2_windows):
    """最近2条模式：按全局时间窗口匹配进展，输出 上周/本周 格式"""
    this_week_window, last_week_window = last2_windows or (None, None)

    this_week_progress = None
    last_week_progress = None

    if this_week_window:
        this_week_progress = WorkProgress.query.filter_by(
            project_id=project.id,
            start_date=this_week_window[0],
            end_date=this_week_window[1]
        ).first()
    if last_week_window:
        last_week_progress = WorkProgress.query.filter_by(
            project_id=project.id,
            start_date=last_week_window[0],
            end_date=last_week_window[1]
        ).first()

    lines = []
    # 第一点：上周
    lines.append(f'上周：{last_week_progress.content if last_week_progress else "无进展。"}')
    # 第二点：本周
    lines.append(f'本周：{this_week_progress.content if this_week_progress else "无进展。"}')
    return '\n'.join(lines)


def _aggregate_issues(project):
    """聚合调度问题，按创建时间倒序，只输出问题描述内容"""
    issues = ProjectIssue.query.filter_by(project_id=project.id)\
        .order_by(ProjectIssue.created_at.desc()).all()
    if not issues:
        return ''

    lines = []
    for i, iss in enumerate(issues, 1):
        desc = iss.issue_description or ''
        if desc:
            lines.append(f'{i}、{desc}')
    return '\n'.join(lines)


def _aggregate_work_roadmap(project, status_filter='pending'):
    """聚合工作路径图，按 sort_order 升序，序号分段
    status_filter: 'pending' 仅待完成 / '' 全部 / 'pending,completed' 多状态逗号分隔"""
    q = WorkRoadmapItem.query.filter_by(project_id=project.id)
    if status_filter:
        codes = [s.strip() for s in status_filter.split(',') if s.strip()]
        if len(codes) == 1:
            q = q.filter_by(status=codes[0])
        elif len(codes) > 1:
            q = q.filter(WorkRoadmapItem.status.in_(codes))
    items = q.order_by(WorkRoadmapItem.sort_order.asc()).all()
    if not items:
        return ''

    # 简化输出：仅保留序号 + 内容，去掉状态标签和日期
    lines = []
    for i, item in enumerate(items, 1):
        lines.append(f'{i}、{item.content}')
    return '\n'.join(lines)


def _resolve_team_leader_names(p):
    """解析专班负责人 ID 列表为中文名列表"""
    _team_names = []
    try:
        _leader_ids = json.loads(getattr(p, 'team_leader_ids', '[]') or '[]')
        if _leader_ids:
            from models import Staff
            _staff_list = Staff.query.filter(Staff.id.in_(_leader_ids)).all()
            _staff_map = {s.id: s.name for s in _staff_list}
            _team_names = [_staff_map.get(sid, str(sid)) for sid in _leader_ids]
    except Exception:
        pass
    return '、'.join(_team_names)


def _resolve_project_row(p, progress_range='', work_path_range='pending', last2_windows=None):
    """将 ConstructionProject 解析为展示用 dict（含字典名称 + 聚合字段）"""
    type_map = {d.code: d.name for d in ConstructionProjectTypeDict.query.all()}
    dispatch_map = {d.code: d.name for d in DispatchStatusDict.query.all()}
    org_map = {d.code: d.name for d in OrganizationDict.query.all()}

    responsible_unit = org_map.get(p.responsible_unit_code, p.responsible_unit_code or '')
    responsible_person = p.responsible_person or ''
    responsible_phone = p.responsible_person_phone or ''

    return {
        'order_no': p.order_no,
        'project_name': p.project_name or '',
        'project_type_code': type_map.get(p.project_type_code, p.project_type_code),
        'dispatch_status_code': dispatch_map.get(p.dispatch_status_code, p.dispatch_status_code),
        'construction_content': p.construction_content or '',
        'construction_unit': p.construction_unit or '',
        'responsible_unit_code': responsible_unit,
        'responsible_person': responsible_person,
        'responsible_person_phone': responsible_phone,
        # V3 组合字段
        '_combined_contact': '\n'.join([responsible_unit, responsible_person, responsible_phone]),
        # 新字段
        'construction_location': p.construction_location or '',
        'start_date': p.start_date or '',
        'end_date': p.end_date or '',
        'funding_source': p.funding_source or '',
        'wuhua_platform': p.wuhua_platform or '',
        '_team_follower_names': _resolve_team_leader_names(p),
        # 聚合字段
        'work_progresses': _aggregate_work_progresses(p, progress_range, last2_windows),
        'issues': _aggregate_issues(p),
        'work_roadmap': _aggregate_work_roadmap(p, work_path_range),
    }


def _fmt_cell(val):
    """格式化单元格值"""
    if isinstance(val, date):
        return val.isoformat()
    return val


# ============================================================
# 导出预览
# ============================================================

@admin_construction_export_bp.route('/construction-export/preview', methods=['POST'])
@dual_login_required
def export_preview():
    """导出预览：返回表头+前3条数据"""
    data = request.get_json() or {}
    project_ids = data.get('project_ids', [])
    template_id = data.get('template_id', 0) or _get_default_template_id()
    progress_range = data.get('progress_range', '').strip()
    work_path_range = data.get('work_path_range', 'pending').strip()

    # 最近2条：预计算时间窗口（限定在勾选项目范围内）
    last2_windows = _get_last2_windows(project_ids) if progress_range == 'last2' else None

    fields = PrintFieldConfig.query.filter_by(template_id=template_id, is_visible=True)\
        .order_by(PrintFieldConfig.sort_order).all()

    headers = [{'field_key': f.field_key, 'field_label': f.field_label, 'column_width': f.column_width}
               for f in fields]

    q = ConstructionProject.query.filter_by(is_deleted=False) \
        .filter(ConstructionProject.dispatch_status_code != 'exited')
    if project_ids:
        q = q.filter(ConstructionProject.id.in_(project_ids))
    projects = q.order_by(ConstructionProject.order_no.asc()).limit(3).all()

    rows = [_resolve_project_row(p, progress_range, work_path_range, last2_windows) for p in projects]

    return jsonify({'code': 0, 'data': {'headers': headers, 'rows': rows, 'total': q.count()}})


# ============================================================
# 导出下载
# ============================================================

@admin_construction_export_bp.route('/construction-export/download', methods=['GET'])
@dual_login_required
def export_download():
    """生成并下载 Excel 文件"""
    ids_str = request.args.get('project_ids', '')
    project_ids = [int(x) for x in ids_str.split(',') if x.strip().isdigit()] if ids_str else []

    template_id = request.args.get('template_id', 0, type=int) or _get_default_template_id()
    progress_range = request.args.get('progress_range', '').strip()
    progress_mode = request.args.get('progress_mode', 'aggregate').strip()  # 'aggregate' | 'progress'
    work_path_range = request.args.get('work_path_range', 'pending').strip()

    # 最近2条：预计算时间窗口（限定在勾选项目范围内）
    last2_windows = _get_last2_windows(project_ids) if progress_range == 'last2' else None

    template = PrintTemplate.query.get(template_id)
    template_name = template.name if template else '在建项目库'

    fields = PrintFieldConfig.query.filter_by(template_id=template_id, is_visible=True)\
        .order_by(PrintFieldConfig.sort_order).all()
    if not fields:
        return jsonify({'code': 1, 'message': '未配置导出字段'}), 400

    q = ConstructionProject.query.filter_by(is_deleted=False) \
        .filter(ConstructionProject.dispatch_status_code != 'exited')
    if project_ids:
        q = q.filter(ConstructionProject.id.in_(project_ids))
    projects = q.order_by(ConstructionProject.order_no.asc()).all()

    # ---- 样式定义（使用打印模板的字体设置） ----
    _tf = (template.title_font_family or template.font_family or '微软雅黑') if template else '微软雅黑'
    _ts = (template.title_font_size or template.header_font_size or 14) if template else 14
    _hf = (template.table_header_font_family or template.font_family or '微软雅黑') if template else '微软雅黑'
    _hs = (template.table_header_font_size or template.font_size or 10) if template else 10
    _cf = (template.cell_font_family or template.font_family or '微软雅黑') if template else '微软雅黑'
    _cs = (template.cell_font_size or template.font_size or 10) if template else 10

    title_font = Font(name=_tf, size=_ts, bold=True, color='1A3A5C')
    title_align = Alignment(horizontal='center', vertical='center')
    header_font = Font(name=_hf, size=_hs, bold=True, color='000000')
    header_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name=_cf, size=_cs)
    cell_align = Alignment(vertical='top', wrap_text=True)
    cell_align_center = Alignment(horizontal='center', vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='333333'),
        right=Side(style='thin', color='333333'),
        top=Side(style='thin', color='333333'),
        bottom=Side(style='thin', color='333333'),
    )
    title_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    wb = Workbook()
    ws = wb.active
    ws.title = template_name[:31]

    if progress_mode == 'progress':
        # ---- 按行导出：工作进展拆分为独立行 ----
        project_fields = [f for f in fields if f.field_key not in ('work_progresses',)]
        progress_cols = [
            {'field_key': 'progress_start_date', 'field_label': '进展开始日期', 'column_width': 120},
            {'field_key': 'progress_end_date', 'field_label': '进展结束日期', 'column_width': 120},
            {'field_key': 'progress_content', 'field_label': '进展内容', 'column_width': 400},
        ]
        total_cols = len(project_fields) + len(progress_cols)

        # 第1行：标题行（合并单元格）
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(row=1, column=1, value=template_name)
        title_cell.font = title_font
        title_cell.alignment = title_align
        title_cell.fill = title_fill
        for c in range(1, total_cols + 1):
            ws.cell(row=1, column=c).border = thin_border

        # 第2行：表头
        for col_idx, f in enumerate(project_fields, 1):
            cell = ws.cell(row=2, column=col_idx, value=f.field_label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = f.column_width / 7
        for col_idx, pc in enumerate(progress_cols, len(project_fields) + 1):
            cell = ws.cell(row=2, column=col_idx, value=pc['field_label'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = pc['column_width'] / 7

        # 数据行 + 合并单元格
        current_row = 3
        for p in projects:
            project_row = _resolve_project_row(p, progress_range=progress_range, work_path_range=work_path_range, last2_windows=last2_windows)
            progresses = WorkProgress.query.filter_by(project_id=p.id)\
                .order_by(WorkProgress.start_date.desc()).all()

            if not progresses:
                # 无进展：一行，项目字段正常填，进展字段留空
                for col_idx, f in enumerate(project_fields, 1):
                    val = _fmt_cell(project_row.get(f.field_key, ''))
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.font = cell_font
                    cell.alignment = cell_align
                    cell.border = thin_border
                for col_idx in range(len(project_fields) + 1, total_cols + 1):
                    cell = ws.cell(row=current_row, column=col_idx, value='')
                    cell.font = cell_font
                    cell.alignment = cell_align_center
                    cell.border = thin_border
                current_row += 1
            else:
                start_row = current_row
                for pg in progresses:
                    # 项目字段
                    for col_idx, f in enumerate(project_fields, 1):
                        val = _fmt_cell(project_row.get(f.field_key, ''))
                        cell = ws.cell(row=current_row, column=col_idx, value=val)
                        cell.font = cell_font
                        cell.alignment = cell_align
                        cell.border = thin_border
                    # 进展字段
                    progress_data = {
                        'progress_start_date': pg.start_date.isoformat() if pg.start_date else '',
                        'progress_end_date': pg.end_date.isoformat() if pg.end_date else '',
                        'progress_content': pg.content or '',
                    }
                    for col_idx, pc in enumerate(progress_cols, len(project_fields) + 1):
                        val = progress_data.get(pc['field_key'], '')
                        cell = ws.cell(row=current_row, column=col_idx, value=val)
                        cell.font = cell_font
                        cell.alignment = cell_align_center if col_idx <= len(project_fields) + 2 else cell_align
                        cell.border = thin_border
                    current_row += 1

                # 合并项目字段单元格（纵向）
                if current_row - start_row > 1:
                    for col_idx in range(1, len(project_fields) + 1):
                        ws.merge_cells(
                            start_row=start_row, start_column=col_idx,
                            end_row=current_row - 1, end_column=col_idx
                        )

        ws.freeze_panes = 'A3'
    else:
        # ---- 聚合导出 ----
        total_cols = len(fields)

        # 第1行：标题行（合并单元格）
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(row=1, column=1, value=template_name)
        title_cell.font = title_font
        title_cell.alignment = title_align
        title_cell.fill = title_fill
        for c in range(1, total_cols + 1):
            ws.cell(row=1, column=c).border = thin_border

        # 第2行：表头
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=2, column=col_idx, value=field.field_label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = field.column_width / 7

        # 数据行
        for row_idx, p in enumerate(projects, 3):
            row_data = _resolve_project_row(p, progress_range=progress_range, work_path_range=work_path_range, last2_windows=last2_windows)
            for col_idx, field in enumerate(fields, 1):
                val = row_data.get(field.field_key, '')
                val = _fmt_cell(val)
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = cell_font
                cell.alignment = cell_align
                cell.border = thin_border

        ws.freeze_panes = 'A3'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{_sanitize_filename(template_name)}.xlsx'
    )
