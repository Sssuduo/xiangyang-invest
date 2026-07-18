import io
import json
from datetime import date, datetime
from flask import request, jsonify, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from models import (
    ImportFieldConfigConstruction,
    ConstructionProject, WorkRoadmapItem, WorkProgress, ProjectIssue,
    ConstructionProjectTypeDict, DispatchStatusDict,
    IssueTypeDict, ResolutionStatusDict, OrganizationDict
)
from extensions import db
from routes import admin_construction_import_bp
from routes.business_auth import dual_login_required, visitor_block
from routes.admin_construction import _renumber_projects


def _parse_date(val):
    """将 ISO 日期字符串转为 date 对象"""
    if not val:
        return None
    try:
        return date.fromisoformat(val)
    except (ValueError, TypeError):
        return None


# ============================================================
# 文本解析：将格式化文本拆分为子表记录
# ============================================================

def _resolve_dict_name_to_code(dict_model, name_or_code):
    """将字典名称或 code 统一解析为 code。优先精确匹配 code，其次匹配 name。"""
    if not name_or_code:
        return None
    val = name_or_code.strip()
    # 优先 code 精确匹配
    item = dict_model.query.filter_by(code=val, is_active=True).first()
    if item:
        return item.code
    # 其次 name 精确匹配
    item = dict_model.query.filter_by(name=val, is_active=True).first()
    if item:
        return item.code
    # 模糊匹配 name
    item = dict_model.query.filter(
        dict_model.name.contains(val),
        dict_model.is_active == True
    ).first()
    if item:
        return item.code
    return None


def parse_roadmap_text(text):
    """
    解析工作路径文本。
    格式：内容|计划日期|状态;内容|计划日期|状态;...
    sort_order 按出现顺序自动分配（1, 2, 3...）
    返回：[{sort_order, content, planned_date, status}, ...]
    """
    results = []
    if not text or not str(text).strip():
        return results

    rows = str(text).strip().split(';')
    for row in rows:
        row = row.strip()
        if not row:
            continue
        parts = [p.strip() for p in row.split('|')]
        if len(parts) < 1:
            continue

        content = parts[0]
        if not content:
            continue  # 内容为必填

        planned_date = parts[1] if len(parts) > 1 and parts[1] else None
        status = parts[2] if len(parts) > 2 and parts[2] else 'pending'
        if status not in ('pending', 'completed', 'cancelled'):
            status = 'pending'

        results.append({
            'sort_order': len(results) + 1,
            'content': content,
            'planned_date': planned_date,
            'status': status,
            'is_delayed': False,
            'delay_reason': '',
            'cancel_reason': ''
        })
    return results


def parse_progress_text(text):
    """
    解析工作进展文本。
    格式：开始日期|结束日期|内容;开始日期|结束日期|内容;...
    返回：[{start_date, end_date, content}, ...]
    """
    results = []
    if not text or not str(text).strip():
        return results

    rows = str(text).strip().split(';')
    for row in rows:
        row = row.strip()
        if not row:
            continue
        parts = [p.strip() for p in row.split('|')]
        if len(parts) < 3:
            continue  # 至少需要开始日期、结束日期、内容

        start_date = parts[0] if parts[0] else None
        end_date = parts[1] if parts[1] else None
        content = parts[2]
        if not content:
            continue  # 内容为必填

        results.append({
            'start_date': start_date,
            'end_date': end_date,
            'content': content
        })
    return results


def parse_issue_text(text):
    """
    解析调度问题文本。
    格式：问题类型|问题描述|解决状态|解决备注|主责部门;...
    返回：[{issue_type_code, issue_description, resolution_status_code, resolution_note, main_department_code}, ...]
    """
    results = []
    if not text or not str(text).strip():
        return results

    rows = str(text).strip().split(';')
    for row in rows:
        row = row.strip()
        if not row:
            continue
        parts = [p.strip() for p in row.split('|')]
        if len(parts) < 2:
            continue  # 至少需要问题类型和问题描述

        issue_type_raw = parts[0]
        issue_description = parts[1]
        if not issue_description:
            continue  # 问题描述为必填

        # 解析问题类型（支持名称或 code）
        issue_type_code = _resolve_dict_name_to_code(IssueTypeDict, issue_type_raw)
        if not issue_type_code and issue_type_raw:
            issue_type_code = issue_type_raw  # 如果找不到匹配，直接用原始值

        # 解析解决状态（支持名称或 code）
        resolution_raw = parts[2] if len(parts) > 2 and parts[2] else 'pending'
        resolution_status_code = _resolve_dict_name_to_code(ResolutionStatusDict, resolution_raw)
        if not resolution_status_code:
            resolution_status_code = 'pending'  # 默认待回应

        resolution_note = parts[3] if len(parts) > 3 and parts[3] else ''
        main_department_code = parts[4] if len(parts) > 4 and parts[4] else ''

        results.append({
            'issue_type_code': issue_type_code,
            'issue_description': issue_description,
            'resolution_status_code': resolution_status_code,
            'resolution_note': resolution_note,
            'main_department_code': main_department_code
        })
    return results


# ============================================================
# 导入字段配置 CRUD
# ============================================================

@admin_construction_import_bp.route('/construction-import/fields', methods=['GET'])
@dual_login_required
def get_import_fields():
    fields = ImportFieldConfigConstruction.query.order_by(ImportFieldConfigConstruction.sort_order).all()
    return jsonify({'code': 0, 'data': [f.to_dict() for f in fields]})


@admin_construction_import_bp.route('/construction-import/fields', methods=['PUT'])
@dual_login_required
@visitor_block
def update_import_fields():
    data = request.get_json()
    if not data or not isinstance(data, list):
        return jsonify({'code': 1, 'message': '格式错误'}), 400
    for item in data:
        f = ImportFieldConfigConstruction.query.get(item.get('id'))
        if f:
            if 'field_label' in item:
                f.field_label = item['field_label']
            if 'is_enabled' in item:
                f.is_enabled = bool(item['is_enabled'])
            if 'is_required' in item:
                f.is_required = bool(item['is_required'])
            if 'sort_order' in item:
                f.sort_order = int(item['sort_order'])
    db.session.commit()
    fields = ImportFieldConfigConstruction.query.order_by(ImportFieldConfigConstruction.sort_order).all()
    return jsonify({'code': 0, 'data': [f.to_dict() for f in fields], 'message': '已保存'})


# ============================================================
# 下载导入模板
# ============================================================

@admin_construction_import_bp.route('/construction-import/template', methods=['GET'])
@dual_login_required
def download_template():
    fields = ImportFieldConfigConstruction.query.filter_by(is_enabled=True)\
        .order_by(ImportFieldConfigConstruction.sort_order).all()
    if not fields:
        return jsonify({'code': 1, 'message': '未配置导入模板'}), 400

    # 预加载字典数据（用于下拉验证）
    project_types = ConstructionProjectTypeDict.query.filter_by(is_active=True).order_by(
        ConstructionProjectTypeDict.sort_order
    ).all()
    dispatch_statuses = DispatchStatusDict.query.filter_by(is_active=True).order_by(
        DispatchStatusDict.sort_order
    ).all()
    organizations = OrganizationDict.query.filter_by(is_active=True).order_by(
        OrganizationDict.sort_order
    ).all()

    # 公共样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='3A7ABD', end_color='3A7ABD', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )
    data_font = Font(name='微软雅黑', size=10)
    data_align = Alignment(vertical='center', wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = '在建项目导入模板'

    # 第1行：表头
    for col_idx, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_idx, value=field.field_label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 第2行：必填/选填标记
    for col_idx, field in enumerate(fields, 1):
        mark = '（必填）' if field.is_required else '（选填）'
        cell = ws.cell(row=2, column=col_idx, value=mark)
        cell.font = Font(name='微软雅黑', size=9, color='C00000' if field.is_required else '909090')

    # 第3行：示例数据
    sample = {
        'project_name': '示例：襄阳农高区智慧农业示范项目',
        'project_type_code': '现代种养类',
        'dispatch_status_code': '调度中',
        'construction_content': '示例：建设智慧农业大棚500亩，配套水肥一体化系统',
        'construction_unit': '示例：某某建设有限公司',
        'responsible_unit_code': '示例：nong_gao_qu_guan_wei_hui',
        'responsible_person': '示例：张三',
        'responsible_person_phone': '示例：13800138000',
        'roadmap_text': '项目签约|2025-07-01|pending;选址确定|2025-08-15|pending;开工建设|2025-10-01|pending',
        'progress_text': '2025-06-01|2025-06-15|完成项目立项审批;2025-06-16|2025-07-01|完成初步设计',
        'issue_text': '资金问题|项目资金缺口约500万|待回应||;用地问题|用地审批进度缓慢|协调中|已协调自然资源局加快审批|',
    }
    for col_idx, field in enumerate(fields, 1):
        val = sample.get(field.field_key, '')
        cell = ws.cell(row=3, column=col_idx, value=val)
        cell.font = Font(name='微软雅黑', size=10, color='909090')
        cell.alignment = data_align
        cell.border = thin_border

    # 列宽
    width_map = {
        'project_name': 38,
        'project_type_code': 16,
        'dispatch_status_code': 14,
        'construction_content': 36,
        'construction_unit': 26,
        'responsible_unit_code': 28,
        'responsible_person': 14,
        'responsible_person_phone': 16,
        'roadmap_text': 52,
        'progress_text': 48,
        'issue_text': 52,
    }
    for col_idx, field in enumerate(fields, 1):
        w = width_map.get(field.field_key, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # 找项目类型和调度状态列索引
    type_col = None
    status_col = None
    resp_unit_col = None
    for col_idx, field in enumerate(fields, 1):
        if field.field_key == 'project_type_code':
            type_col = col_idx
        elif field.field_key == 'dispatch_status_code':
            status_col = col_idx
        elif field.field_key == 'responsible_unit_code':
            resp_unit_col = col_idx

    # 下拉数据验证 — 项目类型
    if project_types and type_col:
        type_names = [d.name for d in project_types]
        ws_hidden = wb.create_sheet('_项目类型')
        ws_hidden.sheet_state = 'hidden'
        for i, name in enumerate(type_names, 1):
            ws_hidden.cell(row=i, column=1, value=name)
        dv = DataValidation(
            type='list',
            formula1=f'=_项目类型!$A$1:$A${len(type_names)}',
            allow_blank=True
        )
        dv.error = '请从下拉列表中选择项目类型'
        dv.errorTitle = '无效类型'
        letter = get_column_letter(type_col)
        dv.add(f'{letter}3:{letter}1000')
        ws.add_data_validation(dv)

    # 下拉数据验证 — 调度状态
    if dispatch_statuses and status_col:
        status_names = [d.name for d in dispatch_statuses]
        ws_hidden2 = wb.create_sheet('_调度状态')
        ws_hidden2.sheet_state = 'hidden'
        for i, name in enumerate(status_names, 1):
            ws_hidden2.cell(row=i, column=1, value=name)
        dv2 = DataValidation(
            type='list',
            formula1=f'=_调度状态!$A$1:$A${len(status_names)}',
            allow_blank=True
        )
        dv2.error = '请从下拉列表中选择调度状态'
        dv2.errorTitle = '无效状态'
        letter2 = get_column_letter(status_col)
        dv2.add(f'{letter2}3:{letter2}1000')
        ws.add_data_validation(dv2)

    # 下拉数据验证 — 责任单位
    if organizations and resp_unit_col:
        org_data = [(d.code, d.name) for d in organizations]
        ws_hidden3 = wb.create_sheet('_责任单位')
        ws_hidden3.sheet_state = 'hidden'
        for i, (code, name) in enumerate(org_data, 1):
            ws_hidden3.cell(row=i, column=1, value=code)
            ws_hidden3.cell(row=i, column=2, value=name)
        dv3 = DataValidation(
            type='list',
            formula1=f'=_责任单位!$A$1:$A${len(org_data)}',
            allow_blank=True
        )
        dv3.error = '请从下拉列表中选择责任单位 code'
        dv3.errorTitle = '无效单位'
        letter3 = get_column_letter(resp_unit_col)
        dv3.add(f'{letter3}3:{letter3}1000')
        ws.add_data_validation(dv3)

    ws.freeze_panes = 'A4'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='在建项目导入模板.xlsx'
    )


# ============================================================
# 导入预览
# ============================================================

def _get_import_fields():
    return ImportFieldConfigConstruction.query.filter_by(is_enabled=True)\
        .order_by(ImportFieldConfigConstruction.sort_order).all()


# 反向映射（field_key -> field_label），用于错误提示
FIELD_LABEL_MAP = {
    'project_name': '项目名称',
    'project_type_code': '项目类型',
    'dispatch_status_code': '调度状态',
    'construction_content': '建设内容',
    'construction_unit': '建设单位',
    'responsible_unit_code': '责任单位',
    'responsible_person': '责任人',
    'responsible_person_phone': '联系电话',
    'roadmap_text': '工作路径',
    'progress_text': '工作进展',
    'issue_text': '调度问题',
}

_PREVIEW_HEADERS = [
    '项目名称', '项目类型', '调度状态', '建设内容', '建设单位',
    '责任单位', '责任人', '工作路径数', '工作进展数', '调度问题数'
]


@admin_construction_import_bp.route('/construction-import/preview', methods=['POST'])
@dual_login_required
def import_preview():
    """上传 Excel 并返回预览数据（含子表文本解析统计）"""
    if 'file' not in request.files:
        return jsonify({'code': 1, 'message': '未找到上传文件'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'code': 1, 'message': '未选择文件'}), 400

    try:
        wb = load_workbook(file, read_only=True)
        ws = wb.active
    except Exception:
        return jsonify({'code': 1, 'message': '无法读取文件，请确认是有效的 Excel 文件'}), 400

    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value is not None:
            headers.append(str(cell.value).strip())

    # 获取配置的导入字段
    import_fields = _get_import_fields()
    expected_headers = [f.field_label for f in import_fields]

    # 校验表头是否一致
    if headers != expected_headers:
        return jsonify({
            'code': 1,
            'message': '模板有误，请选择正确的导入模板',
            'data': {'expected': expected_headers, 'actual': headers}
        }), 400

    # 构建 field_label -> field_key 映射
    field_map = {f.field_label: f.field_key for f in import_fields}
    required_fields = {f.field_key for f in import_fields if f.is_required}

    # 预加载字典数据（用于校验和名称→code 转换）
    type_name_to_code = {
        d.name: d.code
        for d in ConstructionProjectTypeDict.query.filter_by(is_active=True).all()
    }
    type_code_set = set(type_name_to_code.values())

    dispatch_name_to_code = {
        d.name: d.code
        for d in DispatchStatusDict.query.filter_by(is_active=True).all()
    }
    dispatch_code_set = set(dispatch_name_to_code.values())

    # 读取数据行
    raw_rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or str(v).strip() == '' for v in row):
            continue

        row_data = {}
        for col_idx, val in enumerate(row):
            if col_idx < len(headers):
                field_label = headers[col_idx]
                field_key = field_map.get(field_label, '')
                str_val = str(val).strip() if val is not None else ''
                row_data[field_key] = str_val

        # 跳过标记行
        first_val = list(row_data.values())[0] if row_data else ''
        if first_val in ('（必填）', '（选填）'):
            continue

        # 跳过示例行
        if first_val and first_val.startswith('示例：'):
            continue

        raw_rows.append({
            'row': row_idx,
            'data': row_data
        })

    # 解析并校验
    preview_rows = []
    all_errors = []

    for raw in raw_rows:
        row_data = raw['data']
        errors = []

        # 校验必填字段
        for field_key in required_fields:
            val = row_data.get(field_key, '')
            if not val:
                label = FIELD_LABEL_MAP.get(field_key, field_key)
                errors.append(f'第{raw["row"]}行：{label} 为必填项')

        # 校验项目类型（支持名称→code 转换）
        type_val = row_data.get('project_type_code', '')
        if type_val:
            if type_val in type_code_set:
                pass  # 已经是有效 code
            elif type_val in type_name_to_code:
                row_data['project_type_code'] = type_name_to_code[type_val]
            else:
                errors.append(f'第{raw["row"]}行：项目类型「{type_val}」不在字典中')

        # 校验调度状态（支持名称→code 转换）
        dispatch_val = row_data.get('dispatch_status_code', '')
        if dispatch_val:
            if dispatch_val in dispatch_code_set:
                pass
            elif dispatch_val in dispatch_name_to_code:
                row_data['dispatch_status_code'] = dispatch_name_to_code[dispatch_val]
            else:
                errors.append(f'第{raw["row"]}行：调度状态「{dispatch_val}」不在字典中')

        # 解析子表文本
        roadmap_items = parse_roadmap_text(row_data.get('roadmap_text', ''))
        progress_items = parse_progress_text(row_data.get('progress_text', ''))
        issue_items = parse_issue_text(row_data.get('issue_text', ''))

        # 校验子表文本解析
        roadmap_text_raw = row_data.get('roadmap_text', '')
        progress_text_raw = row_data.get('progress_text', '')
        issue_text_raw = row_data.get('issue_text', '')

        if roadmap_text_raw and not roadmap_items:
            errors.append(f'第{raw["row"]}行：工作路径文本格式有误，未能解析出有效记录')
        if progress_text_raw and not progress_items:
            errors.append(f'第{raw["row"]}行：工作进展文本格式有误，未能解析出有效记录')
        if issue_text_raw and not issue_items:
            errors.append(f'第{raw["row"]}行：调度问题文本格式有误，未能解析出有效记录')

        all_errors.extend(errors)

        preview_rows.append({
            'row': raw['row'],
            'data': row_data,
            'parsed': {
                'roadmap_count': len(roadmap_items),
                'roadmap_items': roadmap_items,
                'progress_count': len(progress_items),
                'progress_items': progress_items,
                'issue_count': len(issue_items),
                'issue_items': issue_items,
            },
            'errors': errors,
            '_valid': len(errors) == 0
        })

    valid_count = sum(1 for r in preview_rows if r['_valid'])
    error_count = len(preview_rows) - valid_count

    return jsonify({
        'code': 0,
        'data': {
            'headers': _PREVIEW_HEADERS,
            'rows': preview_rows,
            'total': len(preview_rows),
            'valid_count': valid_count,
            'error_count': error_count,
            'all_errors': all_errors,
        }
    })


# ============================================================
# 执行导入
# ============================================================

@admin_construction_import_bp.route('/construction-import/execute', methods=['POST'])
@dual_login_required
@visitor_block
def import_execute():
    """将预览通过的数据写入数据库"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 1, 'message': '数据为空'}), 400

    rows = data.get('rows', [])
    if not rows:
        return jsonify({'code': 1, 'message': '无数据可导入'}), 400

    valid_rows = [r for r in rows if r.get('_valid')]
    if not valid_rows:
        return jsonify({'code': 1, 'message': '没有有效数据行'}), 400

    # 先闭合现有项目的序号缺口
    _renumber_projects()

    # 获取当前最大 order_no
    max_no_result = db.session.query(db.func.max(ConstructionProject.order_no)).scalar()
    next_order_no = (max_no_result or 0) + 1

    # 预加载字典数据
    type_code_set = {
        d.code for d in ConstructionProjectTypeDict.query.filter_by(is_active=True).all()
    }
    dispatch_code_set = {
        d.code for d in DispatchStatusDict.query.filter_by(is_active=True).all()
    }

    imported = 0
    for row in valid_rows:
        row_data = row['data']

        # 解析字典 code（可能已在上传时转换，这里做兜底）
        project_type_code = row_data.get('project_type_code', '')
        if project_type_code and project_type_code not in type_code_set:
            # 尝试名称匹配
            t = ConstructionProjectTypeDict.query.filter_by(name=project_type_code, is_active=True).first()
            project_type_code = t.code if t else project_type_code

        dispatch_status_code = row_data.get('dispatch_status_code', 'dispatching')
        if dispatch_status_code and dispatch_status_code not in dispatch_code_set:
            t = DispatchStatusDict.query.filter_by(name=dispatch_status_code, is_active=True).first()
            dispatch_status_code = t.code if t else 'dispatching'

        # 创建项目
        project = ConstructionProject(
            order_no=next_order_no,
            project_name=row_data.get('project_name', ''),
            project_type_code=project_type_code,
            dispatch_status_code=dispatch_status_code,
            construction_content=row_data.get('construction_content', ''),
            construction_unit=row_data.get('construction_unit', ''),
            responsible_unit_code=row_data.get('responsible_unit_code', ''),
            responsible_person=row_data.get('responsible_person', ''),
            responsible_person_phone=row_data.get('responsible_person_phone', ''),
        )
        db.session.add(project)
        db.session.flush()  # 获取 project.id
        imported += 1
        if imported % 20 == 0:
            db.session.flush()  # 每 20 条 flush，释放写锁

        # 解析并创建子表数据
        parsed = row.get('parsed', {})

        # 工作路径图
        for item in (parsed.get('roadmap_items') or []):
            db.session.add(WorkRoadmapItem(
                project_id=project.id,
                sort_order=item.get('sort_order', 0),
                content=item.get('content', ''),
                planned_date=_parse_date(item.get('planned_date')),
                actual_date=_parse_date(item.get('actual_date')),
                status=item.get('status', 'pending'),
                is_delayed=item.get('is_delayed', False),
                delay_reason=item.get('delay_reason', ''),
                cancel_reason=item.get('cancel_reason', ''),
            ))

        # 工作进展
        for item in (parsed.get('progress_items') or []):
            db.session.add(WorkProgress(
                project_id=project.id,
                start_date=_parse_date(item.get('start_date')),
                end_date=_parse_date(item.get('end_date')),
                content=item.get('content', ''),
            ))

        # 调度问题
        for item in (parsed.get('issue_items') or []):
            db.session.add(ProjectIssue(
                project_id=project.id,
                issue_type_code=item.get('issue_type_code', ''),
                issue_description=item.get('issue_description', ''),
                resolution_status_code=item.get('resolution_status_code', 'pending'),
                resolution_note=item.get('resolution_note', ''),
                main_department_code=item.get('main_department_code', ''),
            ))

        next_order_no += 1
        imported += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        import traceback
        from flask import current_app
        current_app.logger.error(f'在建项目导入失败: {e}\n{traceback.format_exc()}')
        return jsonify({
            'code': 1,
            'message': f'导入失败（数据库写入错误）：{str(e)}'
        }), 500
    return jsonify({'code': 0, 'message': f'成功导入 {imported} 个项目', 'data': {'count': imported}})
