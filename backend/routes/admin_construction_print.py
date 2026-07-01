"""
在建项目打印 API
提供打印模板 CRUD、字段配置、数据解析（复用导出模块的聚合逻辑）
"""
import json
import re as _re
from datetime import date, datetime
from flask import request, jsonify, send_file
from models import PrintTemplate, PrintFieldConfig, ConstructionProject
from extensions import db
from routes import admin_construction_print_bp
from routes.business_auth import dual_login_required, visitor_block

# 复用导出模块的数据解析函数
from routes.admin_construction_export import _resolve_project_row as _export_resolve_row
from routes.admin_construction_export import _aggregate_work_progresses, _aggregate_issues, _aggregate_work_roadmap
from routes.admin_construction_export import _get_last2_windows


def _sanitize_filename(name):
    """移除文件名中的非法字符（仅去除 Windows 真正禁止的字符）"""
    name = _re.sub(r'[\[\]*?"<>|\\/:]', '', name)
    name = _re.sub(r'_+', '_', name)
    return name.strip() or f'导出文件_{date.today().strftime("%Y%m%d")}'


# ============================================================
# 模板 CRUD
# ============================================================

def _get_default_template_id():
    """获取默认打印模板ID（entity_type='construction' 的第一个模板）"""
    tpl = PrintTemplate.query.filter_by(entity_type='construction').order_by(PrintTemplate.id).first()
    return tpl.id if tpl else 1


@admin_construction_print_bp.route('/construction-print/templates', methods=['GET'])
@dual_login_required
def get_templates():
    """列出所有打印模板"""
    entity_type = request.args.get('entity_type', 'construction').strip()
    templates = PrintTemplate.query.filter_by(entity_type=entity_type)\
        .order_by(PrintTemplate.id.asc()).all()
    return jsonify({'code': 0, 'data': [t.to_dict() for t in templates]})


@admin_construction_print_bp.route('/construction-print/templates', methods=['POST'])
@dual_login_required
@visitor_block
def create_template():
    """新建打印模板"""
    data = request.get_json()
    if not data or not data.get('name'):
        return jsonify({'code': 1, 'message': '模板名称不能为空'}), 400
    entity_type = data.get('entity_type', 'construction')
    template = PrintTemplate(name=data['name'].strip(), entity_type=entity_type)
    db.session.add(template)
    db.session.flush()

    # 从已有模板（排除自身）复制字段配置
    default_template = PrintTemplate.query.filter_by(entity_type=entity_type)\
        .filter(PrintTemplate.id != template.id)\
        .order_by(PrintTemplate.id.asc()).first()
    if default_template:
        default_fields = PrintFieldConfig.query.filter_by(template_id=default_template.id)\
            .order_by(PrintFieldConfig.sort_order).all()
        for f in default_fields:
            db.session.add(PrintFieldConfig(
                template_id=template.id,
                field_key=f.field_key,
                field_label=f.field_label,
                is_visible=True,
                column_width=f.column_width,
                sort_order=f.sort_order
            ))

    db.session.commit()
    return jsonify({'code': 0, 'data': template.to_dict(), 'message': '模板已创建'})


@admin_construction_print_bp.route('/construction-print/templates/<int:template_id>', methods=['PUT'])
@dual_login_required
@visitor_block
def update_template(template_id):
    """更新模板（名称 + 布局参数）"""
    tpl = PrintTemplate.query.get(template_id)
    if not tpl:
        return jsonify({'code': 1, 'message': '模板不存在'}), 404
    data = request.get_json()
    if not data:
        return jsonify({'code': 1, 'message': '请求数据为空'}), 400

    if 'name' in data and data['name']:
        tpl.name = data['name'].strip()
    for key in ('paper_size', 'orientation', 'font_family', 'font_size',
                'header_font_size', 'margin_top', 'margin_bottom', 'margin_left',
                'margin_right', 'title_font_family', 'title_font_size',
                'table_header_font_family', 'table_header_font_size',
                'cell_font_family', 'cell_font_size',
                'sub_title_font_size', 'border_style', 'group_by'):
        if key in data:
            setattr(tpl, key, data[key])
    if 'show_page_number' in data:
        tpl.show_page_number = bool(data['show_page_number'])
    if 'show_print_meta' in data:
        tpl.show_print_meta = bool(data['show_print_meta'])
    if 'title_bold' in data:
        tpl.title_bold = bool(data['title_bold'])
    if 'subtitle_bold' in data:
        tpl.subtitle_bold = bool(data['subtitle_bold'])

    db.session.commit()
    return jsonify({'code': 0, 'data': tpl.to_dict(), 'message': '模板已更新'})


@admin_construction_print_bp.route('/construction-print/templates/<int:template_id>', methods=['DELETE'])
@dual_login_required
@visitor_block
def delete_template(template_id):
    """删除模板及其字段配置"""
    tpl = PrintTemplate.query.get(template_id)
    if not tpl:
        return jsonify({'code': 1, 'message': '模板不存在'}), 404
    remaining = PrintTemplate.query.filter_by(entity_type=tpl.entity_type).count()
    if remaining <= 1:
        return jsonify({'code': 1, 'message': '至少保留一个打印模板'}), 400
    PrintFieldConfig.query.filter_by(template_id=template_id).delete()
    db.session.delete(tpl)
    db.session.commit()
    return jsonify({'code': 0, 'message': '模板已删除'})


# ============================================================
# 字段配置 CRUD
# ============================================================

@admin_construction_print_bp.route('/construction-print/fields', methods=['GET'])
@dual_login_required
def get_print_fields():
    """获取指定模板的打印字段配置"""
    template_id = request.args.get('template_id', 0, type=int) or _get_default_template_id()
    fields = PrintFieldConfig.query.filter_by(template_id=template_id)\
        .order_by(PrintFieldConfig.sort_order).all()
    return jsonify({'code': 0, 'data': [f.to_dict() for f in fields]})


@admin_construction_print_bp.route('/construction-print/fields', methods=['PUT'])
@dual_login_required
@visitor_block
def update_print_fields():
    """批量更新打印字段配置"""
    data = request.get_json()
    if not data or not isinstance(data, list):
        return jsonify({'code': 1, 'message': '请求数据格式错误'}), 400

    template_id = None
    saved_ids = set()
    for item in data:
        field_id = item.get('id')
        if field_id and field_id > 0:
            saved_ids.add(field_id)
            field = PrintFieldConfig.query.get(field_id)
            if field:
                template_id = field.template_id
        elif item.get('_new') and item.get('template_id'):
            template_id = int(item['template_id'])
            field = PrintFieldConfig(
                template_id=template_id,
                field_key=item.get('field_key', ''),
                field_label=item.get('field_label', ''),
                is_visible=item.get('is_visible', True),
                is_custom=item.get('is_custom', True),
                column_width=item.get('column_width', 120),
                sort_order=item.get('sort_order', 0)
            )
            db.session.add(field)
            continue

        if field:
            if 'field_label' in item:
                field.field_label = item['field_label']
            if 'is_visible' in item:
                field.is_visible = bool(item['is_visible'])
            if 'column_width' in item:
                field.column_width = int(item['column_width'])
            if 'sort_order' in item:
                field.sort_order = int(item['sort_order'])

    if template_id:
        existing_ids = {f.id for f in PrintFieldConfig.query.filter_by(template_id=template_id).all()}
        to_delete = existing_ids - saved_ids
        if to_delete:
            PrintFieldConfig.query.filter(PrintFieldConfig.id.in_(to_delete)).delete(synchronize_session=False)

    db.session.commit()
    fields = PrintFieldConfig.query.filter_by(template_id=template_id)\
        .order_by(PrintFieldConfig.sort_order).all() if template_id else []
    return jsonify({'code': 0, 'data': [f.to_dict() for f in fields], 'message': '配置已保存'})


# ============================================================
# 打印数据解析
# ============================================================

# V3: 项目类型优先级顺序
CONSTRUCTION_TYPE_ORDER = ['示范类', '科创类', '现代种养类', '产业类', '基础设施类']


def _group_rows_construction(rows):
    """在建项目分组：先按调度状态，再按项目类型
    返回 [{title, section, is_section, rows}]
    """
    from models import DispatchStatusDict, ConstructionProjectTypeDict

    dispatch_map = {d.code: d.name for d in DispatchStatusDict.query.all()}
    type_map = {d.code: d.name for d in ConstructionProjectTypeDict.query.all()}

    # 分开 调度中 / 不予调度
    dispatching_rows = [r for r in rows if r.get('dispatch_status_code') == dispatch_map.get('dispatching', '调度中')]
    no_dispatch_rows = [r for r in rows if r.get('dispatch_status_code') == dispatch_map.get('no_dispatch', '不予调度')]
    # 兜底：任何不属于以上两类的行归入"其他"
    known = set()
    for r in dispatching_rows + no_dispatch_rows:
        known.add(id(r))
    other_rows = [r for r in rows if id(r) not in known]

    groups = []

    # 调度中部分：按类型子分组
    _append_type_groups(groups, dispatching_rows, 'dispatching')

    # 其他未知状态
    if other_rows:
        _append_type_groups(groups, other_rows, 'other')

    # 不予调度部分：先插入 section 标题，再按类型子分组
    if no_dispatch_rows:
        section_name = dispatch_map.get('no_dispatch', '不予调度')
        groups.append({
            'title': section_name,
            'section': 'no_dispatch',
            'is_section': True,
            'rows': []
        })
        _append_type_groups(groups, no_dispatch_rows, 'no_dispatch')

    if not groups:
        groups.append({'title': None, 'rows': []})

    return groups


def _append_type_groups(groups, rows, section):
    """按项目类型对行进行子分组并追加到 groups 列表"""
    from models import ConstructionProjectTypeDict
    type_map = {d.code: d.name for d in ConstructionProjectTypeDict.query.all()}
    seen_types = set()

    # 按优先级顺序处理
    for type_name in CONSTRUCTION_TYPE_ORDER:
        type_rows = [r for r in rows if r.get('project_type_code') == type_name]
        if type_rows:
            groups.append({
                'title': type_name,
                'section': section,
                'is_section': False,
                'rows': type_rows
            })
            seen_types.add(type_name)

    # 其他未知类型
    other_type_rows = [r for r in rows if r.get('project_type_code') not in seen_types]
    if other_type_rows:
        other_types = []
        for r in other_type_rows:
            t = r.get('project_type_code', '')
            if t not in other_types:
                other_types.append(t)
        for t in other_types:
            t_rows = [r for r in other_type_rows if r.get('project_type_code') == t]
            if t_rows:
                groups.append({
                    'title': t,
                    'section': section,
                    'is_section': False,
                    'rows': t_rows
                })


def _resolve_project_row(p, progress_range='', work_path_range='pending', last2_windows=None):
    """将 ConstructionProject 解析为展示用 dict（复用导出模块的解析逻辑）"""
    return _export_resolve_row(p, progress_range, work_path_range, last2_windows)


@admin_construction_print_bp.route('/construction-print/data', methods=['GET'])
@dual_login_required
def get_print_data():
    """解析打印数据，返回 JSON（模板信息 + 表头 + 数据行）"""
    ids_str = request.args.get('project_ids', '')
    project_ids = [int(x) for x in ids_str.split(',') if x.strip().isdigit()] if ids_str else []

    template_id = request.args.get('template_id', 0, type=int) or _get_default_template_id()
    progress_range = request.args.get('progress_range', '').strip()
    progress_mode = request.args.get('progress_mode', 'aggregate').strip()
    work_path_range = request.args.get('work_path_range', 'pending').strip()

    # 最近2条：预计算时间窗口（限定在勾选项目范围内）
    last2_windows = _get_last2_windows(project_ids) if progress_range == 'last2' else None

    template = PrintTemplate.query.get(template_id)
    if not template:
        return jsonify({'code': 1, 'message': '打印模板不存在'}), 404

    fields = PrintFieldConfig.query.filter_by(template_id=template_id, is_visible=True)\
        .order_by(PrintFieldConfig.sort_order).all()
    if not fields:
        return jsonify({'code': 1, 'message': '未配置打印字段'}), 400

    q = ConstructionProject.query.filter_by(is_deleted=False) \
        .filter(ConstructionProject.dispatch_status_code != 'exited')
    if project_ids:
        q = q.filter(ConstructionProject.id.in_(project_ids))
    projects = q.order_by(ConstructionProject.order_no.asc()).all()

    headers = [{'field_key': f.field_key, 'field_label': f.field_label,
                'column_width': f.column_width} for f in fields]

    if progress_mode == 'progress':
        # 按行展开：每个工作进展拆分为独立行
        from models import WorkProgress
        rows = []
        for p in projects:
            base = _resolve_project_row(p, progress_range, work_path_range, last2_windows)
            progresses = WorkProgress.query.filter_by(project_id=p.id)\
                .order_by(WorkProgress.start_date.desc()).all()
            if progresses:
                for wp in progresses:
                    row = dict(base)
                    start = wp.start_date.isoformat() if wp.start_date else ''
                    end = wp.end_date.isoformat() if wp.end_date else ''
                    row['_progress_item'] = f'{start}~{end}: {wp.content or ""}'
                    rows.append(row)
            else:
                base['_progress_item'] = ''
                rows.append(base)
    else:
        rows = [_resolve_project_row(p, progress_range, work_path_range, last2_windows) for p in projects]

    # V3: 按调度状态+项目类型分组
    groups = _group_rows_construction(rows)

    return jsonify({
        'code': 0,
        'data': {
            'template': template.to_dict(),
            'headers': headers,
            'rows': rows,     # 保留扁平列表兼容
            'groups': groups,  # V3 分组数据
            'total': len(rows),
            'progress_mode': progress_mode,
            'group_mode': 'dispatch_status+project_type'
        }
    })


# ============================================================
# Excel 下载（使用 PrintTemplate 生成）
# ============================================================

def _fill_template_file_from_mappings(abs_path, template, mappings, groups, template_name, progress_range=''):
    """使用上传的 .xlsx 模板 + 列映射填充在建项目数据"""
    import io as _io
    import openpyxl
    from copy import copy as _copy
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.cell.rich_text import TextBlock, CellRichText
    from openpyxl.cell.text import InlineFont
    from openpyxl.utils import column_index_from_string

    wb = openpyxl.load_workbook(abs_path)
    ws = wb.active

    col_map = {m.column_letter: m.field_key for m in mappings if m.field_key}
    col_indices = {letter: column_index_from_string(letter) for letter in col_map}

    data_start = template.data_start_row or 3
    header_row = template.header_row or 2
    # 样式参考行：数据行在第一行数据处（如有分组标题则在 data_start + 1）
    style_ref_row = data_start + 1 if template.has_group_title else data_start

    thin_border = Border(
        left=Side(style='thin', color='333333'), right=Side(style='thin', color='333333'),
        top=Side(style='thin', color='333333'), bottom=Side(style='thin', color='333333'),
    )
    _ss = template.sub_title_font_size or 20
    section_font = Font(name='微软雅黑', size=_ss, bold=True, color='000000')
    section_align = Alignment(horizontal='center', vertical='center')

    # 缓存样式（显式复制，避免 openpyxl copy 不完整）
    style_cache = {}
    for col_letter, col_idx in col_indices.items():
        ref_cell = ws.cell(row=style_ref_row, column=col_idx)
        rf = ref_cell.font; ra = ref_cell.alignment
        style_cache[col_letter] = {
            'font': Font(
                name=rf.name or '宋体',
                size=rf.size or 18,
                bold=rf.bold or False,
                color=rf.color if rf.color else '000000',
            ),
            'alignment': Alignment(
                horizontal=ra.horizontal or 'left',
                vertical=ra.vertical or 'center',
                wrap_text=ra.wrap_text if ra.wrap_text is not None else True,
            ),
            'border': thin_border,
            'fill': _copy(ref_cell.fill) if ref_cell.fill else PatternFill(),
        }

    # 清除并解除合并
    merges_to_remove = [str(m) for m in ws.merged_cells.ranges if m.min_row >= data_start]
    for mr_str in merges_to_remove:
        ws.unmerge_cells(mr_str)
    if ws.max_row >= data_start:
        ws.delete_rows(data_start, ws.max_row - data_start + 1)

    if template.title_row and template.title_row > 0:
        ws.cell(row=template.title_row, column=1).value = template_name

    current_row = data_start
    total_cols = max(col_indices.values()) if col_indices else 10
    # 缓存列宽用于自动行高计算
    col_widths = {}
    for col_letter in col_indices:
        cd = ws.column_dimensions.get(col_letter)
        col_widths[col_letter] = cd.width if cd and cd.width else 10

    # 确定进度列（用于 last2 模式下的富文本字体）
    progress_col_letter = None
    for col_letter, field_key in col_map.items():
        if field_key == 'work_progresses':
            progress_col_letter = col_letter
            break

    # 除进度列外，统一使用 宋体 12pt
    for col_letter in style_cache:
        if col_letter != progress_col_letter:
            style_cache[col_letter]['font'] = Font(name='宋体', size=12, bold=False, color='000000')

    # 序号、项目名称、责任单位/责任人 — 居中对齐
    _center_field_keys = {'order_no', 'project_name', '_combined_contact'}
    for col_letter, field_key in col_map.items():
        if field_key in _center_field_keys and col_letter in style_cache:
            style_cache[col_letter]['alignment'] = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # last2 模式专用字体（TextBlock 需要 InlineFont）
    last2_label_font = InlineFont(rFont='宋体', sz=14.0, b=True, color='000000')
    last2_this_content_font = InlineFont(rFont='宋体', sz=14.0, b=True, color='000000')
    last2_last_content_font = InlineFont(rFont='宋体', sz=12.0, b=False, color='000000')
    last2_other_inline_font = InlineFont(rFont='宋体', sz=12.0, b=False, color='000000')
    # 注意：cell.font 必须用 Font，不能用 InlineFont（InlineFont 仅用于 TextBlock 内部）
    last2_cell_font = Font(name='宋体', size=12, bold=False, color='000000')
    last2_other_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    is_last2 = (progress_range == 'last2')

    def _build_last2_rich_text(text):
        """将最近2条进展文本解析为 CellRichText，实现混排字体
        输入格式: '上周：xxx\\n本周：yyy'
        换行符嵌入末个 TextBlock 的 text 末尾，避免 CellRichText 中出现裸字符串
        本周内容内部的换行也保持宋体14号加粗"""
        blocks = []
        lines = text.split('\n')
        current_section = None  # 'last' | 'this'
        for idx, line in enumerate(lines):
            tail = '\n' if idx < len(lines) - 1 else ''
            if line.startswith('上周：'):
                current_section = 'last'
                blocks.append(TextBlock(font=last2_label_font, text='上周：'))
                blocks.append(TextBlock(font=last2_last_content_font, text=line[3:] + tail))
            elif line.startswith('本周：'):
                current_section = 'this'
                blocks.append(TextBlock(font=last2_label_font, text='本周：'))
                blocks.append(TextBlock(font=last2_this_content_font, text=line[3:] + tail))
            elif current_section == 'this':
                blocks.append(TextBlock(font=last2_this_content_font, text=line + tail))
            elif current_section == 'last':
                blocks.append(TextBlock(font=last2_last_content_font, text=line + tail))
            else:
                blocks.append(TextBlock(font=last2_other_inline_font, text=line + tail))
        return CellRichText(*blocks)

    def _calc_row_height(r):
        """根据单元格内容和列宽自动计算行高，使内容完全展开"""
        max_lines = 1
        for col_letter, col_idx in col_indices.items():
            cell = ws.cell(row=r, column=col_idx)
            val = cell.value
            if not val:
                continue
            # 兼容 CellRichText：提取纯文本计算行数（\n 已嵌入 TextBlock.text）
            if isinstance(val, CellRichText):
                try:
                    text = ''.join(getattr(tb, 'text', str(tb)) for tb in val)
                except Exception:
                    text = str(val)
            else:
                text = str(val)
            if not text:
                continue
            lines = text.split('\n')
            line_count = len(lines)
            cw = col_widths.get(col_letter, 10)
            chars_per_line = max(3, int(cw * 1.5))
            for line in lines:
                if len(line) > chars_per_line:
                    line_count += (len(line) - 1) // chars_per_line
            max_lines = max(max_lines, line_count)
        return max(28, max_lines * 24)

    for group in groups:
        rows = group.get('rows', [])
        if not rows:
            continue

        if group.get('title'):
            # 子标题合并 A-L 列（前12列），M、N 列不参与合并，保留独立单元格
            merge_end_col = total_cols - 2 if total_cols > 2 else total_cols
            # 所有列设置边框
            for c in range(1, total_cols + 1):
                ws.cell(row=current_row, column=c).border = thin_border
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=merge_end_col)
            title_cell = ws.cell(row=current_row, column=1)
            title_cell.value = group['title']
            title_cell.font = section_font
            title_cell.alignment = section_align
            ws.row_dimensions[current_row].height = 34
            current_row += 1

        for row_data in rows:
            for col_letter, field_key in col_map.items():
                col_idx = col_indices[col_letter]
                val = row_data.get(field_key, '')
                if val is None:
                    val = ''
                cell = ws.cell(row=current_row, column=col_idx)

                if is_last2 and col_letter == progress_col_letter and val:
                    # 最近2条模式 — 进度列使用 CellRichText 混排字体
                    cell.value = _build_last2_rich_text(str(val))
                    cell.alignment = style_cache.get(col_letter, {}).get('alignment') or last2_other_align
                    cell.border = style_cache.get(col_letter, {}).get('border', thin_border)
                    cell.fill = style_cache.get(col_letter, {}).get('fill', PatternFill())
                elif is_last2:
                    # 最近2条模式 — 其余列使用 12pt 宋体
                    cell.value = str(val) if not isinstance(val, (int, float)) else val
                    cell.font = last2_cell_font
                    cell.alignment = style_cache.get(col_letter, {}).get('alignment', last2_other_align)
                    cell.border = thin_border
                    cell.fill = PatternFill()
                else:
                    cell.value = str(val) if not isinstance(val, (int, float)) else val
                    style = style_cache.get(col_letter)
                    if style:
                        cell.font = style['font']; cell.alignment = style['alignment']
                        cell.border = style['border']; cell.fill = style['fill']
            ws.row_dimensions[current_row].height = _calc_row_height(current_row)
            current_row += 1

    ws.freeze_panes = f'A{header_row + 1}'
    output = _io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@admin_construction_print_bp.route('/construction-print/download', methods=['GET'])
@dual_login_required
def print_download():
    """使用打印模板生成并下载 Excel 文件"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.cell.rich_text import TextBlock, CellRichText
    from openpyxl.cell.text import InlineFont
    from openpyxl.utils import get_column_letter

    ids_str = request.args.get('project_ids', '')
    project_ids = [int(x) for x in ids_str.split(',') if x.strip().isdigit()] if ids_str else []

    template_id = request.args.get('template_id', 0, type=int) or _get_default_template_id()
    progress_range = request.args.get('progress_range', '').strip()
    progress_mode = request.args.get('progress_mode', 'aggregate').strip()
    work_path_range = request.args.get('work_path_range', 'pending').strip()

    # 最近2条：预计算时间窗口（限定在勾选项目范围内）
    last2_windows = _get_last2_windows(project_ids) if progress_range == 'last2' else None

    template = PrintTemplate.query.get(template_id)
    template_name = template.name if template else '在建项目库'

    fields = PrintFieldConfig.query.filter_by(template_id=template_id, is_visible=True)\
        .order_by(PrintFieldConfig.sort_order).all()
    if not fields:
        return jsonify({'code': 1, 'message': '未配置打印字段'}), 400

    q = ConstructionProject.query.filter_by(is_deleted=False) \
        .filter(ConstructionProject.dispatch_status_code != 'exited')
    if project_ids:
        q = q.filter(ConstructionProject.id.in_(project_ids))
    projects = q.order_by(ConstructionProject.order_no.asc()).all()

    # 字典映射
    from models import ConstructionProjectTypeDict, DispatchStatusDict, OrganizationDict, WorkProgress
    type_map = {d.code: d.name for d in ConstructionProjectTypeDict.query.all()}
    dispatch_map = {d.code: d.name for d in DispatchStatusDict.query.all()}
    org_map = {d.code: d.name for d in OrganizationDict.query.all()}

    # 样式（使用打印模板的字体设置；数据单元格统一 宋体 12pt）
    _tf = template.title_font_family or template.font_family or '微软雅黑'
    _ts = template.title_font_size or template.header_font_size or 14
    _hf = template.table_header_font_family or template.font_family or '微软雅黑'
    _hs = template.table_header_font_size or template.font_size or 10
    _cf = '宋体'
    _cs = 12

    title_font = Font(name=_tf, size=_ts, bold=True, color='1A3A5C')
    title_align = Alignment(horizontal='center', vertical='center')
    header_font = Font(name=_hf, size=_hs, bold=True, color='000000')
    header_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name=_cf, size=_cs)
    cell_align = Alignment(vertical='top', wrap_text=True)
    cell_align_center = Alignment(horizontal='center', vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='333333'),
        right=Side(style='thin', color='333333'),
        top=Side(style='thin', color='333333'),
        bottom=Side(style='thin', color='333333'),
    )
    title_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # 最近2条模式专用字体和辅助函数
    is_last2 = (progress_range == 'last2')
    last2_cell_font = Font(name='宋体', size=12, bold=False, color='000000')

    def _build_last2_rich_text_prog(text):
        """将最近2条进展文本解析为 CellRichText（程序化生成路径用）
        换行符嵌入末个 TextBlock 的 text 末尾，避免 CellRichText 中出现裸字符串
        本周内容内部的换行也保持宋体14号加粗"""
        label_font = InlineFont(rFont='宋体', sz=14.0, b=True, color='000000')
        this_content_font = InlineFont(rFont='宋体', sz=14.0, b=True, color='000000')
        last_content_font = InlineFont(rFont='宋体', sz=12.0, b=False, color='000000')
        other_font = InlineFont(rFont='宋体', sz=12.0, b=False, color='000000')
        blocks = []
        lines_text = text.split('\n')
        current_section = None  # 'last' | 'this'
        for idx, line in enumerate(lines_text):
            tail = '\n' if idx < len(lines_text) - 1 else ''
            if line.startswith('上周：'):
                current_section = 'last'
                blocks.append(TextBlock(font=label_font, text='上周：'))
                blocks.append(TextBlock(font=last_content_font, text=line[3:] + tail))
            elif line.startswith('本周：'):
                current_section = 'this'
                blocks.append(TextBlock(font=label_font, text='本周：'))
                blocks.append(TextBlock(font=this_content_font, text=line[3:] + tail))
            elif current_section == 'this':
                blocks.append(TextBlock(font=this_content_font, text=line + tail))
            elif current_section == 'last':
                blocks.append(TextBlock(font=last_content_font, text=line + tail))
            else:
                blocks.append(TextBlock(font=other_font, text=line + tail))
        return CellRichText(*blocks)

    # ---- 模板文件填充路径 ----
    from models import TemplateFieldMapping
    template_mappings = TemplateFieldMapping.query.filter_by(template_id=template_id)\
        .filter(TemplateFieldMapping.field_key != '').all() if template else []
    use_template_file = template and template.template_file and len(template_mappings) > 0

    if use_template_file:
        all_rows = [_resolve_project_row(p, progress_range, work_path_range, last2_windows) for p in projects]
        groups = _group_rows_construction(all_rows)
        import os as _os
        abs_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), '..',
                                 template.template_file.lstrip('/'))
        output = _fill_template_file_from_mappings(abs_path, template, template_mappings, groups, template_name, progress_range)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{_sanitize_filename(template_name)}.xlsx'
        )

    wb = Workbook()
    ws = wb.active
    ws.title = template_name[:31]

    if progress_mode == 'progress':
        project_fields = [f for f in fields if f.field_key not in ('issues', 'work_roadmap', 'work_progresses')]
        progress_cols = [
            {'field_key': 'progress_start', 'field_label': '进展开始日期', 'column_width': 120},
            {'field_key': 'progress_end', 'field_label': '进展结束日期', 'column_width': 120},
            {'field_key': 'progress_content', 'field_label': '进展内容', 'column_width': 400},
        ]
        total_cols = len(project_fields) + len(progress_cols)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(row=1, column=1, value=template_name)
        title_cell.font = title_font; title_cell.alignment = title_align; title_cell.fill = title_fill
        for c in range(1, total_cols + 1):
            ws.cell(row=1, column=c).border = thin_border

        for col_idx, f in enumerate(project_fields, 1):
            cell = ws.cell(row=2, column=col_idx, value=f.field_label)
            cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align; cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = f.column_width / 7
        for col_idx, pc in enumerate(progress_cols, len(project_fields) + 1):
            cell = ws.cell(row=2, column=col_idx, value=pc['field_label'])
            cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align; cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = pc['column_width'] / 7

        current_row = 3
        for p in projects:
            project_row = _resolve_project_row(p, progress_range, work_path_range, last2_windows)
            progresses = WorkProgress.query.filter_by(project_id=p.id)\
                .order_by(WorkProgress.start_date.desc()).all()
            if not progresses:
                for col_idx, f in enumerate(project_fields, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=project_row.get(f.field_key, ''))
                    cell.font = cell_font; cell.alignment = cell_align; cell.border = thin_border
                for col_idx in range(len(project_fields) + 1, total_cols + 1):
                    cell = ws.cell(row=current_row, column=col_idx, value='')
                    cell.font = cell_font; cell.alignment = cell_align_center; cell.border = thin_border
                current_row += 1
            else:
                start_row = current_row
                for wp in progresses:
                    for col_idx, f in enumerate(project_fields, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=project_row.get(f.field_key, ''))
                        cell.font = cell_font; cell.alignment = cell_align; cell.border = thin_border
                    start_str = wp.start_date.isoformat() if wp.start_date else ''
                    end_str = wp.end_date.isoformat() if wp.end_date else ''
                    progress_data = {
                        'progress_start': start_str,
                        'progress_end': end_str,
                        'progress_content': wp.content or '',
                    }
                    for col_idx, pc in enumerate(progress_cols, len(project_fields) + 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=progress_data.get(pc['field_key'], ''))
                        cell.font = cell_font; cell.alignment = cell_align; cell.border = thin_border
                    current_row += 1
                if current_row - start_row > 1:
                    for col_idx in range(1, len(project_fields) + 1):
                        ws.merge_cells(start_row=start_row, start_column=col_idx,
                                       end_row=current_row - 1, end_column=col_idx)
        ws.freeze_panes = 'A3'
    else:
        # V3: 分组模式写入（调度状态 + 项目类型二级分组）
        total_cols = len(fields)
        # 标题行
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(row=1, column=1, value=template_name)
        title_cell.font = title_font; title_cell.alignment = title_align; title_cell.fill = title_fill
        for c in range(1, total_cols + 1):
            ws.cell(row=1, column=c).border = thin_border

        # 列标题行
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=2, column=col_idx, value=field.field_label)
            cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align; cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = field.column_width / 7

        # 解析所有行并分组
        all_rows = [_resolve_project_row(p, progress_range, work_path_range, last2_windows) for p in projects]
        groups = _group_rows_construction(all_rows)

        # V3 子标题/分区标题样式
        _ss = template.sub_title_font_size or 12
        _sb = template.subtitle_bold if template.subtitle_bold is not None else True
        subtitle_font = Font(name=_tf, size=_ss, bold=_sb, color='000000')
        subtitle_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        section_font = Font(name=_tf, size=_ss, bold=True, color='000000')
        section_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        subtitle_align = Alignment(horizontal='center', vertical='center')

        current_row = 3
        for group in groups:
            if group.get('title'):
                # 子标题合并 A-L 列（前12列），M、N 列不参与合并
                merge_end_col = total_cols - 2 if total_cols > 2 else total_cols
                for c in range(1, total_cols + 1):
                    ws.cell(row=current_row, column=c).border = thin_border
                ws.merge_cells(start_row=current_row, start_column=1,
                               end_row=current_row, end_column=merge_end_col)
                sub_cell = ws.cell(row=current_row, column=1, value=group['title'])
                if group.get('is_section'):
                    sub_cell.font = section_font
                    sub_cell.fill = section_fill
                else:
                    sub_cell.font = subtitle_font
                    sub_cell.fill = subtitle_fill
                sub_cell.alignment = subtitle_align
                current_row += 1

            for row_data in group.get('rows', []):
                for col_idx, field in enumerate(fields, 1):
                    val = row_data.get(field.field_key, '')
                    # 最近2条模式：进度列使用 CellRichText 混排，其余列 12pt 宋体
                    if is_last2 and field.field_key == 'work_progresses' and val:
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.value = _build_last2_rich_text_prog(str(val))
                        cell.alignment = cell_align
                        cell.border = thin_border
                    elif is_last2:
                        cell = ws.cell(row=current_row, column=col_idx, value=val)
                        cell.font = last2_cell_font
                        cell.alignment = cell_align
                        cell.border = thin_border
                    else:
                        cell = ws.cell(row=current_row, column=col_idx, value=val)
                        cell.font = cell_font; cell.alignment = cell_align; cell.border = thin_border
                current_row += 1

        ws.freeze_panes = 'A3'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{_sanitize_filename(template_name)}.xlsx'
    )
