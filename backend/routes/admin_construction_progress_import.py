import io
from datetime import date
from flask import request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from models import ImportFieldConfigWorkProgress, WorkProgress, ConstructionProject
from extensions import db
from routes import admin_construction_progress_import_bp
from routes.business_auth import dual_login_required, visitor_block


def _parse_date(val):
    """将 ISO 日期字符串转为 date 对象"""
    if not val:
        return None
    try:
        return date.fromisoformat(val)
    except (ValueError, TypeError):
        return None


# ============================================================
# 导入字段配置 CRUD
# ============================================================

@admin_construction_progress_import_bp.route('/construction-progress-import/fields', methods=['GET'])
@dual_login_required
def get_import_fields():
    fields = ImportFieldConfigWorkProgress.query.order_by(ImportFieldConfigWorkProgress.sort_order).all()
    return jsonify({'code': 0, 'data': [f.to_dict() for f in fields]})


@admin_construction_progress_import_bp.route('/construction-progress-import/fields', methods=['PUT'])
@dual_login_required
@visitor_block
def update_import_fields():
    data = request.get_json()
    if not data or not isinstance(data, list):
        return jsonify({'code': 1, 'message': '格式错误'}), 400
    for item in data:
        f = ImportFieldConfigWorkProgress.query.get(item.get('id'))
        if f:
            if 'field_label' in item:
                f.field_label = item['field_label']
            if 'is_enabled' in item:
                f.is_enabled = bool(item['is_enabled'])
            if 'is_required' in item:
                f.is_required = bool(item['is_required'])
            if 'sort_order' in item:
                f.sort_order = int(item['sort_order'])
    db.session.commit()
    fields = ImportFieldConfigWorkProgress.query.order_by(ImportFieldConfigWorkProgress.sort_order).all()
    return jsonify({'code': 0, 'data': [f.to_dict() for f in fields], 'message': '已保存'})


# ============================================================
# 下载导入模板
# ============================================================

@admin_construction_progress_import_bp.route('/construction-progress-import/template', methods=['GET'])
@dual_login_required
def download_template():
    fields = ImportFieldConfigWorkProgress.query.filter_by(is_enabled=True)\
        .order_by(ImportFieldConfigWorkProgress.sort_order).all()
    if not fields:
        return jsonify({'code': 1, 'message': '未配置导入模板'}), 400

    # 公共样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='3A7ABD', end_color='3A7ABD', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )
    data_font = Font(name='微软雅黑', size=10)
    data_align = Alignment(vertical='center', wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = '工作进展导入模板'

    # 第1行：表头
    for col_idx, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_idx, value=field.field_label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 第2行：必填/选填标记
    for col_idx, field in enumerate(fields, 1):
        mark = '（必填）' if field.is_required else '（选填）'
        cell = ws.cell(row=2, column=col_idx, value=mark)
        cell.font = Font(name='微软雅黑', size=9, color='C00000' if field.is_required else '909090')

    # 加载在建项目列表（用于所属项目下拉和示例行）
    projects = ConstructionProject.query.filter_by(is_deleted=False)\
        .order_by(ConstructionProject.order_no.asc()).all()
    project_names = [p.project_name for p in projects]

    # 第3行：示例数据（优先使用真实项目名）
    sample_project_name = project_names[0] if project_names else '示例：请先从下拉列表选择项目'
    sample = {
        'project_name': sample_project_name,
        'start_date': '2025-06-01',
        'end_date': '2025-06-15',
        'content': '示例：完成项目立项审批，取得相关批文',
    }
    for col_idx, field in enumerate(fields, 1):
        val = sample.get(field.field_key, '')
        cell = ws.cell(row=3, column=col_idx, value=val)
        cell.font = Font(name='微软雅黑', size=10, color='909090')
        cell.alignment = data_align
        cell.border = thin_border

    # 列宽
    width_map = {
        'project_name': 38,
        'start_date': 14,
        'end_date': 14,
        'content': 48,
    }
    for col_idx, field in enumerate(fields, 1):
        w = width_map.get(field.field_key, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # 找所属项目列索引
    project_name_col = None
    for col_idx, field in enumerate(fields, 1):
        if field.field_key == 'project_name':
            project_name_col = col_idx
            break

    # 下拉数据验证 — 所属项目
    if project_names and project_name_col:
        ws_hidden = wb.create_sheet('_在建项目')
        ws_hidden.sheet_state = 'hidden'
        for i, name in enumerate(project_names, 1):
            ws_hidden.cell(row=i, column=1, value=name)
        dv = DataValidation(
            type='list',
            formula1=f'=_在建项目!$A$1:$A${len(project_names)}',
            allow_blank=True
        )
        dv.error = '请从下拉列表中选择在建项目'
        dv.errorTitle = '无效项目'
        letter = get_column_letter(project_name_col)
        dv.add(f'{letter}3:{letter}1000')
        ws.add_data_validation(dv)

    ws.freeze_panes = 'A4'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='工作进展导入模板.xlsx'
    )


# ============================================================
# 导入预览
# ============================================================

def _get_import_fields():
    return ImportFieldConfigWorkProgress.query.filter_by(is_enabled=True)\
        .order_by(ImportFieldConfigWorkProgress.sort_order).all()


FIELD_LABEL_MAP = {
    'project_name': '所属项目',
    'start_date': '开始日期',
    'end_date': '结束日期',
    'content': '进展内容',
}

_PREVIEW_HEADERS = ['所属项目', '开始日期', '结束日期', '进展内容']


@admin_construction_progress_import_bp.route('/construction-progress-import/preview', methods=['POST'])
@dual_login_required
def import_preview():
    """上传 Excel 并返回预览数据"""
    if 'file' not in request.files:
        return jsonify({'code': 1, 'message': '未找到上传文件'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'code': 1, 'message': '未选择文件'}), 400

    try:
        from openpyxl import load_workbook
        wb = load_workbook(file, read_only=True)
        ws = wb.active
    except Exception:
        return jsonify({'code': 1, 'message': '无法读取文件，请确认是有效的 Excel 文件'}), 400

    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value is not None:
            headers.append(str(cell.value).strip())

    # 获取配置的导入字段
    import_fields = _get_import_fields()
    expected_headers = [f.field_label for f in import_fields]

    # 校验表头是否一致
    if headers != expected_headers:
        return jsonify({
            'code': 1,
            'message': '模板有误，请选择正确的导入模板',
            'data': {'expected': expected_headers, 'actual': headers}
        }), 400

    # 构建 field_label -> field_key 映射
    field_map = {f.field_label: f.field_key for f in import_fields}
    required_fields = {f.field_key for f in import_fields if f.is_required}

    # 预加载项目名称 → id 映射（用于预览校验）
    projects = ConstructionProject.query.filter_by(is_deleted=False).all()
    project_name_to_id = {p.project_name: p.id for p in projects}

    # 读取数据行
    raw_rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or str(v).strip() == '' for v in row):
            continue

        row_data = {}
        for col_idx, val in enumerate(row):
            if col_idx < len(headers):
                field_label = headers[col_idx]
                field_key = field_map.get(field_label, '')
                str_val = str(val).strip() if val is not None else ''
                row_data[field_key] = str_val

        # 跳过标记行、示例行、第3行（模板示例）
        first_val = list(row_data.values())[0] if row_data else ''
        if first_val in ('（必填）', '（选填）'):
            continue
        if first_val and first_val.startswith('示例：'):
            continue
        if row_idx == 3:
            continue

        raw_rows.append({
            'row': row_idx,
            'data': row_data
        })

    # 解析并校验
    preview_rows = []
    all_errors = []

    for raw in raw_rows:
        row_data = raw['data']
        errors = []

        # 校验必填字段
        for field_key in required_fields:
            val = row_data.get(field_key, '')
            if not val:
                label = FIELD_LABEL_MAP.get(field_key, field_key)
                errors.append(f'第{raw["row"]}行：{label} 为必填项')

        # 校验日期格式
        start_date = row_data.get('start_date', '')
        if start_date:
            sd = _parse_date(start_date)
            if sd is None:
                errors.append(f'第{raw["row"]}行：开始日期「{start_date}」格式不正确，应为 YYYY-MM-DD')

        end_date = row_data.get('end_date', '')
        if end_date:
            ed = _parse_date(end_date)
            if ed is None:
                errors.append(f'第{raw["row"]}行：结束日期「{end_date}」格式不正确，应为 YYYY-MM-DD')

        # 校验项目名称是否存在
        project_name = row_data.get('project_name', '')
        if project_name:
            if project_name not in project_name_to_id:
                errors.append(f'第{raw["row"]}行：项目「{project_name}」不存在于在建项目库中')

        all_errors.extend(errors)

        preview_rows.append({
            'row': raw['row'],
            'data': row_data,
            'errors': errors,
            '_valid': len(errors) == 0
        })

    valid_count = sum(1 for r in preview_rows if r['_valid'])
    error_count = len(preview_rows) - valid_count

    return jsonify({
        'code': 0,
        'data': {
            'headers': _PREVIEW_HEADERS,
            'rows': preview_rows,
            'total': len(preview_rows),
            'valid_count': valid_count,
            'error_count': error_count,
            'all_errors': all_errors,
        }
    })


# ============================================================
# 执行导入
# ============================================================

@admin_construction_progress_import_bp.route('/construction-progress-import/execute', methods=['POST'])
@dual_login_required
@visitor_block
def import_execute():
    """将预览通过的数据写入数据库"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 1, 'message': '数据为空'}), 400

    rows = data.get('rows', [])
    if not rows:
        return jsonify({'code': 1, 'message': '无数据可导入'}), 400

    valid_rows = [r for r in rows if r.get('_valid')]
    if not valid_rows:
        return jsonify({'code': 1, 'message': '没有有效数据行'}), 400

    # 预加载项目名称 → id 映射
    projects = ConstructionProject.query.filter_by(is_deleted=False).all()
    project_name_to_id = {p.project_name: p.id for p in projects}

    imported = 0
    skipped = 0
    for row in valid_rows:
        row_data = row['data']

        project_name = row_data.get('project_name', '')
        project_id = project_name_to_id.get(project_name)
        if not project_id:
            skipped += 1
            continue

        start_date = _parse_date(row_data.get('start_date'))
        end_date = _parse_date(row_data.get('end_date'))
        content = row_data.get('content', '')

        if not content or not start_date or not end_date:
            skipped += 1
            continue

        wp = WorkProgress(
            project_id=project_id,
            start_date=start_date,
            end_date=end_date,
            content=content,
        )
        db.session.add(wp)
        imported += 1

    db.session.commit()

    return jsonify({
        'code': 0,
        'message': f'成功导入 {imported} 条工作进展' + (f'，跳过 {skipped} 条' if skipped > 0 else ''),
        'data': {'count': imported, 'skipped': skipped}
    })
