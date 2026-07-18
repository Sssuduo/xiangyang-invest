"""在建项目工作进展「更新导入」— 模板 / 预览 / 执行

与 admin_construction_progress_import.py 区分：
- 模板包含项目编号 + 项目名称 + 内容（三列固定，不走字段配置表）
- 时间段由前端弹窗统一选择，应用于所有导入行
- 内容自动处理"上周：... 本周：..."格式，只保留本周内容
- 导入时记录导入人 user_id + display_name
"""
import io
import re
from datetime import date
from flask import request, jsonify, send_file, session, current_app
from flask_login import current_user
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from models import WorkProgress, ConstructionProject, BusinessUser
from extensions import db
from routes import admin_construction_progress_update_import_bp
from routes.business_auth import dual_login_required, visitor_block


def _parse_date(val):
    """将日期字符串转为 date 对象"""
    if not val:
        return None
    try:
        return date.fromisoformat(val)
    except (ValueError, TypeError):
        return None


def _process_content(raw_content):
    """如果格式为 上周：... 本周：...，只保留本周内容；否则保留全部"""
    if not raw_content:
        return ''
    text = raw_content.strip()
    # 匹配 "上周：xxx 本周：yyy" 或 "上周: xxx 本周: yyy"
    m = re.search(r'上周[：:](.*?)本周[：:](.*)', text, re.DOTALL)
    if m:
        return m.group(2).strip()
    # 也匹配 "本周：..." 单独出现的情况
    m2 = re.search(r'本周[：:](.*)', text, re.DOTALL)
    if m2:
        return m2.group(1).strip()
    return text


def _get_current_user_info():
    """获取当前登录用户信息（兼容 AdminUser 和 BusinessUser）"""
    if current_user.is_authenticated:
        return current_user.id, current_user.display_name or current_user.username
    user_id = session.get('business_user_id')
    if user_id:
        user = BusinessUser.query.get(int(user_id))
        if user:
            return user.id, user.display_name or user.username
    return None, None


# ============================================================
# 下载更新导入模板
# ============================================================

@admin_construction_progress_update_import_bp.route(
    '/construction-progress-update-import/template', methods=['GET']
)
@dual_login_required
def download_template():
    """下载更新导入模板（固定3列：项目编号、项目名称、内容）"""
    # 加载所有在建项目用于下拉验证和示例
    projects = ConstructionProject.query.filter_by(is_deleted=False)\
        .order_by(ConstructionProject.order_no.asc()).all()
    project_names = [p.project_name for p in projects]

    # 公共样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='3A7ABD', end_color='3A7ABD', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )
    data_align = Alignment(vertical='center', wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = '工作进展更新导入模板'

    # 固定三列
    headers = ['项目编号', '项目名称', '进展内容']
    widths = [14, 38, 68]

    # 第1行：表头
    for col_idx, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 第2行：必填标记
    for col_idx in range(1, 4):
        cell = ws.cell(row=2, column=col_idx, value='（必填）')
        cell.font = Font(name='微软雅黑', size=9, color='C00000')

    # 第3行及之后：预填所有在建项目（编号+名称，时间段和内容留空供填写）
    if projects:
        for i, proj in enumerate(projects):
            row_num = 3 + i
            data_row = [
                str(proj.order_no),
                proj.project_name,
                '',
            ]
            for col_idx, val in enumerate(data_row, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = Font(name='微软雅黑', size=10)
                cell.alignment = data_align
                cell.border = thin_border
    else:
        sample_row = [
            '1',
            '示例：请从下拉列表选择项目',
            '本周：完成项目方案评审，通过专家论证；上周：准备评审材料，协调专家时间',
        ]
        for col_idx, val in enumerate(sample_row, 1):
            cell = ws.cell(row=3, column=col_idx, value=val)
            cell.font = Font(name='微软雅黑', size=10, color='909090')
            cell.alignment = data_align
            cell.border = thin_border

    # 列宽
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # 项目名称列（B列）下拉验证
    if project_names:
        ws_hidden = wb.create_sheet('_在建项目')
        ws_hidden.sheet_state = 'hidden'
        for i, name in enumerate(project_names, 1):
            ws_hidden.cell(row=i, column=1, value=name)
        dv = DataValidation(
            type='list',
            formula1=f'=_在建项目!$A$1:$A${len(project_names)}',
            allow_blank=True
        )
        dv.error = '请从下拉列表中选择在建项目'
        dv.errorTitle = '无效项目'
        dv.add('B3:B2000')
        ws.add_data_validation(dv)

    # 冻结表头（冻结前2行）
    ws.freeze_panes = 'A3'

    # 添加说明 sheet
    ws_info = wb.create_sheet('填写说明')
    ws_info.cell(row=1, column=1, value='填写说明').font = Font(name='微软雅黑', size=14, bold=True)
    notes = [
        '1. 项目编号：在建项目库中的 order_no（序号），可在项目中确认',
        '2. 项目名称：必须从下拉列表中选择，确保与系统中的项目名称完全一致',
        '3. 进展内容：',
        '   - 如果格式为"上周：xxx 本周：yyy"，系统会自动只保留"本周"的内容',
        '   - 如果没有"上周/本周"标记，则保留全部内容',
        '4. 时间段在导入弹窗中统一选择，无需在模板中填写',
        '5. 所有列为必填，请勿留空',
        '6. 模板已自动填入所有在建项目的编号和名称，请在对应行填写内容',
        '7. 不需要的行可以整行删除或留空（留空的行导入时自动跳过）',
    ]
    for i, note in enumerate(notes, 3):
        ws_info.cell(row=i, column=1, value=note).font = Font(name='微软雅黑', size=11)
    ws_info.column_dimensions['A'].width = 80

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='工作进展更新导入模板.xlsx'
    )


# ============================================================
# 导入预览
# ============================================================

PREVIEW_HEADERS = ['项目编号', '项目名称', '开始日期', '结束日期', '原始内容', '处理后内容（本周）', '备注']


@admin_construction_progress_update_import_bp.route(
    '/construction-progress-update-import/preview', methods=['POST']
)
@dual_login_required
def import_preview():
    """上传 Excel 并返回预览数据（含内容截取效果展示）。
    日期范围由前端弹窗统一选择，通过 FormData 传入应用于所有行。"""
    if 'file' not in request.files:
        return jsonify({'code': 1, 'message': '未找到上传文件'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'code': 1, 'message': '未选择文件'}), 400

    # 从前端获取全局日期范围
    start_date_str = request.form.get('start_date', '').strip()
    end_date_str = request.form.get('end_date', '').strip()
    if not start_date_str or not end_date_str:
        return jsonify({'code': 1, 'message': '请选择工作进展的时间段（开始日期和结束日期）'}), 400
    start_date = _parse_date(start_date_str)
    end_date = _parse_date(end_date_str)
    if not start_date or not end_date:
        return jsonify({'code': 1, 'message': '日期格式不正确'}), 400

    try:
        from openpyxl import load_workbook
        wb = load_workbook(file, read_only=True)
        ws = wb.active
    except Exception:
        return jsonify({'code': 1, 'message': '无法读取文件，请确认是有效的 Excel 文件'}), 400

    # 读取表头
    raw_headers = []
    for cell in ws[1]:
        if cell.value is not None:
            raw_headers.append(str(cell.value).strip())

    expected = ['项目编号', '项目名称', '进展内容']
    if raw_headers != expected:
        return jsonify({
            'code': 1,
            'message': '模板有误，请选择正确的更新导入模板',
            'data': {'expected': expected, 'actual': raw_headers}
        }), 400

    # 预加载项目数据（编号+名称双重校验）
    projects = ConstructionProject.query.filter_by(is_deleted=False).all()
    project_map = {}  # {project_name: ConstructionProject}
    order_no_to_project = {}  # {order_no: ConstructionProject}
    for p in projects:
        project_map[p.project_name] = p
        order_no_to_project[p.order_no] = p

    # 读取数据行
    raw_rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or str(v).strip() == '' for v in row):
            continue

        vals = [str(v).strip() if v is not None else '' for v in row[:3]]

        # 跳过标记行
        first_val = vals[0] if vals else ''
        if first_val in ('（必填）', '（选填）'):
            continue
        if first_val and first_val.startswith('示例：'):
            continue

        # 跳过内容为空的行（模板中预填了项目信息但用户未填写进展）
        if not vals[2]:
            continue

        raw_rows.append({
            'row': row_idx,
            'data': {
                'order_no': vals[0],
                'project_name': vals[1],
                'content': vals[2],
            }
        })

    # 解析并校验
    preview_rows = []
    all_errors = []

    for raw in raw_rows:
        row_data = raw['data']
        errors = []

        order_no_str = row_data.get('order_no', '')
        project_name = row_data.get('project_name', '')
        raw_content = row_data.get('content', '')

        # 校验必填
        if not order_no_str:
            errors.append(f'第{raw["row"]}行：项目编号 为必填项')
        if not project_name:
            errors.append(f'第{raw["row"]}行：项目名称 为必填项')
        if not raw_content:
            errors.append(f'第{raw["row"]}行：进展内容 为必填项')

        # 校验项目编号
        order_no = None
        if order_no_str:
            try:
                order_no = int(order_no_str)
            except ValueError:
                errors.append(f'第{raw["row"]}行：项目编号「{order_no_str}」不是有效数字')

        # 校验项目名称存在
        project = None
        if project_name:
            project = project_map.get(project_name)
            if not project:
                errors.append(f'第{raw["row"]}行：项目「{project_name}」不存在于在建项目库中')

        # 校验编号与名称一致性
        if order_no is not None and project is not None:
            matched = order_no_to_project.get(order_no)
            if matched is None:
                errors.append(f'第{raw["row"]}行：项目编号 {order_no} 不存在')
            elif matched.project_name != project_name:
                errors.append(
                    f'第{raw["row"]}行：编号 {order_no} 对应项目「{matched.project_name}」，'
                    f'与填写的「{project_name}」不一致'
                )

        # 处理内容
        processed_content = _process_content(raw_content)
        content_note = ''
        if processed_content != raw_content.strip():
            content_note = '已截取"本周"内容'

        all_errors.extend(errors)

        preview_rows.append({
            'row': raw['row'],
            'data': {
                'order_no': order_no_str,
                'project_name': project_name,
                'start_date': start_date.isoformat() if start_date else start_date_str,
                'end_date': end_date.isoformat() if end_date else end_date_str,
                'content': raw_content,
                'processed_content': processed_content,
            },
            'content_note': content_note,
            'has_content_change': processed_content != raw_content.strip(),
            'errors': errors,
            '_valid': len(errors) == 0
        })

    valid_count = sum(1 for r in preview_rows if r['_valid'])
    error_count = len(preview_rows) - valid_count

    return jsonify({
        'code': 0,
        'data': {
            'headers': PREVIEW_HEADERS,
            'rows': preview_rows,
            'total': len(preview_rows),
            'valid_count': valid_count,
            'error_count': error_count,
            'all_errors': all_errors,
        }
    })


# ============================================================
# 执行导入
# ============================================================

@admin_construction_progress_update_import_bp.route(
    '/construction-progress-update-import/execute', methods=['POST']
)
@dual_login_required
@visitor_block
def import_execute():
    """将预览通过的数据写入数据库，记录导入人信息"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 1, 'message': '数据为空'}), 400

    rows = data.get('rows', [])
    if not rows:
        return jsonify({'code': 1, 'message': '无数据可导入'}), 400

    # 全局日期（由前端传入，作为行级日期解析失败时的回退）
    global_start_date = _parse_date(data.get('start_date', ''))
    global_end_date = _parse_date(data.get('end_date', ''))

    valid_rows = [r for r in rows if r.get('_valid')]
    if not valid_rows:
        return jsonify({'code': 1, 'message': '没有有效数据行'}), 400

    # 获取当前登录用户信息
    import_user_id, import_user_name = _get_current_user_info()

    # 预加载项目名称 → id 映射
    projects = ConstructionProject.query.filter_by(is_deleted=False).all()
    project_name_to_id = {p.project_name: p.id for p in projects}

    imported = 0
    skipped = 0
    try:
        for row in valid_rows:
            row_data = row['data']

            project_name = row_data.get('project_name', '')
            project_id = project_name_to_id.get(project_name)
            if not project_id:
                skipped += 1
                continue

            start_date = _parse_date(row_data.get('start_date')) or global_start_date
            end_date = _parse_date(row_data.get('end_date')) or global_end_date

            # 内容处理：优先使用 processed_content，其次再处理一次
            content = row_data.get('processed_content', '')
            if not content:
                content = _process_content(row_data.get('content', ''))

            if not content or not start_date or not end_date:
                skipped += 1
                continue

            wp = WorkProgress(
                project_id=project_id,
                start_date=start_date,
                end_date=end_date,
                content=content,
                import_user_id=import_user_id,
                import_user_name=import_user_name or '',
            )
            db.session.add(wp)
            imported += 1
            if imported % 50 == 0:
                db.session.flush()  # 分批 flush，释放写锁给其他写者

        db.session.commit()
    except Exception as e:
        db.session.rollback()
        import traceback
        current_app.logger.error(f'工作进展更新导入失败: {e}\n{traceback.format_exc()}')
        return jsonify({
            'code': 1,
            'message': f'导入失败（数据库写入错误）：{str(e)}'
        }), 500

    msg = f'成功导入 {imported} 条工作进展'
    if import_user_name:
        msg += f'（导入人：{import_user_name}）'
    if skipped > 0:
        msg += f'，跳过 {skipped} 条'

    return jsonify({
        'code': 0,
        'message': msg,
        'data': {
            'count': imported,
            'skipped': skipped,
            'import_user_id': import_user_id,
            'import_user_name': import_user_name or '',
        }
    })
