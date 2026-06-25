import io
import json
from datetime import datetime, timedelta
from flask import request, jsonify, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from models import EnterpriseDemand, InvestmentProject, DemandTypeDict, OrganizationDict, FollowStatusDict
from models import ImportFieldConfigDemand
from extensions import db
from routes import admin_demand_bp
from routes.business_auth import dual_login_required
from utils import get_current_user_info, log_changes


# ============================================================
# 企业诉求 CRUD
# ============================================================

@admin_demand_bp.route('/demand/demands', methods=['GET'])
@dual_login_required
def list_demands():
    """诉求列表（含搜索/筛选/项目名称）"""
    search = request.args.get('search', '').strip()
    project_id = request.args.get('project_id', '').strip()
    demand_type = request.args.get('demand_type', '').strip()
    project_type = request.args.get('project_type', '').strip()
    status = request.args.get('status', '').strip()
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)

    q = EnterpriseDemand.query.join(InvestmentProject)

    if search:
        like = f'%{search}%'
        q = q.filter(db.or_(
            InvestmentProject.project_name.ilike(like),
            EnterpriseDemand.demand_content.ilike(like)
        ))

    if project_id:
        q = q.filter(EnterpriseDemand.project_id == int(project_id))
    if project_type:
        q = q.filter(InvestmentProject.project_type_code == project_type)
    if demand_type:
        codes = [c.strip() for c in demand_type.split(',') if c.strip()]
        if len(codes) == 1:
            q = q.filter(EnterpriseDemand.demand_type_code == codes[0])
        else:
            q = q.filter(EnterpriseDemand.demand_type_code.in_(codes))
    if status:
        q = q.filter(EnterpriseDemand.status == status)

    total = q.count()
    q = q.order_by(EnterpriseDemand.created_at.desc())
    demands = q.offset((page - 1) * page_size).limit(page_size).all()

    # 解析字典名称（二级显示：一级：二级）
    type_map = DemandTypeDict.build_display_name_map()
    org_map = {d.code: d.name for d in OrganizationDict.query.all()}

    result = []
    for d in demands:
        item = d.to_dict()
        item['project_name'] = d.project.project_name if d.project else ''
        # 支持逗号分隔的多值编码
        codes = [c.strip() for c in (d.demand_type_code or '').split(',') if c.strip()]
        item['demand_type_name'] = '、'.join([type_map.get(c, c) for c in codes]) if codes else ''
        item['unit_name'] = org_map.get(d.unit_code, '')
        result.append(item)

    return jsonify({'code': 0, 'data': result, 'total': total})


@admin_demand_bp.route('/demand/demands', methods=['POST'])
@dual_login_required
def create_demand():
    """新建诉求"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 1, 'message': '请求数据不能为空'}), 400

    if not data.get('demand_content'):
        return jsonify({'code': 1, 'message': '诉求内容为必填字段'}), 400
    if not data.get('project_id'):
        return jsonify({'code': 1, 'message': '所属项目为必填字段'}), 400

    max_order = db.session.query(db.func.max(EnterpriseDemand.sort_order))\
        .filter_by(project_id=int(data['project_id'])).scalar() or 0

    demand = EnterpriseDemand(
        project_id=int(data['project_id']),
        demand_type_code=data.get('demand_type_code', ''),
        demand_content=data['demand_content'],
        resolution=data.get('resolution', ''),
        unit_code=data.get('unit_code', ''),
        status=data.get('status', 'pending'),
        sort_order=max_order + 1
    )
    db.session.add(demand)
    db.session.flush()

    # 审计日志
    user_info = get_current_user_info()
    if user_info:
        changes = {
            'project_id': (None, demand.project_id),
            'demand_type_code': (None, demand.demand_type_code or ''),
            'demand_content': (None, demand.demand_content),
            'resolution': (None, demand.resolution or ''),
            'unit_code': (None, demand.unit_code or ''),
            'status': (None, demand.status),
            'sort_order': (None, demand.sort_order)
        }
        log_changes('enterprise_demands', demand.id, changes, 'create', user_info)

    db.session.commit()
    return jsonify({'code': 0, 'data': demand.to_dict(), 'message': '诉求已创建'})


@admin_demand_bp.route('/demand/demands/<int:demand_id>', methods=['GET'])
@dual_login_required
def get_demand(demand_id):
    """获取诉求详情"""
    demand = EnterpriseDemand.query.filter_by(id=demand_id).first_or_404()
    data = demand.to_dict()
    data['project_name'] = demand.project.project_name if demand.project else ''
    return jsonify({'code': 0, 'data': data})


@admin_demand_bp.route('/demand/demands/<int:demand_id>', methods=['PUT'])
@dual_login_required
def update_demand(demand_id):
    """更新诉求"""
    demand = EnterpriseDemand.query.filter_by(id=demand_id).first_or_404()
    data = request.get_json()
    if not data:
        return jsonify({'code': 1, 'message': '请求数据不能为空'}), 400

    updatable = ['project_id', 'demand_type_code', 'demand_content', 'resolution', 'unit_code', 'status', 'sort_order']

    # 审计日志：保存旧值
    user_info = get_current_user_info()
    old_values = {}
    if user_info:
        for field in updatable:
            old_values[field] = getattr(demand, field) or ''

    for field in updatable:
        if field in data:
            setattr(demand, field, data[field])

    # 审计日志：对比变更
    if user_info:
        changes = {}
        for field in updatable:
            old_val = old_values.get(field, '')
            new_val = getattr(demand, field) or ''
            if str(old_val) != str(new_val):
                changes[field] = (old_val, new_val)
        log_changes('enterprise_demands', demand_id, changes, 'update', user_info)

    db.session.commit()
    return jsonify({'code': 0, 'data': demand.to_dict(), 'message': '诉求已更新'})


@admin_demand_bp.route('/demand/demands/<int:demand_id>', methods=['DELETE'])
@dual_login_required
def delete_demand(demand_id):
    """删除诉求"""
    demand = EnterpriseDemand.query.filter_by(id=demand_id).first_or_404()
    db.session.delete(demand)
    db.session.commit()
    return jsonify({'code': 0, 'message': '诉求已删除'})


@admin_demand_bp.route('/demand/demands/batch-delete', methods=['POST'])
@dual_login_required
def batch_delete_demands():
    """批量删除诉求"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 1, 'message': '请求数据不能为空'}), 400

    ids = data.get('ids', [])
    if not ids or not isinstance(ids, list):
        return jsonify({'code': 1, 'message': '请提供要删除的诉求ID列表'}), 400

    deleted = EnterpriseDemand.query.filter(EnterpriseDemand.id.in_(ids)).delete(synchronize_session=False)
    db.session.commit()

    return jsonify({'code': 0, 'message': f'成功删除 {deleted} 条诉求', 'data': {'count': deleted}})


# ============================================================
# 导入字段配置
# ============================================================

@admin_demand_bp.route('/demand-import/fields', methods=['GET'])
@dual_login_required
def get_import_fields():
    fields = ImportFieldConfigDemand.query.order_by(ImportFieldConfigDemand.sort_order).all()
    return jsonify({'code': 0, 'data': [f.to_dict() for f in fields]})


@admin_demand_bp.route('/demand-import/fields', methods=['PUT'])
@dual_login_required
def update_import_fields():
    data = request.get_json()
    if not data or not isinstance(data, list):
        return jsonify({'code': 1, 'message': '格式错误'}), 400
    for item in data:
        f = ImportFieldConfigDemand.query.get(item.get('id'))
        if f:
            if 'field_label' in item:
                f.field_label = item['field_label']
            if 'is_enabled' in item:
                f.is_enabled = bool(item['is_enabled'])
            if 'is_required' in item:
                f.is_required = bool(item['is_required'])
            if 'sort_order' in item:
                f.sort_order = int(item['sort_order'])
    db.session.commit()
    fields = ImportFieldConfigDemand.query.order_by(ImportFieldConfigDemand.sort_order).all()
    return jsonify({'code': 0, 'data': [f.to_dict() for f in fields], 'message': '已保存'})


# ============================================================
# 项目列表（供下载模板选择）
# ============================================================

@admin_demand_bp.route('/demand-import/projects-for-template', methods=['GET'])
@dual_login_required
def projects_for_template():
    """返回可用于模板的项目列表，支持跟进状态筛选"""
    follow_status = request.args.get('follow_status', '').strip()

    q = InvestmentProject.query.filter_by(is_deleted=False)
    if follow_status:
        q = q.filter_by(follow_status_code=follow_status)

    projects = q.order_by(InvestmentProject.order_no.asc()).all()

    follow_map = {d.code: d.name for d in FollowStatusDict.query.all()}

    result = []
    for p in projects:
        result.append({
            'id': p.id,
            'project_name': p.project_name,
            'follow_status_code': p.follow_status_code,
            'follow_status_name': follow_map.get(p.follow_status_code, ''),
            'invest_enterprise': p.invest_enterprise,
        })

    return jsonify({'code': 0, 'data': result})


# ============================================================
# 下载导入模板
# ============================================================

@admin_demand_bp.route('/demand-import/template', methods=['GET'])
@dual_login_required
def download_template():
    """
    下载导入模板。
    可选参数 project_ids=1,2,3 用于预填选中的项目ID和项目名称到模板，
    同时为项目名称列添加下拉数据验证。
    """
    fields = ImportFieldConfigDemand.query.filter_by(is_enabled=True)\
        .order_by(ImportFieldConfigDemand.sort_order).all()
    if not fields:
        return jsonify({'code': 1, 'message': '未配置导入模板'}), 400

    project_ids_str = request.args.get('project_ids', '')
    project_ids = []
    if project_ids_str:
        try:
            project_ids = [int(x) for x in project_ids_str.split(',') if x.strip().isdigit()]
        except (ValueError, TypeError):
            project_ids = []

    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='3A7ABD', end_color='3A7ABD', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )
    data_font = Font(name='微软雅黑', size=10)
    data_align = Alignment(vertical='center', wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = '企业诉求导入模板'

    # 第1行：表头
    for col_idx, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_idx, value=field.field_label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 列宽
    widths = {'project_id': 10, 'project_name': 26, 'demand_type_code': 16, 'demand_content': 50, 'unit_code': 16, 'status': 12}
    for col_idx, field in enumerate(fields, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(field.field_key, 18)

    # 找到 project_id 和 project_name 的列索引
    id_col = None
    name_col = None
    for col_idx, field in enumerate(fields, 1):
        if field.field_key == 'project_id':
            id_col = col_idx
        elif field.field_key == 'project_name':
            name_col = col_idx

    if project_ids:
        # ========== 有项目列表：预填项目ID + 项目名称 ==========
        selected_projects = InvestmentProject.query.filter(
            InvestmentProject.id.in_(project_ids),
            InvestmentProject.is_deleted == False
        ).order_by(InvestmentProject.order_no.asc()).all()

        all_projects = InvestmentProject.query.filter_by(is_deleted=False)\
            .order_by(InvestmentProject.project_name).all()
        all_project_names = [p.project_name for p in all_projects]

        # 第2行：必填/选填标记
        for col_idx, field in enumerate(fields, 1):
            mark = '（必填）' if field.is_required else '（选填）'
            cell = ws.cell(row=2, column=col_idx, value=mark)
            cell.font = Font(name='微软雅黑', size=9, color='C00000' if field.is_required else '909090')

        # 第3行起：写入选中的项目ID和项目名称
        for i, proj in enumerate(selected_projects):
            row = 3 + i
            if id_col:
                cell = ws.cell(row=row, column=id_col, value=proj.id)
                cell.font = data_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            if name_col:
                cell = ws.cell(row=row, column=name_col, value=proj.project_name)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border
            for col_idx in range(1, len(fields) + 1):
                if col_idx not in (id_col, name_col):
                    cell = ws.cell(row=row, column=col_idx, value='')
                    cell.font = data_font
                    cell.alignment = data_align
                    cell.border = thin_border

        # 下拉数据验证（项目名称列）
        if all_project_names and name_col:
            ws_hidden = wb.create_sheet('项目列表')
            ws_hidden.sheet_state = 'hidden'
            for i, name in enumerate(all_project_names, 1):
                ws_hidden.cell(row=i, column=1, value=name)

            dv = DataValidation(
                type='list',
                formula1=f'=项目列表!$A$1:$A${len(all_project_names)}',
                allow_blank=True
            )
            dv.error = '请从下拉列表中选择项目名称'
            dv.errorTitle = '无效项目'
            dv.prompt = '请选择'
            dv.promptTitle = '所属项目'
            last_row = max(3 + len(selected_projects), 500)
            name_letter = get_column_letter(name_col)
            dv.add(f'{name_letter}3:{name_letter}{last_row}')
            ws.add_data_validation(dv)

        ws.freeze_panes = 'A3'

    else:
        # ========== 无项目列表：空模板 ==========
        sample = {
            'project_id': '',
            'project_name': '示例项目名称',
            'demand_type_code': '用地需求',
            'demand_content': '示例诉求内容',
            'unit_code': '示例对接单位',
            'status': '待处理',
        }
        for col_idx, field in enumerate(fields, 1):
            val = sample.get(field.field_key, '')
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.font = Font(name='微软雅黑', size=10, color='909090')
            cell.border = thin_border

        # 第3行：必填标记
        for col_idx, field in enumerate(fields, 1):
            mark = '（必填）' if field.is_required else '（选填）'
            cell = ws.cell(row=3, column=col_idx, value=mark)
            cell.font = Font(name='微软雅黑', size=9, color='C00000' if field.is_required else '909090')

        ws.freeze_panes = 'A4'

    # ========== 字典下拉数据验证（诉求类型 / 对接单位 / 状态）==========
    # 找到各字典列的索引
    demand_type_col = None
    unit_col = None
    status_col = None
    for col_idx, field in enumerate(fields, 1):
        if field.field_key == 'demand_type_code':
            demand_type_col = col_idx
        elif field.field_key == 'unit_code':
            unit_col = col_idx
        elif field.field_key == 'status':
            status_col = col_idx

    # 确定数据起始行
    data_start_row = 3 if project_ids else 4

    # 诉求类型下拉选项
    if demand_type_col:
        demand_types = DemandTypeDict.query.filter_by(is_active=True)\
            .order_by(DemandTypeDict.sort_order).all()
        demand_type_names = [d.name for d in demand_types]
        if demand_type_names:
            dv = DataValidation(
                type='list',
                formula1='"' + ','.join(demand_type_names) + '"',
                allow_blank=True
            )
            dv.error = '请从下拉列表中选择诉求类型'
            dv.errorTitle = '无效类型'
            col_letter = get_column_letter(demand_type_col)
            dv.add(f'{col_letter}{data_start_row}:{col_letter}1000')
            ws.add_data_validation(dv)

    # 对接单位下拉选项
    if unit_col:
        orgs = OrganizationDict.query.filter_by(is_active=True)\
            .order_by(OrganizationDict.sort_order).all()
        org_names = [o.name for o in orgs]
        if org_names:
            dv = DataValidation(
                type='list',
                formula1='"' + ','.join(org_names) + '"',
                allow_blank=True
            )
            dv.error = '请从下拉列表中选择对接单位'
            dv.errorTitle = '无效单位'
            col_letter = get_column_letter(unit_col)
            dv.add(f'{col_letter}{data_start_row}:{col_letter}1000')
            ws.add_data_validation(dv)

    # 状态下拉选项
    if status_col:
        status_names = ['待处理', '处理中', '已解决']
        dv = DataValidation(
            type='list',
            formula1='"' + ','.join(status_names) + '"',
            allow_blank=True
        )
        dv.error = '请从下拉列表中选择状态'
        dv.errorTitle = '无效状态'
        col_letter = get_column_letter(status_col)
        dv.add(f'{col_letter}{data_start_row}:{col_letter}1000')
        ws.add_data_validation(dv)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='企业诉求导入模板.xlsx'
    )


# ============================================================
# 导入预览
# ============================================================

def _get_demand_import_fields():
    return ImportFieldConfigDemand.query.filter_by(is_enabled=True)\
        .order_by(ImportFieldConfigDemand.sort_order).all()


@admin_demand_bp.route('/demand-import/preview', methods=['POST'])
@dual_login_required
def import_preview():
    """上传 Excel 并返回预览数据"""
    if 'file' not in request.files:
        return jsonify({'code': 1, 'message': '未找到上传文件'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'code': 1, 'message': '未选择文件'}), 400

    try:
        wb = load_workbook(file, read_only=True)
        ws = wb.active
    except Exception:
        return jsonify({'code': 1, 'message': '无法读取文件，请确认是有效的 Excel 文件'}), 400

    headers = []
    for cell in ws[1]:
        if cell.value is not None:
            headers.append(str(cell.value).strip())

    import_fields = _get_demand_import_fields()
    expected_headers = [f.field_label for f in import_fields]

    if headers != expected_headers:
        return jsonify({
            'code': 1,
            'message': '模板有误，请选择正确的导入模板',
            'data': {'expected': expected_headers, 'actual': headers}
        }), 400

    field_map = {f.field_label: f.field_key for f in import_fields}
    all_projects = {p.project_name: p.id for p in InvestmentProject.query.filter_by(is_deleted=False).all()}
    all_projects_by_id = {p.id: p for p in InvestmentProject.query.filter_by(is_deleted=False).all()}

    rows = []
    all_errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or str(v).strip() == '' for v in row):
            continue

        row_data = {}
        for col_idx, val in enumerate(row):
            if col_idx < len(headers):
                field_key = field_map.get(headers[col_idx], '')
                str_val = str(val).strip() if val is not None else ''
                row_data[field_key] = str_val

        first_val = list(row_data.values())[0] if row_data else ''
        if first_val in ('（必填）', '（选填）'):
            continue

        errors = []
        # 校验必填
        for f in import_fields:
            val = row_data.get(f.field_key, '')
            if f.is_required and (val is None or str(val).strip() == ''):
                errors.append(f'第{row_idx}行：{f.field_label} 为必填项')

        # 校验项目存在：优先通过项目ID，回退到项目名称
        pid_str = row_data.get('project_id', '').strip()
        pname = row_data.get('project_name', '').strip()
        if pid_str and pid_str.isdigit():
            pid = int(pid_str)
            if pid not in all_projects_by_id or all_projects_by_id[pid].is_deleted:
                errors.append(f'第{row_idx}行：项目ID「{pid}」在数据库中不存在')
        elif pname and pname not in all_projects:
            errors.append(f'第{row_idx}行：项目「{pname}」在数据库中不存在')
        elif not pid_str and not pname:
            # 如果 project_id 是必填的，空值已在上面被捕获
            pass

        rows.append({
            'row': row_idx,
            'data': row_data,
            'errors': errors,
            '_valid': len(errors) == 0
        })
        all_errors.extend(errors)

    valid_count = sum(1 for r in rows if r['_valid'])
    error_count = len(rows) - valid_count

    return jsonify({
        'code': 0,
        'data': {
            'headers': expected_headers,
            'rows': rows,
            'total': len(rows),
            'valid_count': valid_count,
            'error_count': error_count,
            'all_errors': all_errors
        }
    })


# ============================================================
# 执行导入
# ============================================================

@admin_demand_bp.route('/demand-import/execute', methods=['POST'])
@dual_login_required
def import_execute():
    """将预览通过的数据写入数据库"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 1, 'message': '数据为空'}), 400

    rows = data.get('rows', [])
    if not rows:
        return jsonify({'code': 1, 'message': '无数据可导入'}), 400

    valid_rows = [r for r in rows if r.get('_valid')]
    if not valid_rows:
        return jsonify({'code': 1, 'message': '没有有效数据行'}), 400

    project_map = {p.project_name: p.id for p in InvestmentProject.query.filter_by(is_deleted=False).all()}

    imported = 0
    for row in valid_rows:
        row_data = row['data']

        # 优先通过项目ID直接获取（100%准确对应）
        pid_str = str(row_data.get('project_id', '') or '').strip()
        if pid_str and pid_str.isdigit():
            project_id = int(pid_str)
            proj = InvestmentProject.query.get(project_id)
            if not proj or proj.is_deleted:
                project_id = None
        else:
            # 回退：通过项目名称查找
            project_name = row_data.get('project_name', '').strip()
            project_id = project_map.get(project_name)

        if not project_id:
            continue

        max_order = db.session.query(db.func.max(EnterpriseDemand.sort_order))\
            .filter_by(project_id=project_id).scalar() or 0

        demand = EnterpriseDemand(
            project_id=project_id,
            demand_type_code=row_data.get('demand_type_code', ''),
            demand_content=row_data.get('demand_content', ''),
            resolution=row_data.get('resolution', ''),
            unit_code=row_data.get('unit_code', ''),
            status=row_data.get('status', 'pending'),
            sort_order=max_order + 1
        )
        db.session.add(demand)
        imported += 1

    db.session.commit()
    return jsonify({'code': 0, 'message': f'成功导入 {imported} 条诉求', 'data': {'count': imported}})


# ============================================================
# 数据看板统计
# ============================================================

@admin_demand_bp.route('/demand-stats', methods=['GET'])
@dual_login_required
def demand_stats():
    """企业诉求数据统计（数据看板）"""
    from sqlalchemy import func

    # 项目类型筛选参数
    project_type = request.args.get('project_type', '').strip()

    # 构建诉求查询基类（可选按项目类型过滤，使用子查询避免join冲突）
    def demand_query():
        q = EnterpriseDemand.query
        if project_type:
            proj_ids = db.session.query(InvestmentProject.id).filter(
                InvestmentProject.project_type_code == project_type
            ).subquery()
            q = q.filter(EnterpriseDemand.project_id.in_(proj_ids))
        return q

    # ---- 总体概览 ----
    total_demands = demand_query().count()
    pending_count = demand_query().filter(EnterpriseDemand.status == 'pending').count()
    processing_count = demand_query().filter(EnterpriseDemand.status == 'processing').count()
    resolved_count = demand_query().filter(EnterpriseDemand.status == 'resolved').count()
    projects_with_demands = db.session.query(func.count(func.distinct(
        demand_query().with_entities(EnterpriseDemand.project_id).subquery().c.project_id
    ))).scalar() or 0
    has_resolution = demand_query().filter(
        EnterpriseDemand.resolution.isnot(None),
        EnterpriseDemand.resolution != ''
    ).count()
    resolution_rate = round(has_resolution / total_demands * 100, 1) if total_demands > 0 else 0

    overview = {
        'total_demands': total_demands,
        'pending_count': pending_count,
        'processing_count': processing_count,
        'resolved_count': resolved_count,
        'projects_with_demands': projects_with_demands,
        'resolution_rate': resolution_rate
    }

    # ---- 按诉求类型分布（含各状态下数量 + 二级下钻 + 关联项目）----
    type_map = DemandTypeDict.build_display_name_map()
    # 获取所有字典项（含 parent_code）
    all_type_dicts = DemandTypeDict.query.all()
    type_parent_map = {d.code: (d.parent_code or '') for d in all_type_dicts}
    type_name_simple = {d.code: d.name for d in all_type_dicts}

    type_rows = demand_query().with_entities(
        EnterpriseDemand.demand_type_code,
        EnterpriseDemand.status,
        func.count(EnterpriseDemand.id)
    ).group_by(
        EnterpriseDemand.demand_type_code,
        EnterpriseDemand.status
    ).all()

    # 汇总为 { code: { pending: n, processing: n, resolved: n, count: total } }
    type_stats = {}
    for code, status, cnt in type_rows:
        code = code or '__empty__'
        if code not in type_stats:
            type_stats[code] = {'pending': 0, 'processing': 0, 'resolved': 0, 'count': 0}
        type_stats[code][status] = cnt
        type_stats[code]['count'] += cnt

    # ---- 查询各诉求类型关联的项目列表（用于悬停提示）----
    type_project_rows = demand_query().with_entities(
        EnterpriseDemand.demand_type_code,
        EnterpriseDemand.project_id,
        InvestmentProject.project_name
    ).join(
        InvestmentProject,
        EnterpriseDemand.project_id == InvestmentProject.id
    ).filter(
        EnterpriseDemand.demand_type_code.isnot(None),
        EnterpriseDemand.demand_type_code != ''
    ).distinct().all()

    type_projects = {}
    for code, pid, pname in type_project_rows:
        if code not in type_projects:
            type_projects[code] = []
        if len(type_projects[code]) < 15:  # 最多保留15个项目
            if not any(p['id'] == pid for p in type_projects[code]):
                type_projects[code].append({'id': pid, 'name': pname})

    by_type = []
    for code, s in sorted(type_stats.items(), key=lambda x: x[1]['count'], reverse=True):
        clean_code = code if code != '__empty__' else ''
        parent = type_parent_map.get(clean_code, '')
        by_type.append({
            'code': clean_code,
            'name': type_map.get(code, type_name_simple.get(clean_code, '未分类')) if code != '__empty__' else '未分类',
            'parent_code': parent,
            'count': s['count'],
            'pending': s['pending'],
            'processing': s['processing'],
            'resolved': s['resolved'],
            'projects': type_projects.get(clean_code, [])
        })

    # ---- 一级类型汇总（用于默认视图）----
    level1_stats = {}  # { parent_code: { pending, processing, resolved, count, name } }
    for item in by_type:
        parent = item['parent_code'] or item['code']  # 无父级的视为一级
        if parent not in level1_stats:
            level1_stats[parent] = {'pending': 0, 'processing': 0, 'resolved': 0, 'count': 0,
                                     'name': type_name_simple.get(parent, parent), 'code': parent}
        for k in ['pending', 'processing', 'resolved', 'count']:
            level1_stats[parent][k] += item[k]

    by_type_level1 = sorted(level1_stats.values(), key=lambda x: x['count'], reverse=True)

    # ---- 按状态分布 ----
    status_labels = {'pending': '待处理', 'processing': '处理中', 'resolved': '已解决'}
    status_colors = {'pending': '#f56c6c', 'processing': '#e6a23c', 'resolved': '#67c23a'}
    status_rows = demand_query().with_entities(
        EnterpriseDemand.status,
        func.count(EnterpriseDemand.id)
    ).group_by(EnterpriseDemand.status).all()

    by_status = []
    for status, cnt in status_rows:
        by_status.append({
            'status': status,
            'label': status_labels.get(status, status),
            'color': status_colors.get(status, '#909399'),
            'count': cnt
        })
    # 保证顺序
    by_status.sort(key=lambda x: ['pending', 'processing', 'resolved'].index(x['status']) if x['status'] in ['pending', 'processing', 'resolved'] else 99)

    # ---- 按对接单位分布（Top 15）----
    org_map = {d.code: d.name for d in OrganizationDict.query.all()}
    unit_rows = demand_query().with_entities(
        EnterpriseDemand.unit_code,
        func.count(EnterpriseDemand.id)
    ).filter(
        EnterpriseDemand.unit_code.isnot(None),
        EnterpriseDemand.unit_code != ''
    ).group_by(EnterpriseDemand.unit_code).order_by(
        func.count(EnterpriseDemand.id).desc()
    ).limit(15).all()

    # 查询各单位的诉求类型明细（用于悬停展示）—— 仅限 Top 15 单位
    unit_codes = [code for code, _ in unit_rows]
    unit_type_rows = []
    if unit_codes:
        unit_type_rows = demand_query().with_entities(
            EnterpriseDemand.unit_code,
            EnterpriseDemand.demand_type_code,
            func.count(EnterpriseDemand.id)
        ).filter(
            EnterpriseDemand.unit_code.in_(unit_codes)
        ).group_by(
            EnterpriseDemand.unit_code,
            EnterpriseDemand.demand_type_code
        ).order_by(
            EnterpriseDemand.unit_code,
            func.count(EnterpriseDemand.id).desc()
        ).all()

    # 构建排序查找表：dict_code → (parent_code, sort_order, simple_name)
    type_sort_map = {}
    for d in all_type_dicts:
        type_sort_map[d.code] = {
            'parent_code': d.parent_code or '',
            'sort_order': d.sort_order,
            'simple_name': d.name
        }
    # 一级类型的 sort_order 查找（用于分组排序）
    level1_sort = {d.code: d.sort_order for d in all_type_dicts if not d.parent_code}

    # 查询各单位各类型的关联项目（用于悬停展示）
    unit_type_projects = {}
    if unit_codes:
        utp_rows = demand_query().with_entities(
            EnterpriseDemand.unit_code,
            EnterpriseDemand.demand_type_code,
            EnterpriseDemand.project_id,
            InvestmentProject.project_name
        ).join(
            InvestmentProject, EnterpriseDemand.project_id == InvestmentProject.id
        ).filter(
            EnterpriseDemand.unit_code.in_(unit_codes),
            EnterpriseDemand.demand_type_code.isnot(None),
            EnterpriseDemand.demand_type_code != ''
        ).distinct().all()
        for ucode, dcode, pid, pname in utp_rows:
            key = (ucode, dcode)
            if key not in unit_type_projects:
                unit_type_projects[key] = []
            if len(unit_type_projects[key]) < 10:
                if not any(p['id'] == pid for p in unit_type_projects[key]):
                    unit_type_projects[key].append({'id': pid, 'name': pname})

    unit_types = {}
    for code, dcode, cnt in unit_type_rows:
        dcode = dcode or '__empty__'
        if code not in unit_types:
            unit_types[code] = []
        tinfo = type_sort_map.get(dcode, {})
        parent = tinfo.get('parent_code', '')
        unit_types[code].append({
            'name': type_map.get(dcode, type_name_simple.get(dcode, '未分类')),
            'parent_code': parent,
            'parent_sort_order': level1_sort.get(parent, 99),
            'sort_order': tinfo.get('sort_order', 0),
            'count': cnt,
            'projects': unit_type_projects.get((code, dcode), [])
        })

    by_unit = []
    for code, cnt in unit_rows:
        by_unit.append({
            'code': code,
            'name': org_map.get(code, code),
            'count': cnt,
            'types': unit_types.get(code, [])
        })

    # ---- 按项目分布（Top 10）----
    project_rows = demand_query().with_entities(
        EnterpriseDemand.project_id,
        func.count(EnterpriseDemand.id).label('cnt')
    ).group_by(EnterpriseDemand.project_id).order_by(
        func.count(EnterpriseDemand.id).desc()
    ).limit(10).all()

    project_ids = [row[0] for row in project_rows]
    project_map = {}
    if project_ids:
        projects = InvestmentProject.query.filter(InvestmentProject.id.in_(project_ids)).all()
        project_map = {p.id: p.project_name for p in projects}

    by_project = []
    for pid, cnt in project_rows:
        by_project.append({
            'project_id': pid,
            'project_name': project_map.get(pid, f'项目#{pid}'),
            'count': cnt
        })

    # ---- 月度趋势（最近12个月）----
    now = datetime.utcnow()
    twelve_months_ago = now - timedelta(days=365)

    month_rows = demand_query().with_entities(
        func.strftime('%Y-%m', EnterpriseDemand.created_at).label('month'),
        func.count(EnterpriseDemand.id)
    ).filter(
        EnterpriseDemand.created_at >= twelve_months_ago
    ).group_by('month').order_by('month').all()

    # 补全缺失的月份为 0
    month_dict = {m: c for m, c in month_rows}
    by_month = []
    cursor = datetime(twelve_months_ago.year, twelve_months_ago.month, 1)
    end = datetime(now.year, now.month, 1)
    while cursor <= end:
        key = cursor.strftime('%Y-%m')
        by_month.append({'month': key, 'count': month_dict.get(key, 0)})
        # 下一个月
        if cursor.month == 12:
            cursor = cursor.replace(year=cursor.year + 1, month=1)
        else:
            cursor = cursor.replace(month=cursor.month + 1)

    return jsonify({
        'code': 0,
        'data': {
            'overview': overview,
            'by_type': by_type,
            'by_type_level1': by_type_level1,
            'by_status': by_status,
            'by_unit': by_unit,
            'by_project': by_project,
            'by_month': by_month
        }
    })
