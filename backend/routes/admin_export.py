import io
import json
import re as _re
from datetime import date, datetime, timedelta
from flask import request, jsonify, send_file
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from models import ExportTemplate, ExportFieldConfig, PrintTemplate, PrintFieldConfig
from models import InvestmentProject, EnterpriseDemand, InvestmentActivity
from models import FollowStatusDict, MeetingStatusDict, OrganizationDict, ProjectTypeDict, DemandTypeDict
from extensions import db
from routes import admin_export_bp


def _sanitize_filename(name):
    """移除文件名中的非法字符（仅去除 Windows 真正禁止的字符）"""
    name = _re.sub(r'[\[\]*?"<>|\\/:]', '', name)
    name = _re.sub(r'_+', '_', name)
    return name.strip() or f'导出文件_{date.today().strftime("%Y%m%d")}'


# ============================================================
# 模板 CRUD
# ============================================================

@admin_export_bp.route('/export/templates', methods=['GET'])
@login_required
def get_templates():
    """列出所有打印模板（统一模板）"""
    entity_type = request.args.get('entity_type', 'investment').strip()
    templates = PrintTemplate.query.filter_by(entity_type=entity_type)\
        .order_by(PrintTemplate.id.asc()).all()
    return jsonify({'code': 0, 'data': [t.to_dict() for t in templates]})


@admin_export_bp.route('/export/templates', methods=['POST'])
@login_required
def create_template():
    """新建打印模板"""
    data = request.get_json()
    if not data or not data.get('name'):
        return jsonify({'code': 1, 'message': '模板名称不能为空'}), 400
    entity_type = data.get('entity_type', 'investment')
    template = PrintTemplate(name=data['name'].strip(), entity_type=entity_type)
    db.session.add(template)
    db.session.flush()

    # 从已有模板（排除自身）复制字段配置
    default_template = PrintTemplate.query.filter_by(entity_type=entity_type)\
        .filter(PrintTemplate.id != template.id)\
        .order_by(PrintTemplate.id.asc()).first()
    if default_template:
        default_fields = PrintFieldConfig.query.filter_by(template_id=default_template.id)\
            .order_by(PrintFieldConfig.sort_order).all()
        for f in default_fields:
            db.session.add(PrintFieldConfig(
                template_id=template.id,
                field_key=f.field_key,
                field_label=f.field_label,
                is_visible=True,
                column_width=f.column_width,
                sort_order=f.sort_order
            ))

    db.session.commit()
    return jsonify({'code': 0, 'data': template.to_dict(), 'message': '模板已创建'})


@admin_export_bp.route('/export/templates/<int:template_id>', methods=['PUT'])
@login_required
def update_template(template_id):
    """重命名模板"""
    tpl = PrintTemplate.query.get(template_id)
    if not tpl:
        return jsonify({'code': 1, 'message': '模板不存在'}), 404
    data = request.get_json()
    if not data or not data.get('name'):
        return jsonify({'code': 1, 'message': '模板名称不能为空'}), 400
    tpl.name = data['name'].strip()
    db.session.commit()
    return jsonify({'code': 0, 'data': tpl.to_dict(), 'message': '模板已更新'})


@admin_export_bp.route('/export/templates/<int:template_id>', methods=['DELETE'])
@login_required
def delete_template(template_id):
    """删除模板及其字段配置"""
    tpl = PrintTemplate.query.get(template_id)
    if not tpl:
        return jsonify({'code': 1, 'message': '模板不存在'}), 404
    # 至少保留一个模板
    remaining = PrintTemplate.query.filter_by(entity_type=tpl.entity_type).count()
    if remaining <= 1:
        return jsonify({'code': 1, 'message': '至少保留一个打印模板'}), 400
    PrintFieldConfig.query.filter_by(template_id=template_id).delete()
    db.session.delete(tpl)
    db.session.commit()
    return jsonify({'code': 0, 'message': '模板已删除'})


# ============================================================
# 导出字段配置 CRUD
# ============================================================

def _get_default_template_id():
    """获取默认模板ID（entity_type='investment' 的第一个模板）"""
    tpl = PrintTemplate.query.filter_by(entity_type='investment').order_by(PrintTemplate.id).first()
    return tpl.id if tpl else 1


@admin_export_bp.route('/export/fields', methods=['GET'])
@login_required
def get_export_fields():
    """获取指定模板的打印字段配置"""
    template_id = request.args.get('template_id', 0, type=int) or _get_default_template_id()
    fields = PrintFieldConfig.query.filter_by(template_id=template_id)\
        .order_by(PrintFieldConfig.sort_order).all()
    return jsonify({'code': 0, 'data': [f.to_dict() for f in fields]})


@admin_export_bp.route('/export/fields', methods=['PUT'])
@login_required
def update_export_fields():
    """批量更新打印字段配置"""
    data = request.get_json()
    if not data or not isinstance(data, list):
        return jsonify({'code': 1, 'message': '请求数据格式错误'}), 400

    template_id = None
    saved_ids = set()
    for item in data:
        field_id = item.get('id')
        if field_id and field_id > 0:
            saved_ids.add(field_id)
            field = PrintFieldConfig.query.get(field_id)
            if field:
                template_id = field.template_id
        elif item.get('_new') and item.get('template_id'):
            # 新建字段（自定义列）
            template_id = int(item['template_id'])
            field = PrintFieldConfig(
                template_id=template_id,
                field_key=item.get('field_key', ''),
                field_label=item.get('field_label', ''),
                is_visible=item.get('is_visible', True),
                is_custom=item.get('is_custom', True),
                column_width=item.get('column_width', 120),
                sort_order=item.get('sort_order', 0)
            )
            db.session.add(field)
            continue

        if field:
            if 'field_label' in item:
                field.field_label = item['field_label']
            if 'is_visible' in item:
                field.is_visible = bool(item['is_visible'])
            if 'column_width' in item:
                field.column_width = int(item['column_width'])
            if 'sort_order' in item:
                field.sort_order = int(item['sort_order'])

    # 删除在前端被移除的自定义字段
    if template_id:
        existing_ids = {f.id for f in PrintFieldConfig.query.filter_by(template_id=template_id).all()}
        to_delete = existing_ids - saved_ids
        if to_delete:
            PrintFieldConfig.query.filter(PrintFieldConfig.id.in_(to_delete)).delete(synchronize_session=False)

    db.session.commit()
    fields = PrintFieldConfig.query.filter_by(template_id=template_id)\
        .order_by(PrintFieldConfig.sort_order).all() if template_id else []
    return jsonify({'code': 0, 'data': [f.to_dict() for f in fields], 'message': '配置已保存'})


# ============================================================
# V15.3: 动态范围 统一过滤选项解析（兼容旧 activity_range / progress_range）
# ============================================================

def _parse_activity_filter(
    mode=None, count=None, time_mode=None, months=None, start=None, end=None,
    # 兼容旧参数
    activity_range=None, progress_range=None
):
    """统一新旧参数，返回标准过滤选项 dict。

    新参数（前端 V15.3+）:
        mode: 'count' | 'time' | ''
        count: int    （mode=count 时）
        time_mode: 'month' | 'daterange'  （mode=time 时）
        months: int   （time_mode=month 时，1个月=30天）
        start, end: 'YYYY-MM-DD'  （time_mode=daterange 时，含当天结束）

    旧参数（兼容）:
        activity_range: '' | 'last5' | 'last1m' | 'last3m'
        progress_range: '' | 'last2' | 'last5' | 'last1m' | 'last3m'

    返回:
        {'mode': '', 'count': 0, 'time_mode': '', 'months': 0, 'start_date': '', 'end_date': ''}
    """
    # 1) 旧参数 → 新参数
    if activity_range == 'last5':
        mode, count = 'count', 5
    elif activity_range == 'last1m':
        mode, time_mode, months = 'time', 'month', 1
    elif activity_range == 'last3m':
        mode, time_mode, months = 'time', 'month', 3

    if progress_range == 'last2':
        mode, count = 'count', 2
    elif progress_range == 'last5':
        mode, count = 'count', 5
    elif progress_range == 'last1m':
        mode, time_mode, months = 'time', 'month', 1
    elif progress_range == 'last3m':
        mode, time_mode, months = 'time', 'month', 3

    # 2) 构造默认 dict
    return {
        'mode': mode or '',
        'count': int(count) if count else 0,
        'time_mode': time_mode or '',
        'months': int(months) if months else 0,
        'start_date': start or '',
        'end_date': end or '',
    }


# ============================================================
# 导出预览
# ============================================================

def _aggregate_activities(project, activity_range='', filter_opts=None):
    """聚合项目招商动态：按日期倒序抓取最新数据，输出时反转从远到近排列

    兼容模式：
      - 旧调用方: _aggregate_activities(project, 'last5')
      - 新调用方: _aggregate_activities(project, '', filter_opts=_parse_activity_filter(...))
    """
    # 兼容：仅传旧 activity_range 时，内部转换成 filter_opts
    if not filter_opts:
        filter_opts = _parse_activity_filter(activity_range=activity_range)

    q = InvestmentActivity.query.filter_by(project_id=project.id)
    mode = filter_opts.get('mode', '')

    if mode == 'count':
        q = q.order_by(InvestmentActivity.date.desc())
        if filter_opts.get('count', 0) > 0:
            q = q.limit(filter_opts['count'])
    elif mode == 'time':
        if filter_opts.get('time_mode') == 'month' and filter_opts.get('months', 0) > 0:
            since = datetime.utcnow() - timedelta(days=30 * filter_opts['months'])
            q = q.filter(InvestmentActivity.date >= since)
        elif filter_opts.get('time_mode') == 'daterange':
            if filter_opts.get('start_date'):
                q = q.filter(InvestmentActivity.date >= filter_opts['start_date'])
            if filter_opts.get('end_date'):
                # 含当天：结束日期视作当天 23:59:59
                end_dt = datetime.strptime(filter_opts['end_date'], '%Y-%m-%d') + timedelta(days=1, seconds=-1)
                q = q.filter(InvestmentActivity.date <= end_dt)
    # mode 为空 → 全量

    activities = q.order_by(InvestmentActivity.date.desc()).all()
    if not activities:
        return ''

    # 反转列表：从远到近输出
    activities.reverse()

    lines = []
    for i, a in enumerate(activities, 1):
        content = a.content or ''
        lines.append(f'{i}、{content}')
    return '\n'.join(lines)


def _aggregate_demands(project, demand_status=''):
    """聚合企业诉求，按 sort_order，序号分段。
    demand_status: 逗号分隔的状态码过滤，为空则不过滤（默认仅待回应+协调中）"""
    q = EnterpriseDemand.query.filter_by(project_id=project.id)
    if demand_status:
        status_list = [s.strip() for s in demand_status.split(',') if s.strip()]
        if status_list:
            q = q.filter(EnterpriseDemand.status.in_(status_list))
    demands = q.order_by(EnterpriseDemand.sort_order.asc()).all()
    if not demands:
        return '暂无'

    type_map = DemandTypeDict.build_display_name_map()
    lines = []
    for i, d in enumerate(demands, 1):
        codes = [c.strip() for c in (d.demand_type_code or '').split(',') if c.strip()]
        type_name = '、'.join([type_map.get(c, c) for c in codes]) if codes else (d.demand_type_code or '诉求')
        content = d.demand_content or '暂无'
        lines.append(f'{i}、[{type_name}] {content}')
    return '\n'.join(lines)


def _aggregate_resolution(project, demand_status=''):
    """聚合解决措施，按 sort_order，仅含 resolution 非空的诉求。
    demand_status: 逗号分隔的状态码过滤，为空则不过滤"""
    q = EnterpriseDemand.query.filter_by(project_id=project.id)\
        .filter(EnterpriseDemand.resolution != '')\
        .filter(EnterpriseDemand.resolution.isnot(None))
    if demand_status:
        status_list = [s.strip() for s in demand_status.split(',') if s.strip()]
        if status_list:
            q = q.filter(EnterpriseDemand.status.in_(status_list))
    demands = q.order_by(EnterpriseDemand.sort_order.asc()).all()
    if not demands:
        return '暂无'

    lines = []
    for i, d in enumerate(demands, 1):
        lines.append(f'{i}、{d.resolution}')
    return '\n'.join(lines)


def _resolve_team_leader_names(p):
    """解析专班负责人 ID 列表为中文名列表"""
    _team_names = []
    try:
        _leader_ids = json.loads(getattr(p, 'team_leader_ids', '[]') or '[]')
        if _leader_ids:
            from models import Staff
            _staff_list = Staff.query.filter(Staff.id.in_(_leader_ids)).all()
            _staff_map = {s.id: s.name for s in _staff_list}
            _team_names = [_staff_map.get(sid, str(sid)) for sid in _leader_ids]
    except Exception:
        pass
    return '、'.join(_team_names)


def _resolve_project_row(p, activity_range='', demand_status='', filter_opts=None):
    """将项目对象解析为展示用 dict（含字典名称 + 聚合字段）
    demand_status: 逗号分隔的状态码过滤诉求，默认不过滤
    filter_opts: V15.3 可选，标准过滤选项 dict（与 activity_range 兼容）"""
    follow_map = {d.code: d.name for d in FollowStatusDict.query.all()}
    meeting_map = {d.code: d.name for d in MeetingStatusDict.query.all()}
    org_map = {d.code: d.name for d in OrganizationDict.query.all()}
    type_map = {d.code: d.name for d in ProjectTypeDict.query.all()}

    responsible_unit = org_map.get(p.responsible_unit_code, p.responsible_unit_code)
    person_in_charge = p.person_in_charge or ''
    person_in_charge_phone = getattr(p, 'person_in_charge_phone', '') or ''

    return {
        'order_no': p.order_no,
        'project_name': p.project_name,
        'project_type_code': type_map.get(p.project_type_code, p.project_type_code),
        'invest_enterprise': p.invest_enterprise,
        'enterprise_info': p.enterprise_info or '',
        'project_content': p.project_content or '',
        'invest_amount': float(p.invest_amount) if p.invest_amount else 0,
        'invest_amount_yi': round(float(p.invest_amount) / 10000, 4) if p.invest_amount else 0,
        'follow_status_code': follow_map.get(p.follow_status_code, p.follow_status_code),
        'meeting_status_code': meeting_map.get(p.meeting_status_code, p.meeting_status_code),
        'recommend_unit_code': org_map.get(p.recommend_unit_code, p.recommend_unit_code or ''),
        'responsible_unit_code': responsible_unit,
        'person_in_charge': person_in_charge,
        'person_in_charge_phone': person_in_charge_phone,
        # V3 组合字段
        '_combined_contact': '\n'.join([responsible_unit, person_in_charge, person_in_charge_phone]),
        # V15.3: 农高区谋划（在谈）项目库 专用组合字段
        '_contact_person_phone': '\n'.join([person_in_charge, person_in_charge_phone]) if (person_in_charge or person_in_charge_phone) else '',
        '_settle_location': '',  # 数据库目前无此字段，预留列（导出时空字符串）
        'project_doc': p.project_doc or '',
        'investment_plan': p.investment_plan or '',
        'conclusion': p.conclusion or '',
        'first_contact_date': p.first_contact_date.isoformat() if p.first_contact_date else '',
        # 聚合字段
        'activities': _aggregate_activities(p, activity_range, filter_opts=filter_opts),
        'demands': _aggregate_demands(p, demand_status),
        'resolution': _aggregate_resolution(p, demand_status),
        # A3 模板字段——专班跟进人
        '_team_followers': _resolve_team_leader_names(p),
    }


@admin_export_bp.route('/export/preview', methods=['POST'])
@login_required
def export_preview():
    """导出预览：返回表头+前3条数据"""
    data = request.get_json() or {}
    project_ids = data.get('project_ids', [])
    template_id = data.get('template_id', 0) or _get_default_template_id()

    fields = PrintFieldConfig.query.filter_by(template_id=template_id, is_visible=True)\
        .order_by(PrintFieldConfig.sort_order).all()

    headers = [{'field_key': f.field_key, 'field_label': f.field_label, 'column_width': f.column_width}
               for f in fields]

    q = InvestmentProject.query.filter_by(is_deleted=False)
    if project_ids:
        q = q.filter(InvestmentProject.id.in_(project_ids))
    projects = q.order_by(InvestmentProject.order_no.asc()).limit(3).all()

    rows = [_resolve_project_row(p) for p in projects]

    return jsonify({'code': 0, 'data': {'headers': headers, 'rows': rows, 'total': q.count()}})


@admin_export_bp.route('/export/download', methods=['GET'])
@login_required
def export_download():
    """生成并下载 Excel 文件（使用统一打印模板）"""
    ids_str = request.args.get('project_ids', '')
    project_ids = [int(x) for x in ids_str.split(',') if x.strip().isdigit()] if ids_str else []

    template_id = request.args.get('template_id', 0, type=int) or _get_default_template_id()
    activity_range = request.args.get('activity_range', '').strip()
    demand_mode = request.args.get('demand_mode', 'aggregate').strip()

    template = PrintTemplate.query.get(template_id)
    template_name = template.name if template else '招商项目库'

    fields = PrintFieldConfig.query.filter_by(template_id=template_id, is_visible=True)\
        .order_by(PrintFieldConfig.sort_order).all()
    if not fields:
        return jsonify({'code': 1, 'message': '未配置打印字段'}), 400

    q = InvestmentProject.query.filter_by(is_deleted=False)
    if project_ids:
        q = q.filter(InvestmentProject.id.in_(project_ids))
    projects = q.order_by(InvestmentProject.order_no.asc()).all()

    # 字典映射
    follow_map = {d.code: d.name for d in FollowStatusDict.query.all()}
    meeting_map = {d.code: d.name for d in MeetingStatusDict.query.all()}
    org_map = {d.code: d.name for d in OrganizationDict.query.all()}
    demand_type_map = DemandTypeDict.build_display_name_map()

    # ---- 样式定义（使用打印模板的字体设置） ----
    _tf = template.title_font_family or template.font_family or '微软雅黑'
    _ts = template.title_font_size or template.header_font_size or 14
    _hf = template.table_header_font_family or template.font_family or '微软雅黑'
    _hs = template.table_header_font_size or template.font_size or 10
    _cf = template.cell_font_family or template.font_family or '微软雅黑'
    _cs = template.cell_font_size or template.font_size or 10

    title_font = Font(name=_tf, size=_ts, bold=True, color='1A3A5C')
    title_align = Alignment(horizontal='center', vertical='center')
    header_font = Font(name=_hf, size=_hs, bold=True, color='000000')
    header_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name=_cf, size=_cs)
    cell_align = Alignment(vertical='top', wrap_text=True)
    cell_align_center = Alignment(horizontal='center', vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='333333'),
        right=Side(style='thin', color='333333'),
        top=Side(style='thin', color='333333'),
        bottom=Side(style='thin', color='333333'),
    )
    title_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    wb = Workbook()
    ws = wb.active
    ws.title = template_name[:31]

    if demand_mode == 'row':
        # ---- 按行导出：诉求拆分为独立行 ----
        project_fields = [f for f in fields if f.field_key not in ('demands', 'resolution')]
        demand_cols = [
            {'field_key': 'demand_type', 'field_label': '诉求类型', 'column_width': 120},
            {'field_key': 'demand_unit', 'field_label': '对接部门', 'column_width': 140},
            {'field_key': 'demand_content', 'field_label': '诉求内容', 'column_width': 400},
            {'field_key': 'demand_resolution', 'field_label': '解决措施', 'column_width': 400},
        ]
        total_cols = len(project_fields) + len(demand_cols)

        # 第1行：标题行（合并单元格）
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(row=1, column=1, value=template_name)
        title_cell.font = title_font
        title_cell.alignment = title_align
        title_cell.fill = title_fill
        for c in range(1, total_cols + 1):
            ws.cell(row=1, column=c).border = thin_border

        # 第2行：表头
        for col_idx, f in enumerate(project_fields, 1):
            cell = ws.cell(row=2, column=col_idx, value=f.field_label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = f.column_width / 7
        for col_idx, dc in enumerate(demand_cols, len(project_fields) + 1):
            cell = ws.cell(row=2, column=col_idx, value=dc['field_label'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = dc['column_width'] / 7

        # 数据行 + 合并单元格
        current_row = 3
        for p in projects:
            project_row = _resolve_project_row(p, activity_range=activity_range)
            demands = EnterpriseDemand.query.filter_by(project_id=p.id)\
                .order_by(EnterpriseDemand.sort_order.asc()).all()

            if not demands:
                # 无诉求：一行，项目字段正常填，诉求字段填"暂无"
                for col_idx, f in enumerate(project_fields, 1):
                    val = _fmt_cell(project_row.get(f.field_key, ''))
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.font = cell_font
                    cell.alignment = cell_align
                    cell.border = thin_border
                for col_idx in range(len(project_fields) + 1, total_cols + 1):
                    cell = ws.cell(row=current_row, column=col_idx, value='暂无')
                    cell.font = cell_font
                    cell.alignment = cell_align_center
                    cell.border = thin_border
                current_row += 1
            else:
                start_row = current_row
                for d in demands:
                    # 项目字段
                    for col_idx, f in enumerate(project_fields, 1):
                        val = _fmt_cell(project_row.get(f.field_key, ''))
                        cell = ws.cell(row=current_row, column=col_idx, value=val)
                        cell.font = cell_font
                        cell.alignment = cell_align
                        cell.border = thin_border
                    # 诉求字段
                    dt_codes = [c.strip() for c in (d.demand_type_code or '').split(',') if c.strip()]
                    demand_data = {
                        'demand_type': '、'.join([demand_type_map.get(c, c) for c in dt_codes]) if dt_codes else '暂无',
                        'demand_unit': org_map.get(d.unit_code, d.unit_code or ''),
                        'demand_content': d.demand_content or '暂无',
                        'demand_resolution': d.resolution or '暂无',
                    }
                    for col_idx, dc in enumerate(demand_cols, len(project_fields) + 1):
                        val = demand_data.get(dc['field_key'], '')
                        cell = ws.cell(row=current_row, column=col_idx, value=val)
                        cell.font = cell_font
                        cell.alignment = cell_align_center if col_idx <= len(project_fields) + 2 else cell_align
                        cell.border = thin_border
                    current_row += 1

                # 合并项目字段单元格（纵向）
                if current_row - start_row > 1:
                    for col_idx in range(1, len(project_fields) + 1):
                        ws.merge_cells(
                            start_row=start_row, start_column=col_idx,
                            end_row=current_row - 1, end_column=col_idx
                        )

        ws.freeze_panes = 'A3'
    else:
        # ---- 聚合导出（原有逻辑） ----
        total_cols = len(fields)

        # 第1行：标题行（合并单元格）
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(row=1, column=1, value=template_name)
        title_cell.font = title_font
        title_cell.alignment = title_align
        title_cell.fill = title_fill
        for c in range(1, total_cols + 1):
            ws.cell(row=1, column=c).border = thin_border

        # 第2行：表头
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=2, column=col_idx, value=field.field_label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = field.column_width / 7

        # 数据行
        for row_idx, p in enumerate(projects, 3):
            row_data = _resolve_project_row(p, activity_range=activity_range)
            for col_idx, field in enumerate(fields, 1):
                val = row_data.get(field.field_key, '')
                val = _fmt_cell(val)
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = cell_font
                cell.alignment = cell_align
                cell.border = thin_border

        ws.freeze_panes = 'A3'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{_sanitize_filename(template_name)}.xlsx'
    )


def _fmt_cell(val):
    """格式化单元格值"""
    if isinstance(val, date):
        return val.isoformat()
    return val
