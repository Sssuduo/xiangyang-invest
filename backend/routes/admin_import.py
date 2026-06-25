import io
import json
from datetime import date
from flask import request, jsonify, send_file
from flask_login import login_required
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from models import ImportFieldConfig, InvestmentProject, EnterpriseDemand
from models import FollowStatusDict, MeetingStatusDict, OrganizationDict, ProjectTypeDict
from extensions import db
from routes import admin_import_bp


# ============================================================
# 导入字段配置 CRUD
# ============================================================

@admin_import_bp.route('/import/fields', methods=['GET'])
@login_required
def get_import_fields():
    fields = ImportFieldConfig.query.order_by(ImportFieldConfig.sort_order).all()
    return jsonify({'code': 0, 'data': [f.to_dict() for f in fields]})


@admin_import_bp.route('/import/fields', methods=['PUT'])
@login_required
def update_import_fields():
    data = request.get_json()
    if not data or not isinstance(data, list):
        return jsonify({'code': 1, 'message': '格式错误'}), 400
    for item in data:
        f = ImportFieldConfig.query.get(item.get('id'))
        if f:
            if 'field_label' in item:
                f.field_label = item['field_label']
            if 'is_enabled' in item:
                f.is_enabled = bool(item['is_enabled'])
            if 'is_required' in item:
                f.is_required = bool(item['is_required'])
            if 'sort_order' in item:
                f.sort_order = int(item['sort_order'])
    db.session.commit()
    fields = ImportFieldConfig.query.order_by(ImportFieldConfig.sort_order).all()
    return jsonify({'code': 0, 'data': [f.to_dict() for f in fields], 'message': '已保存'})


# ============================================================
# 下载导入模板
# ============================================================

@admin_import_bp.route('/import/template', methods=['GET'])
@login_required
def download_template():
    fields = ImportFieldConfig.query.filter_by(is_enabled=True)\
        .order_by(ImportFieldConfig.sort_order).all()
    if not fields:
        return jsonify({'code': 1, 'message': '未配置导入模板'}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = '项目导入模板'

    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='3A7ABD', end_color='3A7ABD', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )

    # 写入表头
    for col_idx, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_idx, value=field.field_label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    # 写入示例数据行
    sample = {
        'order_no': '1',
        'project_name': '示例项目名称',
        'project_type_code': '设施农业',
        'invest_enterprise': '示例企业名称',
        'enterprise_info': '企业简介示例',
        'project_content': '项目内容描述',
        'invest_amount': '5000',
        'follow_status_code': '重点跟进',
        'meeting_status_code': '未上会',
        'recommend_unit_code': '区招商服务中心',
        'responsible_unit_code': '农高区创建专班',
        'person_in_charge': '张三',
        'first_contact_date': '2026-01-01',
    }
    for col_idx, field in enumerate(fields, 1):
        val = sample.get(field.field_key, '')
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.font = Font(name='微软雅黑', size=10, color='909090')
        cell.border = thin_border

    # 必填标记（第3行说明）
    for col_idx, field in enumerate(fields, 1):
        mark = '（必填）' if field.is_required else '（选填）'
        cell = ws.cell(row=3, column=col_idx, value=mark)
        cell.font = Font(name='微软雅黑', size=9, color='C00000' if field.is_required else '909090')

    ws.freeze_panes = 'A4'

    # ---- 字典字段添加下拉验证 ----
    # 构建字典 field_key -> [选项列表]
    dict_options = {}

    # 跟进状态
    follow_list = [d.name for d in FollowStatusDict.query.order_by(FollowStatusDict.sort_order).all()]
    if follow_list:
        dict_options['follow_status_code'] = follow_list

    # 上会状态
    meeting_list = [d.name for d in MeetingStatusDict.query.order_by(MeetingStatusDict.sort_order).all()]
    if meeting_list:
        dict_options['meeting_status_code'] = meeting_list

    # 项目类型
    type_list = [d.name for d in ProjectTypeDict.query.filter_by(is_active=True).order_by(ProjectTypeDict.sort_order).all()]
    if type_list:
        dict_options['project_type_code'] = type_list

    # 单位（推介单位、责任单位共用）
    org_list = [d.name for d in OrganizationDict.query.filter_by(is_active=True).order_by(OrganizationDict.sort_order).all()]
    if org_list:
        dict_options['recommend_unit_code'] = org_list
        dict_options['responsible_unit_code'] = org_list

    # 为每个字典字段列添加下拉验证（数据行从第4行到第1000行）
    for col_idx, field in enumerate(fields, 1):
        options = dict_options.get(field.field_key)
        if options:
            col_letter = get_column_letter(col_idx)
            options_str = ','.join(options)
            dv = DataValidation(type='list', formula1=f'"{options_str}"', allow_blank=True)
            dv.error = '请从下拉列表中选择'
            dv.errorTitle = '输入错误'
            dv.prompt = '请选择'
            dv.promptTitle = field.field_label
            dv.sqref = f'{col_letter}4:{col_letter}1000'
            ws.add_data_validation(dv)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='项目导入模板.xlsx'
    )


# ============================================================
# 导入预览
# ============================================================

def _get_import_fields():
    return ImportFieldConfig.query.filter_by(is_enabled=True)\
        .order_by(ImportFieldConfig.sort_order).all()


def _get_dict_maps():
    return {
        'follow': {d.name: d.code for d in FollowStatusDict.query.all()},
        'meeting': {d.name: d.code for d in MeetingStatusDict.query.all()},
        'org': {d.name: d.code for d in OrganizationDict.query.all()},
        'type': {d.name: d.code for d in ProjectTypeDict.query.all()},
    }


def _resolve_code(field_key, name_val, maps):
    """将中文名转换为 code"""
    if field_key == 'follow_status_code':
        return maps['follow'].get(name_val, name_val)
    elif field_key == 'meeting_status_code':
        return maps['meeting'].get(name_val, name_val)
    elif field_key in ('recommend_unit_code', 'responsible_unit_code'):
        return maps['org'].get(name_val, name_val)
    elif field_key == 'project_type_code':
        return maps['type'].get(name_val, name_val)
    return name_val


def _validate_row(row_data, field_configs, existing_names, row_idx):
    """校验单行数据，返回错误列表"""
    errors = []
    for f in field_configs:
        val = row_data.get(f.field_key, '')
        # 必填检查
        if f.is_required and (val is None or str(val).strip() == ''):
            errors.append(f'第{row_idx}行：{f.field_label} 为必填项')
        # 类型校验
        if val is not None and str(val).strip() != '' and f.field_key == 'invest_amount':
            try:
                float(val)
            except (ValueError, TypeError):
                errors.append(f'第{row_idx}行：投资金额格式有误，应为数字')
        if f.field_key == 'first_contact_date' and val and str(val).strip():
            try:
                date.fromisoformat(str(val).strip())
            except (ValueError, TypeError):
                errors.append(f'第{row_idx}行：首次对接时间格式有误，应为 YYYY-MM-DD')
        if f.field_key == 'order_no' and val is not None and str(val).strip():
            try:
                int(val)
            except (ValueError, TypeError):
                errors.append(f'第{row_idx}行：顺序号格式有误，应为整数')
    # 项目名称唯一性检查（与已有数据对比）
    pname = str(row_data.get('project_name', '') or '').strip()
    if pname and pname in existing_names:
        errors.append(f'第{row_idx}行：项目名称「{pname}」已存在')
    return errors


@admin_import_bp.route('/import/preview', methods=['POST'])
@login_required
def import_preview():
    """上传 Excel 并返回预览数据"""
    if 'file' not in request.files:
        return jsonify({'code': 1, 'message': '未找到上传文件'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'code': 1, 'message': '未选择文件'}), 400

    try:
        wb = load_workbook(file, read_only=True)
        ws = wb.active
    except Exception:
        return jsonify({'code': 1, 'message': '无法读取文件，请确认是有效的 Excel 文件'}), 400

    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value is not None:
            headers.append(str(cell.value).strip())

    # 获取配置的导入字段
    import_fields = _get_import_fields()
    expected_headers = [f.field_label for f in import_fields]

    # 校验表头是否一致
    if headers != expected_headers:
        return jsonify({
            'code': 1,
            'message': '模板有误，请选择正确的导入模板',
            'data': {'expected': expected_headers, 'actual': headers}
        }), 400

    # 构建 field_key -> field_label 映射
    field_map = {f.field_label: f.field_key for f in import_fields}
    field_configs = {f.field_key: f for f in import_fields}

    # 获取已存在的项目名称
    existing_names = set(
        p.project_name for p in InvestmentProject.query.filter_by(is_deleted=False).all()
    )

    # 读取数据行（跳过表头、示例行、标记行）
    rows = []
    all_errors = []
    dict_maps = _get_dict_maps()

    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        # 跳过完全空行
        if all(v is None or str(v).strip() == '' for v in row):
            continue

        row_data = {}
        for col_idx, val in enumerate(row):
            if col_idx < len(headers):
                field_label = headers[col_idx]
                field_key = field_map.get(field_label, '')
                str_val = str(val).strip() if val is not None else ''
                # 字典字段：中文名转 code
                str_val = _resolve_code(field_key, str_val, dict_maps)
                row_data[field_key] = str_val

        # 校验
        row_errors = _validate_row(row_data, import_fields, existing_names, row_idx)
        # 同行内的名称唯一检查
        pname = row_data.get('project_name', '')
        if pname:
            # 检查前面行是否已有同名
            for prev in rows:
                if prev['data'].get('project_name') == pname and not any(
                    e.startswith(f'第{row_idx}行：项目名称') for e in row_errors
                ):
                    row_errors.append(f'第{row_idx}行：项目名称「{pname}」与第{prev["row"]}行重复')

        all_errors.extend(row_errors)
        rows.append({
            'row': row_idx,
            'data': row_data,
            'errors': row_errors,
            '_valid': len(row_errors) == 0
        })

    # 统计
    valid_count = sum(1 for r in rows if r['_valid'])
    error_count = len(rows) - valid_count

    return jsonify({
        'code': 0,
        'data': {
            'headers': expected_headers,
            'rows': rows,
            'total': len(rows),
            'valid_count': valid_count,
            'error_count': error_count,
            'all_errors': all_errors
        }
    })


# ============================================================
# 执行导入
# ============================================================

@admin_import_bp.route('/import/execute', methods=['POST'])
@login_required
def import_execute():
    """将预览通过的数据写入数据库"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 1, 'message': '数据为空'}), 400

    rows = data.get('rows', [])
    if not rows:
        return jsonify({'code': 1, 'message': '无数据可导入'}), 400

    # 过滤出有效行
    valid_rows = [r for r in rows if r.get('_valid')]
    if not valid_rows:
        return jsonify({'code': 1, 'message': '没有有效数据行'}), 400

    import_fields = _get_import_fields()
    field_keys = {f.field_key for f in import_fields}

    imported = 0
    for row in valid_rows:
        row_data = row['data']

        project = InvestmentProject(
            order_no=int(row_data.get('order_no', 0) or 0),
            project_name=row_data.get('project_name', ''),
            invest_enterprise=row_data.get('invest_enterprise', ''),
            enterprise_info=row_data.get('enterprise_info', ''),
            project_content=row_data.get('project_content', ''),
            invest_amount=float(row_data.get('invest_amount', 0) or 0),
            follow_status_code=row_data.get('follow_status_code', 'follow_keep'),
            meeting_status_code=row_data.get('meeting_status_code', 'not_meeting'),
            recommend_unit_code=row_data.get('recommend_unit_code', ''),
            responsible_unit_code=row_data.get('responsible_unit_code', ''),
            project_type_code=row_data.get('project_type_code', ''),
            person_in_charge=row_data.get('person_in_charge', ''),
            person_in_charge_phone=row_data.get('person_in_charge_phone', ''),
            project_doc=row_data.get('project_doc', ''),
            first_contact_date=_parse_date(row_data.get('first_contact_date'))
        )
        db.session.add(project)
        imported += 1

    db.session.commit()
    return jsonify({'code': 0, 'message': f'成功导入 {imported} 条记录', 'data': {'count': imported}})


def _parse_date(val):
    if not val:
        return None
    if isinstance(val, date):
        return val
    try:
        return date.fromisoformat(str(val).strip())
    except (ValueError, TypeError):
        return None
