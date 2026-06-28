"""
招商项目打印 API
提供打印模板 CRUD、字段配置、数据解析（复用导出模块的聚合逻辑）
"""
import re as _re
from datetime import date, datetime
from flask import request, jsonify, send_file
from models import PrintTemplate, PrintFieldConfig, InvestmentProject
from extensions import db
from routes import admin_print_bp
from routes.business_auth import dual_login_required, visitor_block

# 复用导出模块的数据解析函数
from routes.admin_export import _resolve_project_row as _export_resolve_row
from routes.admin_export import _aggregate_activities, _aggregate_demands, _aggregate_resolution


def _sanitize_filename(name):
    """移除文件名中的非法字符（仅去除 Windows 真正禁止的字符）"""
    name = _re.sub(r'[\[\]*?"<>|\\/:]', '', name)
    name = _re.sub(r'_+', '_', name)
    return name.strip() or f'导出文件_{date.today().strftime("%Y%m%d")}'


# ============================================================
# 模板 CRUD
# ============================================================

def _get_default_template_id():
    """获取默认打印模板ID（entity_type='investment' 的第一个模板）"""
    tpl = PrintTemplate.query.filter_by(entity_type='investment').order_by(PrintTemplate.id).first()
    return tpl.id if tpl else 1


@admin_print_bp.route('/print/templates', methods=['GET'])
@dual_login_required
def get_templates():
    """列出所有打印模板"""
    entity_type = request.args.get('entity_type', 'investment').strip()
    templates = PrintTemplate.query.filter_by(entity_type=entity_type)\
        .order_by(PrintTemplate.id.asc()).all()
    return jsonify({'code': 0, 'data': [t.to_dict() for t in templates]})


@admin_print_bp.route('/print/templates', methods=['POST'])
@dual_login_required
@visitor_block
def create_template():
    """新建打印模板"""
    data = request.get_json()
    if not data or not data.get('name'):
        return jsonify({'code': 1, 'message': '模板名称不能为空'}), 400
    entity_type = data.get('entity_type', 'investment')
    template = PrintTemplate(name=data['name'].strip(), entity_type=entity_type)
    db.session.add(template)
    db.session.flush()

    # 从已有模板（排除自身）复制字段配置
    default_template = PrintTemplate.query.filter_by(entity_type=entity_type)\
        .filter(PrintTemplate.id != template.id)\
        .order_by(PrintTemplate.id.asc()).first()
    if default_template:
        default_fields = PrintFieldConfig.query.filter_by(template_id=default_template.id)\
            .order_by(PrintFieldConfig.sort_order).all()
        for f in default_fields:
            db.session.add(PrintFieldConfig(
                template_id=template.id,
                field_key=f.field_key,
                field_label=f.field_label,
                is_visible=True,
                column_width=f.column_width,
                sort_order=f.sort_order
            ))

    db.session.commit()
    return jsonify({'code': 0, 'data': template.to_dict(), 'message': '模板已创建'})


@admin_print_bp.route('/print/templates/<int:template_id>', methods=['PUT'])
@dual_login_required
@visitor_block
def update_template(template_id):
    """更新模板（名称 + 布局参数）"""
    tpl = PrintTemplate.query.get(template_id)
    if not tpl:
        return jsonify({'code': 1, 'message': '模板不存在'}), 404
    data = request.get_json()
    if not data:
        return jsonify({'code': 1, 'message': '请求数据为空'}), 400

    # 更新名称
    if 'name' in data and data['name']:
        tpl.name = data['name'].strip()
    # 更新布局参数
    for key in ('paper_size', 'orientation', 'font_family', 'font_size',
                'header_font_size', 'margin_top', 'margin_bottom', 'margin_left',
                'margin_right', 'title_font_family', 'title_font_size',
                'table_header_font_family', 'table_header_font_size',
                'cell_font_family', 'cell_font_size',
                'sub_title_font_size', 'border_style', 'group_by'):
        if key in data:
            setattr(tpl, key, data[key])
    if 'show_page_number' in data:
        tpl.show_page_number = bool(data['show_page_number'])
    if 'show_print_meta' in data:
        tpl.show_print_meta = bool(data['show_print_meta'])
    if 'title_bold' in data:
        tpl.title_bold = bool(data['title_bold'])
    if 'subtitle_bold' in data:
        tpl.subtitle_bold = bool(data['subtitle_bold'])

    db.session.commit()
    return jsonify({'code': 0, 'data': tpl.to_dict(), 'message': '模板已更新'})


@admin_print_bp.route('/print/templates/<int:template_id>', methods=['DELETE'])
@dual_login_required
@visitor_block
def delete_template(template_id):
    """删除模板及其字段配置"""
    tpl = PrintTemplate.query.get(template_id)
    if not tpl:
        return jsonify({'code': 1, 'message': '模板不存在'}), 404
    # 至少保留一个模板
    remaining = PrintTemplate.query.filter_by(entity_type=tpl.entity_type).count()
    if remaining <= 1:
        return jsonify({'code': 1, 'message': '至少保留一个打印模板'}), 400
    PrintFieldConfig.query.filter_by(template_id=template_id).delete()
    db.session.delete(tpl)
    db.session.commit()
    return jsonify({'code': 0, 'message': '模板已删除'})


# ============================================================
# 字段配置 CRUD
# ============================================================

@admin_print_bp.route('/print/fields', methods=['GET'])
@dual_login_required
def get_print_fields():
    """获取指定模板的打印字段配置"""
    template_id = request.args.get('template_id', 0, type=int) or _get_default_template_id()
    fields = PrintFieldConfig.query.filter_by(template_id=template_id)\
        .order_by(PrintFieldConfig.sort_order).all()
    return jsonify({'code': 0, 'data': [f.to_dict() for f in fields]})


@admin_print_bp.route('/print/fields', methods=['PUT'])
@dual_login_required
@visitor_block
def update_print_fields():
    """批量更新打印字段配置"""
    data = request.get_json()
    if not data or not isinstance(data, list):
        return jsonify({'code': 1, 'message': '请求数据格式错误'}), 400

    template_id = None
    saved_ids = set()
    for item in data:
        field_id = item.get('id')
        if field_id and field_id > 0:
            saved_ids.add(field_id)
            field = PrintFieldConfig.query.get(field_id)
            if field:
                template_id = field.template_id
        elif item.get('_new') and item.get('template_id'):
            template_id = int(item['template_id'])
            field = PrintFieldConfig(
                template_id=template_id,
                field_key=item.get('field_key', ''),
                field_label=item.get('field_label', ''),
                is_visible=item.get('is_visible', True),
                is_custom=item.get('is_custom', True),
                column_width=item.get('column_width', 120),
                sort_order=item.get('sort_order', 0)
            )
            db.session.add(field)
            continue

        if field:
            if 'field_label' in item:
                field.field_label = item['field_label']
            if 'is_visible' in item:
                field.is_visible = bool(item['is_visible'])
            if 'column_width' in item:
                field.column_width = int(item['column_width'])
            if 'sort_order' in item:
                field.sort_order = int(item['sort_order'])

    # 删除在前端被移除的自定义字段
    if template_id:
        existing_ids = {f.id for f in PrintFieldConfig.query.filter_by(template_id=template_id).all()}
        to_delete = existing_ids - saved_ids
        if to_delete:
            PrintFieldConfig.query.filter(PrintFieldConfig.id.in_(to_delete)).delete(synchronize_session=False)

    db.session.commit()
    fields = PrintFieldConfig.query.filter_by(template_id=template_id)\
        .order_by(PrintFieldConfig.sort_order).all() if template_id else []
    return jsonify({'code': 0, 'data': [f.to_dict() for f in fields], 'message': '配置已保存'})


# ============================================================
# 模板文件上传 + 列映射
# ============================================================

import os as _os
TEMPLATE_UPLOAD_DIR = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), '..', 'static', 'uploads', 'templates')


def _parse_template_headers(file_path, header_row):
    """从上传的 .xlsx 模板中解析列标题，返回 [{column_letter, column_header}]"""
    import openpyxl
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    columns = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        header_text = str(cell.value).strip() if cell.value is not None else ''
        if header_text:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            columns.append({'column_letter': col_letter, 'column_header': header_text})
    wb.close()
    return columns


@admin_print_bp.route('/print/templates/<int:template_id>/upload', methods=['POST'])
@dual_login_required
@visitor_block
def upload_template_file(template_id):
    """上传 .xlsx 模板文件，自动解析列标题"""
    tpl = PrintTemplate.query.get(template_id)
    if not tpl:
        return jsonify({'code': 1, 'message': '模板不存在'}), 404

    if 'file' not in request.files:
        return jsonify({'code': 1, 'message': '未找到上传文件'}), 400

    file = request.files['file']
    if not file.filename or not file.filename.lower().endswith('.xlsx'):
        return jsonify({'code': 1, 'message': '仅支持 .xlsx 文件'}), 400

    try:
        # 保存文件
        from utils.image_upload import save_uploaded_image
        _os.makedirs(TEMPLATE_UPLOAD_DIR, exist_ok=True)
        # save_uploaded_image 返回 /static/uploads/{name}，需补上 /templates/ 子目录
        saved_rel = save_uploaded_image(file, TEMPLATE_UPLOAD_DIR)
        file_path = saved_rel.replace('/static/uploads/', '/static/uploads/templates/')

        # 解析列标题
        abs_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), '..', file_path.lstrip('/'))
        header_row = tpl.header_row or 2
        columns = _parse_template_headers(abs_path, header_row)

        # 清除旧映射，创建新映射
        from models import TemplateFieldMapping
        TemplateFieldMapping.query.filter_by(template_id=template_id).delete()

        for idx, col in enumerate(columns):
            db.session.add(TemplateFieldMapping(
                template_id=template_id,
                column_letter=col['column_letter'],
                column_header=col['column_header'],
                field_key='',
                sort_order=idx,
            ))

        # 更新模板记录
        tpl.template_file = file_path

        db.session.commit()

        mappings = TemplateFieldMapping.query.filter_by(template_id=template_id)\
            .order_by(TemplateFieldMapping.sort_order).all()
        return jsonify({'code': 0, 'data': [m.to_dict() for m in mappings],
                        'message': f'已解析 {len(columns)} 个列标题'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 1, 'message': f'上传失败: {str(e)}'}), 500


@admin_print_bp.route('/print/templates/<int:template_id>/mappings', methods=['GET'])
@dual_login_required
def get_template_mappings(template_id):
    """获取模板的列映射列表 + 可用系统字段"""
    tpl = PrintTemplate.query.get(template_id)
    if not tpl:
        return jsonify({'code': 1, 'message': '模板不存在'}), 404

    from models import TemplateFieldMapping
    mappings = TemplateFieldMapping.query.filter_by(template_id=template_id)\
        .order_by(TemplateFieldMapping.sort_order).all()

    # 可用系统字段（已有的 PrintFieldConfig 列表，用作下拉选项）
    available_fields = PrintFieldConfig.query.filter_by(template_id=template_id, is_visible=True)\
        .order_by(PrintFieldConfig.sort_order).all()

    return jsonify({'code': 0, 'data': {
        'mappings': [m.to_dict() for m in mappings],
        'available_fields': [{'field_key': f.field_key, 'field_label': f.field_label}
                             for f in available_fields],
    }})


@admin_print_bp.route('/print/templates/<int:template_id>/mappings', methods=['PUT'])
@dual_login_required
@visitor_block
def update_template_mappings(template_id):
    """批量更新列→字段映射关系"""
    tpl = PrintTemplate.query.get(template_id)
    if not tpl:
        return jsonify({'code': 1, 'message': '模板不存在'}), 404

    data = request.get_json()
    if not isinstance(data, list):
        return jsonify({'code': 1, 'message': '请求格式错误'}), 400

    from models import TemplateFieldMapping
    updated = 0
    for item in data:
        col_letter = item.get('column_letter', '')
        field_key = item.get('field_key', '')
        m = TemplateFieldMapping.query.filter_by(
            template_id=template_id, column_letter=col_letter).first()
        if m:
            m.field_key = field_key
            updated += 1

    db.session.commit()
    return jsonify({'code': 0, 'message': f'已更新 {updated} 条映射'})


@admin_print_bp.route('/print/templates/<int:template_id>/file', methods=['GET'])
@dual_login_required
def download_template_file(template_id):
    """下载当前模板文件"""
    tpl = PrintTemplate.query.get(template_id)
    if not tpl or not tpl.template_file:
        return jsonify({'code': 1, 'message': '模板文件不存在'}), 404

    abs_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), '..', tpl.template_file.lstrip('/'))
    if not _os.path.exists(abs_path):
        return jsonify({'code': 1, 'message': '模板文件未找到'}), 404

    return send_file(abs_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'{_sanitize_filename(tpl.name)}.xlsx')


@admin_print_bp.route('/print/templates/<int:template_id>/file', methods=['DELETE'])
@dual_login_required
@visitor_block
def delete_template_file(template_id):
    """删除模板文件及映射"""
    tpl = PrintTemplate.query.get(template_id)
    if not tpl:
        return jsonify({'code': 1, 'message': '模板不存在'}), 404

    from models import TemplateFieldMapping
    if tpl.template_file:
        abs_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), '..', tpl.template_file.lstrip('/'))
        if _os.path.exists(abs_path):
            _os.remove(abs_path)
        tpl.template_file = ''

    TemplateFieldMapping.query.filter_by(template_id=template_id).delete()
    db.session.commit()
    return jsonify({'code': 0, 'message': '模板文件已删除'})


# ============================================================
# 打印数据解析
# ============================================================

# V3 跟进状态分组优先级（越靠前越优先）
FOLLOW_STATUS_ORDER = ['重点跟进', '保持对接', '不再跟进']


def _group_rows_by_status(rows):
    """按跟进状态对行进行分组，返回 [{title, rows}]"""
    groups = []
    seen_statuses = set()

    # 按优先级顺序处理
    for status in FOLLOW_STATUS_ORDER:
        group_rows = [r for r in rows if r.get('follow_status_code') == status]
        if group_rows:
            groups.append({'title': status, 'rows': group_rows})
            seen_statuses.add(status)

    # 其他未知状态
    other_rows = [r for r in rows if r.get('follow_status_code') not in seen_statuses]
    if other_rows:
        # 收集未知状态名并按出现顺序分组
        other_statuses = []
        for r in other_rows:
            s = r.get('follow_status_code', '')
            if s not in other_statuses:
                other_statuses.append(s)
        for s in other_statuses:
            s_rows = [r for r in other_rows if r.get('follow_status_code') == s]
            if s_rows:
                groups.append({'title': s, 'rows': s_rows})

    # 如果没有分组（所有行无 follow_status_code），返回空 groups
    if not groups:
        groups.append({'title': None, 'rows': rows})

    return groups


def _resolve_project_row(p, activity_range='', demand_status=''):
    """将项目对象解析为展示用 dict（复用导出模块的解析逻辑）"""
    return _export_resolve_row(p, activity_range, demand_status)


@admin_print_bp.route('/print/data', methods=['GET'])
@dual_login_required
def get_print_data():
    """解析打印数据，返回 JSON（模板信息 + 表头 + 数据行）"""
    ids_str = request.args.get('project_ids', '')
    project_ids = [int(x) for x in ids_str.split(',') if x.strip().isdigit()] if ids_str else []

    template_id = request.args.get('template_id', 0, type=int) or _get_default_template_id()
    activity_range = request.args.get('activity_range', '').strip()
    demand_mode = request.args.get('demand_mode', 'aggregate').strip()
    demand_status = request.args.get('demand_status', 'pending,processing').strip()

    template = PrintTemplate.query.get(template_id)
    if not template:
        return jsonify({'code': 1, 'message': '打印模板不存在'}), 404

    fields = PrintFieldConfig.query.filter_by(template_id=template_id, is_visible=True)\
        .order_by(PrintFieldConfig.sort_order).all()
    if not fields:
        return jsonify({'code': 1, 'message': '未配置打印字段'}), 400

    q = InvestmentProject.query.filter_by(is_deleted=False)
    if project_ids:
        q = q.filter(InvestmentProject.id.in_(project_ids))
    projects = q.order_by(InvestmentProject.order_no.asc()).all()

    headers = [{'field_key': f.field_key, 'field_label': f.field_label,
                'column_width': f.column_width} for f in fields]

    if demand_mode == 'row':
        # 按行展开：每个诉求拆分为独立行，项目字段通过纵向合并
        from models import EnterpriseDemand, DemandTypeDict
        demand_type_map = DemandTypeDict.build_display_name_map()

        rows = []
        for p in projects:
            base = _resolve_project_row(p, activity_range, demand_status)
            q_demands = EnterpriseDemand.query.filter_by(project_id=p.id)
            if demand_status:
                status_list = [s.strip() for s in demand_status.split(',') if s.strip()]
                if status_list:
                    q_demands = q_demands.filter(EnterpriseDemand.status.in_(status_list))
            demands = q_demands.order_by(EnterpriseDemand.sort_order.asc()).all()
            if demands:
                for d in demands:
                    codes = [c.strip() for c in (d.demand_type_code or '').split(',') if c.strip()]
                    type_name = '、'.join([demand_type_map.get(c, c) for c in codes]) if codes else ''
                    row = dict(base)
                    # 子表字段附加到行
                    row['_demand_type'] = type_name
                    row['_demand_unit'] = d.demand_unit or ''
                    row['_demand_content'] = d.demand_content or ''
                    row['_demand_resolution'] = d.resolution or ''
                    rows.append(row)
            else:
                base['_demand_type'] = ''
                base['_demand_unit'] = ''
                base['_demand_content'] = ''
                base['_demand_resolution'] = ''
                rows.append(base)
    else:
        # 聚合模式
        rows = [_resolve_project_row(p, activity_range, demand_status) for p in projects]

    # V3: 按跟进状态分组
    groups = _group_rows_by_status(rows)

    return jsonify({
        'code': 0,
        'data': {
            'template': template.to_dict(),
            'headers': headers,
            'rows': rows,     # 保留扁平列表兼容
            'groups': groups,  # V3 分组数据
            'total': len(rows),
            'demand_mode': demand_mode,
            'group_mode': 'follow_status'
        }
    })


# ============================================================
# Excel 下载（使用 PrintTemplate 生成）
# ============================================================

def _fill_template_file(template, mappings, groups, template_name):
    """使用上传的 .xlsx 模板 + 列映射填充数据，保留原文件全部格式"""
    import io as _io
    import openpyxl
    from copy import copy as _copy
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter, column_index_from_string

    abs_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), '..',
                             template.template_file.lstrip('/'))
    wb = openpyxl.load_workbook(abs_path)
    ws = wb.active

    # 列映射字典: {列字母: field_key}
    col_map = {m.column_letter: m.field_key for m in mappings if m.field_key}
    col_indices = {letter: column_index_from_string(letter) for letter in col_map}

    # 从模板数据行复制样式
    data_start = template.data_start_row or 3
    header_row = template.header_row or 2
    # 样式参考行：数据行在第一行数据处（如有分组标题则在 data_start + 1）
    style_ref_row = data_start + 1 if template.has_group_title else data_start

    # 缓存模板数据行的样式（列字母 → 样式），显式复制避免 openpyxl copy 不完整
    thin_border = Border(
        left=Side(style='thin', color='333333'),
        right=Side(style='thin', color='333333'),
        top=Side(style='thin', color='333333'),
        bottom=Side(style='thin', color='333333'),
    )
    style_cache = {}
    for col_letter, col_idx in col_indices.items():
        ref_cell = ws.cell(row=style_ref_row, column=col_idx)
        rf = ref_cell.font
        ra = ref_cell.alignment
        style_cache[col_letter] = {
            'font': Font(
                name=rf.name or '宋体',
                size=rf.size or 18,
                bold=rf.bold or False,
                color=rf.color if rf.color else '000000',
            ),
            'alignment': Alignment(
                horizontal=ra.horizontal or 'left',
                vertical=ra.vertical or 'center',
                wrap_text=ra.wrap_text if ra.wrap_text is not None else True,
            ),
            'border': _copy(ref_cell.border) if ref_cell.border else thin_border,
            'fill': _copy(ref_cell.fill) if ref_cell.fill else PatternFill(),
        }

    # 缓存分组标题行样式
    if template.has_group_title:
        # 假设分组标题在 data_start 行或附近
        section_style = {
            'font': Font(name='宋体', size=28, bold=True, color='000000'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': thin_border,
        }

    # 清除数据行（从 data_start 起）
    merges_to_remove = []
    for merge_range in ws.merged_cells.ranges:
        if merge_range.min_row >= data_start:
            merges_to_remove.append(str(merge_range))
    for mr_str in merges_to_remove:
        ws.unmerge_cells(mr_str)

    if ws.max_row >= data_start:
        ws.delete_rows(data_start, ws.max_row - data_start + 1)

    # 更新标题
    if template.title_row and template.title_row > 0:
        ws.cell(row=template.title_row, column=1).value = template_name

    # 填充数据
    current_row = data_start
    total_cols = max(col_indices.values()) if col_indices else 10
    # 缓存列宽用于自动行高计算
    col_widths = {}
    for col_letter in col_indices:
        cd = ws.column_dimensions.get(col_letter)
        col_widths[col_letter] = cd.width if cd and cd.width else 10

    def _calc_row_height(r):
        """根据单元格内容和列宽自动计算行高，使内容完全展开"""
        max_lines = 1
        for col_letter, col_idx in col_indices.items():
            cell = ws.cell(row=r, column=col_idx)
            text = str(cell.value) if cell.value else ''
            if not text:
                continue
            lines = text.split('\n')
            line_count = len(lines)
            # 估算每行最大字符数：列宽 * 1.5（中文字符较宽）
            cw = col_widths.get(col_letter, 10)
            chars_per_line = max(3, int(cw * 1.5))
            for line in lines:
                if len(line) > chars_per_line:
                    line_count += (len(line) - 1) // chars_per_line
            max_lines = max(max_lines, line_count)
        return max(28, max_lines * 24)

    for group in groups:
        group_title = group.get('title')
        rows = group.get('rows', [])
        if not rows:
            continue

        # 分组标题行
        if group_title and template.has_group_title:
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=total_cols - 1)
            title_cell = ws.cell(row=current_row, column=1)
            title_cell.value = group_title
            title_cell.font = section_style['font']
            title_cell.alignment = section_style['alignment']
            for c in range(1, total_cols + 1):
                ws.cell(row=current_row, column=c).border = section_style['border']
            ws.row_dimensions[current_row].height = 34
            current_row += 1

        # 数据行
        for row_data in rows:
            for col_letter, field_key in col_map.items():
                col_idx = col_indices[col_letter]
                val = row_data.get(field_key, '')
                if val is None:
                    val = ''
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = str(val) if not isinstance(val, (int, float)) else val

                style = style_cache.get(col_letter)
                if style:
                    cell.font = style['font']
                    cell.alignment = style['alignment']
                    cell.border = style['border']
                    cell.fill = style['fill']
            ws.row_dimensions[current_row].height = _calc_row_height(current_row)
            current_row += 1

    # 恢复冻结窗格
    ws.freeze_panes = f'A{header_row + 1}'

    output = _io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@admin_print_bp.route('/print/download', methods=['GET'])
@dual_login_required
def print_download():
    """使用打印模板生成并下载 Excel 文件"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    ids_str = request.args.get('project_ids', '')
    project_ids = [int(x) for x in ids_str.split(',') if x.strip().isdigit()] if ids_str else []

    template_id = request.args.get('template_id', 0, type=int) or _get_default_template_id()
    activity_range = request.args.get('activity_range', '').strip()
    demand_mode = request.args.get('demand_mode', 'aggregate').strip()
    # 诉求状态过滤：默认仅待回应+协调中（不含已回应），传空字符串表示全部
    demand_status = request.args.get('demand_status', 'pending,processing').strip()

    template = PrintTemplate.query.get(template_id)
    template_name = template.name if template else '招商项目库'

    fields = PrintFieldConfig.query.filter_by(template_id=template_id, is_visible=True)\
        .order_by(PrintFieldConfig.sort_order).all()
    if not fields:
        return jsonify({'code': 1, 'message': '未配置打印字段'}), 400

    q = InvestmentProject.query.filter_by(is_deleted=False)
    if project_ids:
        q = q.filter(InvestmentProject.id.in_(project_ids))
    projects = q.order_by(InvestmentProject.order_no.asc()).all()

    # 字典映射
    from models import FollowStatusDict, MeetingStatusDict, OrganizationDict, DemandTypeDict, EnterpriseDemand
    follow_map = {d.code: d.name for d in FollowStatusDict.query.all()}
    meeting_map = {d.code: d.name for d in MeetingStatusDict.query.all()}
    org_map = {d.code: d.name for d in OrganizationDict.query.all()}
    demand_type_map = DemandTypeDict.build_display_name_map()

    # 样式定义（使用打印模板的字体设置）
    _tf = template.title_font_family or template.font_family or '微软雅黑'
    _ts = template.title_font_size or template.header_font_size or 14
    _hf = template.table_header_font_family or template.font_family or '微软雅黑'
    _hs = template.table_header_font_size or template.font_size or 10
    _cf = template.cell_font_family or template.font_family or '微软雅黑'
    _cs = template.cell_font_size or template.font_size or 10

    title_font = Font(name=_tf, size=_ts, bold=True, color='1A3A5C')
    title_align = Alignment(horizontal='center', vertical='center')
    header_font = Font(name=_hf, size=_hs, bold=True, color='000000')
    header_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name=_cf, size=_cs)
    cell_align = Alignment(vertical='top', wrap_text=True)
    cell_align_center = Alignment(horizontal='center', vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='333333'),
        right=Side(style='thin', color='333333'),
        top=Side(style='thin', color='333333'),
        bottom=Side(style='thin', color='333333'),
    )
    title_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # ---- 模板文件填充路径（上传的 .xlsx 模板 + 列映射） ----
    from models import TemplateFieldMapping
    template_mappings = TemplateFieldMapping.query.filter_by(template_id=template_id)\
        .filter(TemplateFieldMapping.field_key != '').all() if template else []
    use_template_file = template and template.template_file and len(template_mappings) > 0

    if use_template_file:
        all_rows = [_resolve_project_row(p, activity_range, demand_status) for p in projects]
        groups = _group_rows_by_status(all_rows)
        output = _fill_template_file(template, template_mappings, groups, template_name)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{_sanitize_filename(template_name)}.xlsx'
        )

    wb = Workbook()
    ws = wb.active
    ws.title = template_name[:31]

    if demand_mode == 'row':
        project_fields = [f for f in fields if f.field_key not in ('demands', 'resolution')]
        demand_cols = [
            {'field_key': 'demand_type', 'field_label': '诉求类型', 'column_width': 120},
            {'field_key': 'demand_unit', 'field_label': '对接部门', 'column_width': 140},
            {'field_key': 'demand_content', 'field_label': '诉求内容', 'column_width': 400},
            {'field_key': 'demand_resolution', 'field_label': '解决措施', 'column_width': 400},
        ]
        total_cols = len(project_fields) + len(demand_cols)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(row=1, column=1, value=template_name)
        title_cell.font = title_font
        title_cell.alignment = title_align
        title_cell.fill = title_fill
        for c in range(1, total_cols + 1):
            ws.cell(row=1, column=c).border = thin_border

        for col_idx, f in enumerate(project_fields, 1):
            cell = ws.cell(row=2, column=col_idx, value=f.field_label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = f.column_width / 7
        for col_idx, dc in enumerate(demand_cols, len(project_fields) + 1):
            cell = ws.cell(row=2, column=col_idx, value=dc['field_label'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = dc['column_width'] / 7

        current_row = 3
        for p in projects:
            project_row = _resolve_project_row(p, activity_range, demand_status)
            q_demands = EnterpriseDemand.query.filter_by(project_id=p.id)
            if demand_status:
                status_list = [s.strip() for s in demand_status.split(',') if s.strip()]
                if status_list:
                    q_demands = q_demands.filter(EnterpriseDemand.status.in_(status_list))
            demands = q_demands.order_by(EnterpriseDemand.sort_order.asc()).all()
            if not demands:
                for col_idx, f in enumerate(project_fields, 1):
                    val = project_row.get(f.field_key, '')
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.font = cell_font; cell.alignment = cell_align; cell.border = thin_border
                for col_idx in range(len(project_fields) + 1, total_cols + 1):
                    cell = ws.cell(row=current_row, column=col_idx, value='')
                    cell.font = cell_font; cell.alignment = cell_align_center; cell.border = thin_border
                current_row += 1
            else:
                start_row = current_row
                for d in demands:
                    for col_idx, f in enumerate(project_fields, 1):
                        val = project_row.get(f.field_key, '')
                        cell = ws.cell(row=current_row, column=col_idx, value=val)
                        cell.font = cell_font; cell.alignment = cell_align; cell.border = thin_border
                    dt_codes = [c.strip() for c in (d.demand_type_code or '').split(',') if c.strip()]
                    demand_data = {
                        'demand_type': '、'.join([demand_type_map.get(c, c) for c in dt_codes]) if dt_codes else '',
                        'demand_unit': org_map.get(d.unit_code, d.unit_code or ''),
                        'demand_content': d.demand_content or '',
                        'demand_resolution': d.resolution or '',
                    }
                    for col_idx, dc in enumerate(demand_cols, len(project_fields) + 1):
                        val = demand_data.get(dc['field_key'], '')
                        cell = ws.cell(row=current_row, column=col_idx, value=val)
                        cell.font = cell_font
                        cell.alignment = cell_align_center if col_idx <= len(project_fields) + 2 else cell_align
                        cell.border = thin_border
                    current_row += 1
                if current_row - start_row > 1:
                    for col_idx in range(1, len(project_fields) + 1):
                        ws.merge_cells(start_row=start_row, start_column=col_idx,
                                       end_row=current_row - 1, end_column=col_idx)
        ws.freeze_panes = 'A3'
    else:
        # V3: 分组模式写入
        total_cols = len(fields)
        # 标题行
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(row=1, column=1, value=template_name)
        title_cell.font = title_font
        title_cell.alignment = title_align
        title_cell.fill = title_fill
        for c in range(1, total_cols + 1):
            ws.cell(row=1, column=c).border = thin_border

        # 列标题行
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=2, column=col_idx, value=field.field_label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = field.column_width / 7

        # 解析所有行并分组
        all_rows = [_resolve_project_row(p, activity_range, demand_status) for p in projects]
        groups = _group_rows_by_status(all_rows)

        # V3 子标题样式
        _ss = template.sub_title_font_size or 12
        _sb = template.subtitle_bold if template.subtitle_bold is not None else True
        subtitle_font = Font(name=_tf, size=_ss, bold=_sb, color='000000')
        subtitle_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        subtitle_align = Alignment(horizontal='center', vertical='center')

        current_row = 3
        for group in groups:
            if group.get('title'):
                # 子标题合并行
                ws.merge_cells(start_row=current_row, start_column=1,
                               end_row=current_row, end_column=total_cols)
                sub_cell = ws.cell(row=current_row, column=1, value=group['title'])
                sub_cell.font = subtitle_font
                sub_cell.fill = subtitle_fill
                sub_cell.alignment = subtitle_align
                for c in range(1, total_cols + 1):
                    ws.cell(row=current_row, column=c).border = thin_border
                current_row += 1

            for row_data in group.get('rows', []):
                for col_idx, field in enumerate(fields, 1):
                    val = row_data.get(field.field_key, '')
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.font = cell_font
                    cell.alignment = cell_align
                    cell.border = thin_border
                current_row += 1

        ws.freeze_panes = 'A3'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{_sanitize_filename(template_name)}.xlsx'
    )
